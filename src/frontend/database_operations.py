# database_operations.py
import csv
import logging
from pathlib import Path
from typing import List, Dict, Optional
from mysql.connector.cursor import MySQLCursor
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from frontend.database_connection import DatabaseConnection
from frontend.database_config import DatabaseConfig

class DatabaseOperations:
    """Handles all database operations including MySQL, CSV, and Excel interactions."""
    
    def __init__(self, db_config: DatabaseConfig):
        self.db_config = db_config
        self.logger = logging.getLogger(__name__)

    # Project Info Operations
    def save_project_info(self, cursor, data: Dict, common_id: str) -> int:
        """Save project information to database."""
        query = """
            INSERT INTO project_info (
                common_id, project_title, section, unit_force, unit_length, 
                unit_time, model_type, element_type, borehole_type, 
                borehole, design_approach
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        values = (
            common_id,
            data.get("Project Title"),
            data.get("Section"),
            data.get("Unit Force"),
            data.get("Unit Length"),
            data.get("Unit Time"),
            data.get("Model Type"),
            data.get("Element Type"),
            data.get("Borehole Type"),
            data.get("Borehole"),
            data.get("Design Approach")
        )
        cursor.execute(query, values)
        return cursor.lastrowid

    # Borehole Operations
    def save_borehole_data(self, cursor, data: List[Dict], common_id: str) -> None:
        """Save borehole data to database."""
        for borehole_set in data:
            query = """
                INSERT INTO borehole_data (
                    common_id, SoilType, DrainType, SPT, TopDepth, BottomDepth, 
                    gammaUnsat, gammaSat, Eref, nu, cref, phi, kx, ky, Rinter, K0Primary
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            values = (
                common_id,
                borehole_set.get("Soil Type"),
                borehole_set.get("Drain Type"),
                borehole_set.get("SPT"),
                borehole_set.get("Top Depth"),
                borehole_set.get("Bottom Depth"),
                borehole_set.get("Gamma Unsat"),
                borehole_set.get("Gamma Sat"),
                borehole_set.get("E ref"),
                borehole_set.get("Nu"),
                borehole_set.get("C '"),
                borehole_set.get("Phi '"),
                borehole_set.get("Kx"),
                borehole_set.get("Ky"),
                borehole_set.get("R inter"),
                borehole_set.get("K0 Primary")
            )
            cursor.execute(query, values)
   
    def get_soil_material_names(self) -> List[str]:
        """Get distinct soil material names from database."""
        with DatabaseConnection(self.db_config) as db:
            db.cursor.execute("SELECT DISTINCT MaterialName FROM soil_properties")
            results = db.cursor.fetchall()
            return [result['MaterialName'] for result in results]

    # CSV Operations
    def import_project_info_from_csv(self, csv_file_path: str) -> Dict:
        """Import project info from CSV file."""
        try:
            with open(csv_file_path) as file:
                reader = csv.DictReader(file)
                return next(reader)
        except Exception as e:
            self.logger.error(f"Error reading project CSV: {str(e)}")
            raise

    def import_borehole_from_csv(self, csv_file_path: str) -> List[Dict]:
        """Import borehole data from CSV file."""
        try:
            df = pd.read_csv(csv_file_path)
            # Add validation and transformation logic here
            return df.to_dict('records')
        except Exception as e:
            self.logger.error(f"Error reading borehole CSV: {str(e)}")
            raise

    # Common Operations
    def save_to_csv(self, data: List[Dict], filename: str, headers: List[str]) -> None:
        """Save data to CSV file."""
        with open(filename, 'a', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=headers)
            if file.tell() == 0:
                writer.writeheader()
            for row in data:
                writer.writerow(row)
    def update_excel(self, filename: str, sheet_name: str, data: List[Dict]) -> None:
      """Update Excel file with new data, creating 'Plate Properties' and 'Soil Properties' if updating 'Project Info'."""
      try:
        file_path = Path(filename)
        
        # Load or create workbook
        if file_path.exists():
            workbook = openpyxl.load_workbook(filename)
        else:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)

        # Load soil properties color mapping if working with Borehole sheet
        soil_color_mapping = {}
        if sheet_name == "Borehole":
            try:
                if 'Soil Properties' in workbook.sheetnames:
                    soil_props_sheet = workbook['Soil Properties']
                    # Find MaterialName and Colour column indices
                    headers = [cell.value for cell in soil_props_sheet[1]]
                    material_col_idx = None
                    colour_col_idx = None
                    
                    for idx, header in enumerate(headers):
                        if header == 'MaterialName':
                            material_col_idx = idx
                        elif header == 'Colour':
                            colour_col_idx = idx
                    
                    # Build color mapping dictionary
                    if material_col_idx is not None and colour_col_idx is not None:
                        for row in soil_props_sheet.iter_rows(min_row=2, values_only=True):
                            if row[material_col_idx]:
                                material_name = str(row[material_col_idx]).strip()
                                colour_value = row[colour_col_idx] if colour_col_idx < len(row) else 15236578
                                soil_color_mapping[material_name] = colour_value
                        
                        print(f"DEBUG: Loaded color mapping for {len(soil_color_mapping)} soil types")
                else:
                    print("DEBUG: Soil Properties sheet not found, using default colors")
            except Exception as e:
                print(f"WARNING: Could not load soil color mapping: {e}")

        # Handle vertical format sheets (Project Info, Geometry Info)
        if sheet_name in ["Project Info", "Geometry Info"]:
            # [Keep existing Project Info and Geometry Info code unchanged]
            if sheet_name in workbook.sheetnames:
                idx = workbook.sheetnames.index(sheet_name)
                workbook.remove(workbook[sheet_name])
                sheet = workbook.create_sheet(sheet_name, idx)
            else:
                sheet = workbook.create_sheet(sheet_name)

            sheet.cell(row=1, column=1, value="Parameters")
            sheet.cell(row=1, column=2, value="Value")

            parameter_order = {
                "Project Info": [
                    "ProjectTitle", "Section", "UnitForce", "UnitLength", "UnitTime", "ModelType", 
                    "ElementType", "BoreholeType", "Borehole", "DesignApproach"
                ],
                "Geometry Info": [
                    "NoOfStrut", "ExcavationBelowStrut", "OverExcavation", "WallType", "Material",
                    "MemberSize", "Spacing", "Borehole_x_coordinate", "GroundWatertable",
                    "x_min_coordinate", "y_min_coordinate", "x_max_coordinate", "y_max_coordinate"
                ]
            }

            parameter_mapping = {
                "Project Info": {
                    "Project Title": "ProjectTitle", "Section": "Section", "Unit Force": "UnitForce",
                    "Unit Length": "UnitLength", "Unit Time": "UnitTime", "Model Type": "ModelType",
                    "Element Type": "ElementType", "Borehole Type": "BoreholeType",
                    "Borehole": "Borehole", "Design Approach": "DesignApproach"
                },
                "Geometry Info": {
                    "No of Strut": "NoOfStrut", "Excavation Below Strut": "ExcavationBelowStrut",
                    "Over Excavation": "OverExcavation", "Wall Type": "WallType", "Material": "Material",
                    "Member Size": "MemberSize", "Spacing": "Spacing",
                    "Borehole X Coordinate": "Borehole_x_coordinate",
                    "Ground Water Table": "GroundWatertable", "x_min": "x_min_coordinate",
                    "y_min": "y_min_coordinate", "x_max": "x_max_coordinate", "y_max": "y_max_coordinate"
                }
            }

            if data:
                row_num = 2
                sheet_data = data[0]
                if 'common_id' in sheet_data:
                    sheet_data = {k: v for k, v in sheet_data.items() if k != 'common_id'}

                for param in parameter_order[sheet_name]:
                    original_key = next((k for k, v in parameter_mapping[sheet_name].items() if v == param), None)
                    if original_key and original_key in sheet_data:
                        sheet.cell(row=row_num, column=1, value=param)
                        sheet.cell(row=row_num, column=2, value=sheet_data[original_key])
                        row_num += 1

                for col in range(1, 3):
                    column_letter = get_column_letter(col)
                    max_length = 0
                    for cell in sheet[column_letter]:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    sheet.column_dimensions[column_letter].width = max_length + 2

        # Special handling for Borehole sheet with color matching
        elif sheet_name == "Borehole":
            if sheet_name in workbook.sheetnames:
                idx = workbook.sheetnames.index(sheet_name)
                workbook.remove(workbook[sheet_name])
                sheet = workbook.create_sheet(sheet_name, idx)
            else:
                sheet = workbook.create_sheet(sheet_name)
            
            columns = [
                "SoilType", "SoilModel", "DrainageType", "SPT", 
                "Top", "Bottom", "gammaUnsat", "gammaSat", "Eref", "nu", 
                "cref", "phi", "kx", "ky", "Strength", "Rinter", 
                "K0Determination", "K0Primary", "Colour"
            ]
            
            key_mapping = {
                "SoilType": ["MaterialName", "SoilType", "Soil Type", "soil_type", "soiltype"],
                "SoilModel": ["SoilModel", "Soil Model", "soil_model"],
                "DrainageType": ["DrainageType", "DrainType", "Drain Type", "drainage_type", "drain_type"],
                "SPT": ["SPT", "spt", "Spt"],
                "Top": ["Top", "top", "TopDepth", "top_depth", "Top Depth"],
                "Bottom": ["Bottom", "bottom", "BottomDepth", "bottom_depth", "Bottom Depth"],
                "gammaUnsat": ["gammaUnsat", "Gamma Unsat", "gamma_unsat", "GammaUnsat"],
                "gammaSat": ["gammaSat", "Gamma Sat", "gamma_sat", "GammaSat"],
                "Eref": ["Eref", "E ref", "e_ref", "ERef"],
                "nu": ["nu", "Nu", "NU", "Poisson"],
                "cref": ["cref", "C '", "c_ref", "CRef", "Cohesion"],
                "phi": ["phi", "Phi '", "phi_ref", "Phi", "FrictionAngle"],
                "kx": ["kx", "Kx", "KX"],
                "ky": ["ky", "Ky", "KY"],
                "Strength": ["Strength", "strength"],
                "Rinter": ["Rinter", "R inter", "r_inter", "RInter"],
                "K0Determination": ["K0Determination", "K0 Determination", "k0_determination"],
                "K0Primary": ["K0Primary", "K0 Primary", "k0_primary", "K0"],
                "Colour": ["Colour", "Color", "colour", "color"]
            }
            
            numeric_columns = ["SPT", "gammaUnsat", "gammaSat", "Eref", "nu", "cref", 
                              "phi", "kx", "ky", "Rinter", "K0Primary", "Top", "Bottom"]
            
            default_values = {
                "SoilModel": "Mohrcoulomb",
                "Strength": "Manual",
                "K0Determination": "Manual",
                "Colour": 15236578
            }
            
            # Add headers
            sheet.append(columns)
            
            # Helper function to find the actual key in data
            def find_actual_key(row_data, target_column):
                possible_keys = key_mapping.get(target_column, [target_column])
                for key in possible_keys:
                    if key in row_data:
                        return key
                return None
            
            # Extract and add data rows with color matching
            for row_idx, row_data in enumerate(data, 2):
                row_data_filtered = {k: v for k, v in row_data.items() if k != 'common_id'}
                
                # Get soil type for color matching
                soil_type_key = find_actual_key(row_data_filtered, "SoilType")
                soil_type_value = row_data_filtered.get(soil_type_key, "") if soil_type_key else ""
                
                # Match color from Soil Properties based on soil type
                matched_color = None
                if soil_type_value and soil_color_mapping:
                    soil_type_clean = str(soil_type_value).strip()
                    matched_color = soil_color_mapping.get(soil_type_clean)
                    if matched_color:
                        print(f"DEBUG: Matched color {matched_color} for soil type '{soil_type_clean}'")
                
                for col_idx, column in enumerate(columns, 1):
                    actual_key = find_actual_key(row_data_filtered, column)
                    
                    # Special handling for Colour column
                    if column == "Colour":
                        if matched_color is not None:
                            value = matched_color
                        elif actual_key and row_data_filtered.get(actual_key):
                            value = row_data_filtered.get(actual_key)
                        else:
                            value = default_values.get(column, 15236578)
                    else:
                        if actual_key:
                            value = row_data_filtered.get(actual_key, "")
                        elif column in default_values:
                            value = default_values[column]
                        else:
                            value = ""
                    
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Format numeric columns
                    if column in numeric_columns and value != "":
                        try:
                            numeric_value = float(value)
                            cell.value = numeric_value
                            cell.number_format = 'General'
                            cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                        except (ValueError, TypeError):
                            pass
            
            # Auto-adjust column widths
            for col_idx, column in enumerate(columns, 1):
                column_letter = get_column_letter(col_idx)
                max_length = len(column)
                for cell in sheet[column_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[column_letter].width = max_length + 2

        # Handle Excavation Details sheet (keep existing code)
        elif sheet_name == "Excavation Details":
            # [Keep existing Excavation Details code unchanged]
            if sheet_name in workbook.sheetnames:
                idx = workbook.sheetnames.index(sheet_name)
                workbook.remove(workbook[sheet_name])
                sheet = workbook.create_sheet(sheet_name, idx)
            else:
                sheet = workbook.create_sheet(sheet_name)
            
            if data:
                columns = list(data[0].keys())
                if 'common_id' in columns:
                    columns.remove('common_id')
            else:
                columns = []
            
            sheet.append(columns)
            
            for row_idx, row_data in enumerate(data, 2):
                row_data_filtered = {k: v for k, v in row_data.items() if k != 'common_id'}
                
                for col_idx, column in enumerate(columns, 1):
                    value = row_data_filtered.get(column, "")
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    
                    try:
                        if value != "" and (isinstance(value, (int, float)) or 
                                           (isinstance(value, str) and value.replace('.', '', 1).replace('-', '', 1).isdigit())):
                            numeric_value = float(value)
                            cell.value = numeric_value
                            cell.number_format = 'General'
                            cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                    except (ValueError, TypeError):
                        pass
            
            for col_idx, column in enumerate(columns, 1):
                column_letter = get_column_letter(col_idx)
                max_length = len(column)
                for cell in sheet[column_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[column_letter].width = max_length + 2

        # Handle normal horizontal format for all other sheets
        else:
            if sheet_name in workbook.sheetnames:
                idx = workbook.sheetnames.index(sheet_name)
                workbook.remove(workbook[sheet_name])
                sheet = workbook.create_sheet(sheet_name, idx)
            else:
                sheet = workbook.create_sheet(sheet_name)
                
            headers = list(data[0].keys()) if data else []
            if 'common_id' in headers:
                headers.remove('common_id')
            sheet.append(headers)

            for row_idx, row in enumerate(data, 2):
                row_without_common_id = {k: v for k, v in row.items() if k != 'common_id'}
                
                for col_idx, (header, value) in enumerate(row_without_common_id.items(), 1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    
                    try:
                        if value != "" and (isinstance(value, (int, float)) or 
                                         (isinstance(value, str) and value.replace('.', '', 1).replace('-', '', 1).isdigit())):
                            numeric_value = float(value)
                            cell.value = numeric_value
                            cell.number_format = 'General'
                            cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                    except (ValueError, TypeError):
                        pass

        workbook.save(filename)
      except Exception as e:
        self.logger.error(f"Error updating Excel: {str(e)}")
        raise
    def ensure_input_data_sheets(self, filename: str) -> None:
      """Ensure Input_Data.xlsx exists with required sheets and create Soil_DB.xlsx if needed."""
      try:
        # Define soil data once to reuse
        soil_headers = [
            "MaterialName", "SoilModel", "DrainageType", "SPT", "gammaUnsat", "gammaSat", 
            "Eref", "nu", "cref", "phi", "kx", "ky", "Strength", "Rinter", 
            "K0Determination", "K0Primary", "Colour"
        ]

        soil_data = [
            ["Fill", "Mohrcoulomb", "Drain", 10, 17, 18, 10000, 0.3, 1, 30, 1.00E-06, 1.00E-06, "Manual", 0.7, "Manual", 0.5, 9881855],
            ["F1", "Mohrcoulomb", "Drain", 10, 17, 18, 6000, 0.3, 1, 30, 1.00E-06, 1.00E-06, "Manual", 0.7, "Manual", 0.5, 65535],
            ["F2", "Mohrcoulomb", "Undrain", 10, 17, 18, 5000, 0.3, 10, 25, 1.00E-06, 1.00E-06, "Manual", 0.7, "Manual", 0.5, 16428800],
            ["OA(E)", "Mohrcoulomb", "Drain", 10, 18.5, 19.5, 33333.33333, 0.3, 2, 30, 1.00E-07, 1.00E-07, "Manual", 0.7, "Manual", 0.7, 11488255],
            ["OA(D)", "Mohrcoulomb", "Drain", 10, 18, 19, 25000, 0.3, 4, 32, 1.00E-06, 1.00E-06, "Manual", 0.7, "Manual", 0.7, 9843450],
            ["OA(C)", "Mohrcoulomb", "Drain", 10, 19, 20, 33333.33333, 0.3, 10, 33, 1.00E-07, 1.00E-07, "Manual", 0.7, "Manual", 0.7, 9843400],
            ["OA(B)", "Mohrcoulomb", "Drain", 10, 19, 20, 75000, 0.3, 8, 32, 1.00E-06, 1.00E-06, "Manual", 0.7, "Manual", 0.7, 9843360],
            ["OA(A)", "Mohrcoulomb", "Drain", 10, 19, 20, 150000, 0.3, 10, 32, 1.00E-06, 1.00E-06, "Manual", 0.7, "Manual", 0.8, 8204925],
            ["G(II)", "Mohrcoulomb", "Drain", 10, 25, 26, 5000000, 0.3, 300, 60, 1.00E-07, 1.00E-07, "Manual", 0.7, "Manual", 0.8, 12850],
            ["G(III)", "Mohrcoulomb", "Drain", 10, 25, 26, 2000000, 0.3, 100, 60, 1.00E-07, 1.00E-07, "Manual", 0.7, "Manual", 0.8, 1657675],
            ["G(IV)", "Mohrcoulomb", "Drain", 10, 25, 26, 500000, 0.3, 30, 45, 1.00E-07, 1.00E-07, "Manual", 0.7, "Manual", 0.8, 3302500],
            ["G(V)", "Mohrcoulomb", "Drain", 10, 18, 19, 25000, 0.3, 5, 32, 1.00E-07, 1.00E-07, "Manual", 0.7, "Manual", 0.8, 4953750],
            ["G(VI)", "Mohrcoulomb", "Drain", 10, 18, 19, 25000, 0.3, 5, 32, 1.00E-07, 1.00E-07, "Manual", 0.7, "Manual", 0.8, 6598575],
            ["S(II)", "Mohrcoulomb", "Drain", 10, 25, 26, 695652.1739, 0.3, 50, 30, 5.00E-08, 5.00E-08, "Manual", 0.7, "Manual", 0.8, 19200],
            ["S(III)", "Mohrcoulomb", "Drain", 10, 24.5, 25.5, 800000, 0.3, 60, 40, 1.00E-07, 1.00E-07, "Manual", 0.7, "Manual", 0.8, 1664025],
            ["S(IV)", "Mohrcoulomb", "Drain", 10, 22, 23, 500000, 0.3, 40, 35, 1.00E-07, 1.00E-07, "Manual", 0.7, "Manual", 0.8, 1676825],
            ["S(V)", "Mohrcoulomb", "Drain", 10, 20.5, 21.5, 35000, 0.3, 20, 30, 1.00E-07, 1.00E-07, "Manual", 0.7, "Manual", 0.8, 1689625],
            ["S(VI)", "Mohrcoulomb", "Drain", 10, 19.5, 20.5, 35000, 0.3, 5, 30, 1.00E-07, 1.00E-07, "Manual", 0.7, "Manual", 0.8, 64000]
        ]

        # Handle Input_Data.xlsx
        if not Path(filename).exists():
            workbook = openpyxl.Workbook()
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            
            # Create Soil Properties sheet
            soil_sheet = workbook.create_sheet("Soil Properties")
            soil_sheet.append(soil_headers)
            for row in soil_data:
                soil_sheet.append(row)
            
            workbook.save(filename)
        else:
            workbook = openpyxl.load_workbook(filename)
            sheet_names = workbook.sheetnames
            
            # Check and create Soil Properties if missing
            if "Soil Properties" not in sheet_names:
                soil_sheet = workbook.create_sheet("Soil Properties")
                soil_sheet.append(soil_headers)
                for row in soil_data:
                    soil_sheet.append(row)
                workbook.save(filename)

        # Handle Soil_DB.xlsx
        soil_db_filename = Path(filename).parent / "Soil_DB.xlsx"
        
        if not soil_db_filename.exists():
            print(f"Creating new Soil_DB.xlsx at {soil_db_filename}")
            # Create new Soil_DB.xlsx file
            soil_workbook = openpyxl.Workbook()
            # Remove default sheet
            if 'Sheet' in soil_workbook.sheetnames:
                del soil_workbook['Sheet']
            
            # Create Soil Properties sheet
            soil_db_sheet = soil_workbook.create_sheet("Soil Properties")
            soil_db_sheet.append(soil_headers)
            for row in soil_data:
                soil_db_sheet.append(row)
            
            soil_workbook.save(soil_db_filename)
        else:
            print(f"Ensuring Soil_DB.xlsx at {soil_db_filename}")
            # Check if Soil Properties sheet exists in existing Soil_DB.xlsx
            soil_workbook = openpyxl.load_workbook(soil_db_filename)
            if "Soil Properties" not in soil_workbook.sheetnames:
                soil_db_sheet = soil_workbook.create_sheet("Soil Properties")
                soil_db_sheet.append(soil_headers)
                for row in soil_data:
                    soil_db_sheet.append(row)
                soil_workbook.save(soil_db_filename)

      except Exception as e:
        self.logger.error(f"Error ensuring input data sheets: {str(e)}")
        raise
    def save_geometry_data(self, cursor, common_id: str, data: Dict) -> None:
      """
      Save geometry data to database with proper type conversion and validation.
    
      Args:
        cursor: Database cursor
        common_id: Unique identifier for the project
        data: Dictionary containing geometry data
      """
      try:
        # Type conversion function with error handling
        def convert_value(value, expected_type):
            if value is None:
                return None
            try:
                if expected_type == float:
            # Check if the value is already a float
                    if isinstance(value, float):
                        return value
            # Convert to string first only if not already a float, then strip
                    return float(str(value).strip() if isinstance(value, str) else value)
                elif expected_type == int:
            # Check if the value is already an int
                    if isinstance(value, int):
                        return value
            # Convert to string first only if not already an int, then strip
                    return int(str(value).strip() if isinstance(value, str) else value)
                elif expected_type == str:
            # Check if the value is already a string
                    if isinstance(value, str):
                        return value.strip()
            # Convert to string if not already a string
                    return str(value)
                return value
            except (ValueError, TypeError) as e:
                self.logger.warning(f"Error converting value '{value}' to {expected_type.__name__}: {str(e)}")
                return None

        # Define expected types for each field
        field_types = {
            "Excavation Type": str,
            "Wall Top Level": float,
            "Excavation Depth": float,
            "Excavation Width": float,
            "Toe Level": float,
            "No of Strut": int,
            "Strut Type": str,
            "Excavation Below Strut": float,
            "Over Excavation": float,
            "Wall Type": str,
            "Material": str,
            "Member Size": str,
            "Spacing": float,
            "Borehole X Coordinate": float,
            "Ground Water Table": float,
            "x_min": float,
            "y_min": float,
            "x_max": float,
            "y_max": float
        }

        # Convert and validate all values
        converted_data = {
            key: convert_value(data.get(key), field_types.get(key, str))
            for key in field_types.keys()
        }

        # Insert into geometry table
        query = """
            INSERT INTO geometry (
                common_id, excavation_type, wall_top_level, excavation_depth,
                excavation_width, toe_level, no_of_strut, strut_type,
                excavation_below_strut, over_excavation, wall_type, material,
                member_size, spacing, borehole_x_coordinate, ground_water_table,
                x_min_coordinate, y_min_coordinate, x_max_coordinate, y_max_coordinate
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        values = (
            common_id,
            converted_data.get("Excavation Type"),
            converted_data.get("Wall Top Level"),
            converted_data.get("Excavation Depth"),
            converted_data.get("Excavation Width"),
            converted_data.get("Toe Level"),
            converted_data.get("No of Strut"),
            converted_data.get("Strut Type"),
            converted_data.get("Excavation Below Strut"),
            converted_data.get("Over Excavation"),
            converted_data.get("Wall Type"),
            converted_data.get("Material"),
            converted_data.get("Member Size"),
            converted_data.get("Spacing"),
            converted_data.get("Borehole X Coordinate"),
            converted_data.get("Ground Water Table"),
            converted_data.get("x_min"),
            converted_data.get("y_min"),
            converted_data.get("x_max"),
            converted_data.get("y_max")
        )
        cursor.execute(query, values)
        
      except Exception as e:
        self.logger.error(f"Error saving geometry data: {str(e)}")
        raise

    def save_wall_details(self, cursor, common_id: str, wall_data: Dict) -> None:
      """Save wall details to database."""
      try:
        query = """
            INSERT INTO erss_wall_details 
            (common_id, MaterialName, WallName, x_Top, y_Top, x_Bottom, y_Bottom)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """
        cursor.execute(query, [
            common_id,
            wall_data['MaterialName'],
            wall_data['WallName'],
            wall_data['x_Top'],
            wall_data['y_Top'],
            wall_data['x_Bottom'],
            wall_data['y_Bottom']
        ])
      except Exception as e:
        self.logger.error(f"Error saving wall details: {str(e)}")
        raise

    def save_strut_details(self, cursor, common_id: str, strut_data: Dict) -> None:
      """Save strut details to database."""
      try:
        query = """
            INSERT INTO StrutDetails 
            (common_id, MaterialName, StrutName, x_Left, y_Left, x_Right, y_Right, Type)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """
        cursor.execute(query, [
            common_id,
            strut_data['MaterialName'],
            strut_data['StrutName'],
            strut_data['x_Left'],
            strut_data['y_Left'],
            strut_data['x_Right'],
            strut_data['y_Right'],
            strut_data['Type']
        ])
      except Exception as e:
        self.logger.error(f"Error saving strut details: {str(e)}")
        raise

    def save_anchor_properties(self, cursor, common_id: str, anchor_data: Dict) -> None:
      """Save anchor properties to database."""
      try:
        query = """
            INSERT INTO anchor_properties 
            (common_id, MaterialName, Elasticity, EA, Lspacing, Colour)
            VALUES (?, ?, ?, ?, ?, ?)
        """
        cursor.execute(query, [
            common_id,
            anchor_data['MaterialName'],
            anchor_data['Elasticity'],
            anchor_data['EA'],
            anchor_data['Lspacing'],
            anchor_data['Colour']
        ])
      except Exception as e:
        self.logger.error(f"Error saving anchor properties: {str(e)}")
        raise

    def save_lineload_data(self, cursor, common_id: str, lineload_data: Dict) -> None:
      """Save lineload data to database."""
      try:
        query = """
            INSERT INTO lineload 
            (common_id, LoadName, x_start, y_start, x_end, y_end, qx_start, qy_start, Distribution)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        cursor.execute(query, [
            common_id,
            lineload_data['LoadName'],
            lineload_data['x_start'],
            lineload_data['y_start'],
            lineload_data['x_end'],
            lineload_data['y_end'],
            lineload_data['qx_start'],
            lineload_data['qy_start'],
            lineload_data['Distribution']
        ])
      except Exception as e:
        self.logger.error(f"Error saving lineload data: {str(e)}")
        raise

    def save_excavation_data(self, cursor: MySQLCursor, common_id: str, data: List[Dict], over_excavation: float = None) -> None:
        """
        Save excavation data to database with proper error handling and validation.
        
        Args:
            cursor: Database cursor
            common_id: Unique identifier for the project
            data: List of dictionaries containing excavation stage data
            over_excavation: Optional over-excavation depth
        """
        try:
            # Validate input data
            if not isinstance(data, list) or not data:
                raise ValueError("Excavation data must be a non-empty list")
            
            required_fields = ["Stage No", "Stage Name", "From", "To"]
            for stage in data:
                missing_fields = [field for field in required_fields if field not in stage]
                if missing_fields:
                    raise ValueError(f"Missing required fields: {', '.join(missing_fields)}")

            # Save main stages
            for stage in data:
                query = """
                    INSERT INTO ExcavationStages 
                    (common_id, StageNo, StageName, `From`, `To`)
                    VALUES (?, ?, ?, ?, ?)
                """
                try:
                    cursor.execute(query, (
                        common_id,
                        int(stage["Stage No"]),
                        str(stage["Stage Name"]),
                        float(stage["From"]),
                        float(stage["To"])
                    ))
                except ValueError as ve:
                    raise ValueError(f"Invalid data type in stage {stage['Stage No']}: {str(ve)}")

            # Save over-excavation if provided
            if over_excavation is not None:
                if not isinstance(over_excavation, (int, float)) or over_excavation < 0:
                    raise ValueError("Over-excavation must be a non-negative number")
                
                last_to = float(data[-1]["To"])
                query = """
                    INSERT INTO ExcavationStages 
                    (common_id, StageNo, StageName, `From`, `To`)
                    VALUES (?, ?, ?, ?, ?)
                """
                cursor.execute(query, (
                    common_id,
                    len(data) + 1,
                    "Overexcavation",
                    last_to,
                    last_to - over_excavation
                ))

        except Exception as e:
            self.logger.error(f"Error saving excavation data: {str(e)}")
            raise

    def get_plate_properties(self, cursor, material_name: str) -> Dict:
      """Get plate properties from database."""
      try:
        cursor.execute("""
            SELECT Elasticity, EA, Colour
            FROM plate_properties 
            WHERE MaterialName = ?
        """, (material_name,))
        return cursor.fetchone() or {}
      except Exception as e:
        self.logger.error(f"Error getting plate properties: {str(e)}")
        return {}