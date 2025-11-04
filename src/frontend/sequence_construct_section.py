# Standard Library 
import os
import logging
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Union
import flet as ft
# Third-Party Library 
import openpyxl
import csv
import sys
from openpyxl.utils import get_column_letter

# Local Module 
from frontend.database_config import DatabaseConfig
from frontend.database_connection import DatabaseConnection
from frontend.form_section import FormSection, FormField
from frontend.form_manager import FormManager
class SequenceConstructSection(FormSection):
    """Manages construction sequence form sections for excavation projects."""
    def __init__(self, db_config: DatabaseConfig):
        """Initialize with database configuration."""
        self.db_config = db_config
        self.excavation_type = None
        self.row_configurations = []  # Store row-specific configurations
        self.phase_to_config_map = {}  # Map phase names to their configurations
        self.phase_usage_count = {}  # Track how many times each phase has been used
        self.preview_usage_count = {}
        self.phase_element_mapping = {  # Define the sequence of element names for each phase
            'Activate Live Load': ['LL_Left', 'LL_Right'],
            'Activate_ERSS_Wall': ['Wall_Left', 'Wall_Right']
        }
        # NEW: Add mapping for display names vs database names
        self.phase_display_to_db_mapping = {}  # Maps display names to database names
        self.phase_db_to_display_mapping = {}  # Maps database names to display names
        self.excavation_data_cache = []  # Cache excavation data for reference
        self.last_geometry_hash = None  # Track changes in geometry data

    def import_from_list(self, data):
      """Import data from a list of dictionaries and store it in the section"""
      try:
        print(f"DEBUG: import_from_list called with {len(data)} records")
        
        # Reset phase usage tracking
        self.reset_phase_usage_tracking()
        
        # Store the data directly - the section should maintain its own data
        if hasattr(self, 'section_data'):
            self.section_data = data
        
        # Convert list data to column format if needed for internal processing
        column_data = {}
        for item in data:
            for key, value in item.items():
                if key not in column_data:
                    column_data[key] = []
                column_data[key].append(value)
        
        # Store column data for any methods that might need it
        if hasattr(self, 'column_data'):
            self.column_data = column_data
        
        # If this section has a form content or UI that needs updating,
        # we should trigger that here. However, since update_form_content
        # doesn't exist, we'll need to handle this differently.
        
        # Log successful import
        print(f"DEBUG: Successfully imported {len(data)} records to sequence construct section")
        print(f"DEBUG: Sample data: {data[:2] if len(data) > 1 else data}")
        
      except Exception as e:
        print(f"ERROR in import_from_list: {e}")
        import traceback
        traceback.print_exc()
        raise

    def reset_phase_usage_tracking(self):
        print("DEBUG: Resetting phase usage tracking")
        print(f"DEBUG: Previous usage counts: {self.phase_usage_count}")
        self.phase_usage_count = {}
        # Also reset preview counter
        self.preview_usage_count = {}
        print("DEBUG: Phase usage tracking reset complete")

    def get_next_element_name_for_phase(self, phase_name: str, element_type: str) -> str:
      if phase_name not in self.phase_usage_count:
        self.phase_usage_count[phase_name] = 0
      current_count = self.phase_usage_count[phase_name]
      ordinal = self.get_ordinal_number(current_count + 1)
      print(f"DEBUG: Getting element name for phase='{phase_name}' {ordinal} time (current count: {current_count})")
    
      if phase_name in self.phase_element_mapping:
        available_names = self.phase_element_mapping[phase_name]
        if current_count < len(available_names):
            selected_name = available_names[current_count]
            self.phase_usage_count[phase_name] += 1
            new_count = self.phase_usage_count[phase_name]
            print(f"DEBUG: Selected '{selected_name}' for phase '{phase_name}' (usage count updated: {current_count} -> {new_count})")
            
            # Format for display if it's an excavation
            display_name = self.format_element_name_for_display(selected_name, element_type)
            return display_name
        else:
            selected_name = available_names[-1]
            print(f"DEBUG: Max elements reached for phase '{phase_name}', returning last element: '{selected_name}'")
            display_name = self.format_element_name_for_display(selected_name, element_type)
            return display_name
      else:
        matching_configs = self.phase_to_config_map.get(phase_name, [])
        if matching_configs:
            if current_count < len(matching_configs):
                selected_name = matching_configs[current_count]['element_name']
                self.phase_usage_count[phase_name] += 1
                new_count = self.phase_usage_count[phase_name]
                print(f"DEBUG: Selected '{selected_name}' for phase '{phase_name}' from config (usage count updated: {current_count} -> {new_count})")
                
                # Format for display if it's an excavation
                display_name = self.format_element_name_for_display(selected_name, element_type)
                return display_name
            else:
                selected_name = matching_configs[-1]['element_name']
                print(f"DEBUG: Max configs reached for phase '{phase_name}', returning last config element: '{selected_name}'")
                display_name = self.format_element_name_for_display(selected_name, element_type)
                return display_name
        else:
            print(f"DEBUG: No mapping/config found for phase '{phase_name}', using fallback: '{element_type}'")
            return element_type

    def get_ordinal_number(self, num: int) -> str:
        if 10 <= num % 100 <= 20:
            suffix = 'th'
        else:
            suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(num % 10, 'th')
        return f"{num}{suffix}"

    def preview_next_element_name_for_phase(self, phase_name: str, element_type: str) -> str:
      if phase_name not in self.preview_usage_count:
        self.preview_usage_count[phase_name] = 0
      current_count = self.preview_usage_count[phase_name]
      ordinal = self.get_ordinal_number(current_count + 1)
      print(f"DEBUG: User viewing phase name='{phase_name}' for {ordinal} time (preview)")
    
      if phase_name in self.phase_element_mapping:
        available_names = self.phase_element_mapping[phase_name]
        if current_count < len(available_names):
            selected_name = available_names[current_count]
            self.preview_usage_count[phase_name] += 1
            print(f"DEBUG: Will show element name: '{selected_name}' for phase '{phase_name}' (preview count updated to {self.preview_usage_count[phase_name]})")
            
            # Format for display if it's an excavation
            display_name = self.format_element_name_for_display(selected_name, element_type)
            return display_name
        else:
            selected_name = available_names[-1]
            print(f"DEBUG: Max usage reached, showing last element name: '{selected_name}' for phase '{phase_name}'")
            display_name = self.format_element_name_for_display(selected_name, element_type)
            return display_name
      else:
        matching_configs = self.phase_to_config_map.get(phase_name, [])
        if matching_configs:
            if current_count < len(matching_configs):
                selected_name = matching_configs[current_count]['element_name']
                self.preview_usage_count[phase_name] += 1
                print(f"DEBUG: Will show element name: '{selected_name}' for phase '{phase_name}' (from config, preview count updated to {self.preview_usage_count[phase_name]})")
                
                # Format for display if it's an excavation
                display_name = self.format_element_name_for_display(selected_name, element_type)
                return display_name
            else:
                selected_name = matching_configs[-1]['element_name']
                print(f"DEBUG: Max config usage reached, showing last element name: '{selected_name}' for phase '{phase_name}'")
                display_name = self.format_element_name_for_display(selected_name, element_type)
                return display_name
        else:
            print(f"DEBUG: No mapping found for phase '{phase_name}', using element_type as fallback")
            return element_type
    def cache_strut_types(self, geometry_data: Dict):
        """Cache strut types from geometry data for later use"""
        self.strut_type_map = {}
        self.global_strut_type = None
        
        print(f"DEBUG: Available geometry keys: {list(geometry_data.keys())}")
        
        # Try different possible key names for strut types
        possible_keys = ['Strut Type', 'Type of Strut', 'Strut_Type', 'strutType', 'strut_type', 'Strut type']
        strut_types = None
        
        for key in possible_keys:
            if key in geometry_data:
                strut_types = geometry_data.get(key)
                print(f"DEBUG: Found strut types under key '{key}': {strut_types}")
                break
        
        if strut_types is None:
            print("DEBUG: No strut type data found in geometry_data")
            print(f"DEBUG: Full geometry_data structure: {geometry_data}")
            return
        
        # Get number of struts
        num_struts = geometry_data.get('No of Strut', '0')
        try:
            num_struts = int(num_struts)
        except (ValueError, TypeError):
            num_struts = 0
            print(f"DEBUG: Invalid number of struts: {geometry_data.get('No of Strut')}")
        
        # Handle both list and single value cases
        if isinstance(strut_types, list):
            # If it's a list, map each index to corresponding strut
            for i, strut_type in enumerate(strut_types, 1):
                strut_name = f'strut_{i}'
                if strut_type:
                    strut_type_clean = str(strut_type).strip().lower()
                    self.strut_type_map[strut_name] = strut_type_clean
                    print(f"DEBUG: Cached strut type - {strut_name}: '{strut_type_clean}'")
                else:
                    print(f"DEBUG: No type found for {strut_name}, skipping")
        else:
            # Single value - apply to all struts
            strut_type_clean = str(strut_types).strip().lower()
            self.global_strut_type = strut_type_clean
            print(f"DEBUG: Single strut type found: '{strut_type_clean}' - will apply to all {num_struts} struts")
            
            # Cache for all struts
            for i in range(1, num_struts + 1):
                strut_name = f'strut_{i}'
                self.strut_type_map[strut_name] = strut_type_clean
                print(f"DEBUG: Cached strut type - {strut_name}: '{strut_type_clean}'")
        
        print(f"DEBUG: Final strut_type_map: {self.strut_type_map}")
    # Rest of the methods remain the same...
    def calculate_rows_and_options(self, previous_data: Dict = None) -> List[Dict]:
      self.reset_phase_usage_tracking()
      geometry_data = previous_data.get('geometry', {}) if previous_data else {}
      excavation_data = previous_data.get('excavation', []) if previous_data else []
      self.cache_strut_types(geometry_data)  # This will now properly cache all struts
      self.excavation_data_cache = excavation_data
      excavation_type = geometry_data.get('Excavation Type', 'Single wall')
      self.excavation_type = excavation_type
      num_struts = geometry_data.get('No of Strut', '0')
    
      row_configurations = []
      self.phase_to_config_map = {}
      self.phase_display_to_db_mapping = {}
      self.phase_db_to_display_mapping = {}
    
      no_of_rows = 0
    
      # Step 1: Add Live Load and ERSS Wall rows first (unchanged)
      if excavation_type == 'Single wall':
        no_of_rows += 2
        row1_configurations = {
            'row_number': len(row_configurations) + 1,
            'phase_name': 'Activate Live Load',
            'element_type': 'Line Load',
            'element_name': 'LL_Left',
            'action': 'Activate'
        }
        row_configurations.append(row1_configurations)
        
        row2_configurations = {
            'row_number': len(row_configurations) + 1,
            'phase_name': 'Activate_ERSS_Wall',
            'element_type': 'ERSS Wall',
            'element_name': 'Wall_Left',
            'action': 'Activate'
        }
        row_configurations.append(row2_configurations)
        
        for config in [row1_configurations, row2_configurations]:
            phase_key = config['phase_name']
            if phase_key not in self.phase_to_config_map:
                self.phase_to_config_map[phase_key] = []
            self.phase_to_config_map[phase_key].append(config)
            
      elif excavation_type == 'Double wall':
        no_of_rows += 4
        row1_configurations = {
            'row_number': len(row_configurations) + 1,
            'phase_name': 'Activate Live Load',
            'element_type': 'Line Load',
            'element_name': 'LL_Left',
            'action': 'Activate'
        }
        row_configurations.append(row1_configurations)
        
        row2_configurations = {
            'row_number': len(row_configurations) + 1,
            'phase_name': 'Activate Live Load',
            'element_type': 'Line Load',
            'element_name': 'LL_Right',
            'action': 'Activate'
        }
        row_configurations.append(row2_configurations)
        
        row3_configurations = {
            'row_number': len(row_configurations) + 1,
            'phase_name': 'Activate_ERSS_Wall',
            'element_type': 'ERSS Wall',
            'element_name': 'Wall_Left',
            'action': 'Activate'
        }
        row_configurations.append(row3_configurations)
        
        row4_configurations = {
            'row_number': len(row_configurations) + 1,
            'phase_name': 'Activate_ERSS_Wall',
            'element_type': 'ERSS Wall',
            'element_name': 'Wall_Right',
            'action': 'Activate'
        }
        row_configurations.append(row4_configurations)
        
        for config in [row1_configurations, row2_configurations, row3_configurations, row4_configurations]:
            phase_key = config['phase_name']
            if phase_key not in self.phase_to_config_map:
                self.phase_to_config_map[phase_key] = []
            self.phase_to_config_map[phase_key].append(config)
    
    # Step 2: Prepare excavation and strut data
      excavation_configs = []
      strut_configs = []
    
      # Prepare excavation configurations
      if excavation_data:
        try:
            stage_numbers = [int(stage.get('Stage No', '1')) for stage in excavation_data if stage.get('Stage No')]
            max_stage = max(stage_numbers) if stage_numbers else 0
            
            for i in range(1, max_stage + 1):
                stage_data = None
                for stage in excavation_data:
                    if int(stage.get('Stage No', '0')) == i:
                        stage_data = stage
                        break
                
                db_phase_name = f'Excavation Stage {i}'
                if stage_data and 'To' in stage_data:
                    display_phase_name = f'Excavate to {stage_data["To"]}'
                else:
                    display_phase_name = db_phase_name
                
                self.phase_display_to_db_mapping[display_phase_name] = db_phase_name
                self.phase_db_to_display_mapping[db_phase_name] = display_phase_name
                
                excavation_row_config = {
                    'row_number': 0,  # Will be updated later
                    'phase_name': db_phase_name,
                    'display_phase_name': display_phase_name,
                    'element_type': 'Excavation',
                    'element_name': str(i),
                    'action': 'Deactivate'
                }
                excavation_configs.append(excavation_row_config)
            
            # Add over excavation
            db_over_phase_name = 'Over Excavation'
            display_over_phase_name = 'Over Excavation'
            
            self.phase_display_to_db_mapping[display_over_phase_name] = db_over_phase_name
            self.phase_db_to_display_mapping[db_over_phase_name] = display_over_phase_name
            
            over_excavation_row_config = {
                'row_number': 0,  # Will be updated later
                'phase_name': db_over_phase_name,
                'display_phase_name': display_over_phase_name,
                'element_type': 'Excavation',
                'element_name': str(max_stage + 1),
                'action': 'Deactivate'
            }
            excavation_configs.append(over_excavation_row_config)
            
        except (ValueError, AttributeError) as e:
            print(f"DEBUG: Error processing excavation data: {e}")
    
    # Prepare strut configurations
      if num_struts:
        try:
            n = int(num_struts)
            for i in range(1, n + 1):
                strut_row_config = {
                    'row_number': 0,  # Will be updated later
                    'phase_name': f'Install Strut{i}',
                    'element_type': 'Strut',
                    'element_name': f'strut_{i}',
                    'action': 'Activate'
                }
                strut_configs.append(strut_row_config)
        except ValueError:
            print(f"DEBUG: Invalid number of struts: {num_struts}")
    
    # Step 3: Interleave excavation and strut rows
    # Pattern: Excavation Stage 1, Strut 1, Excavation Stage 2, Strut 2, etc.
      max_items = max(len(excavation_configs), len(strut_configs))
    
      for i in range(max_items):
        # Add excavation stage if available (but not over excavation yet)
        if i < len(excavation_configs) - 1:  # -1 to exclude over excavation
            excavation_config = excavation_configs[i]
            excavation_config['row_number'] = len(row_configurations) + 1
            row_configurations.append(excavation_config)
            no_of_rows += 1
            
            # Update phase mapping
            phase_key = excavation_config['phase_name']
            if phase_key not in self.phase_to_config_map:
                self.phase_to_config_map[phase_key] = []
            self.phase_to_config_map[phase_key].append(excavation_config)
        
        # Add strut if available
        if i < len(strut_configs):
            strut_config = strut_configs[i]
            strut_config['row_number'] = len(row_configurations) + 1
            row_configurations.append(strut_config)
            no_of_rows += 1
            
            # Update phase mapping
            phase_key = strut_config['phase_name']
            if phase_key not in self.phase_to_config_map:
                self.phase_to_config_map[phase_key] = []
            self.phase_to_config_map[phase_key].append(strut_config)
    
    # Step 4: Add over excavation at the end
      if excavation_configs and len(excavation_configs) > 0:
        over_excavation = excavation_configs[-1]  # Last item is over excavation
        over_excavation['row_number'] = len(row_configurations) + 1
        row_configurations.append(over_excavation)
        no_of_rows += 1
        
        # Update phase mapping
        phase_key = over_excavation['phase_name']
        if phase_key not in self.phase_to_config_map:
            self.phase_to_config_map[phase_key] = []
        self.phase_to_config_map[phase_key].append(over_excavation)
    
      self.row_configurations = row_configurations
      print(f"DEBUG: Final sequence order:")
      for i, config in enumerate(row_configurations):
        display_name = config.get('display_phase_name', config['phase_name'])
        print(f"  {i+1}. {display_name} - {config['element_type']} ({config['element_name']})")
    
      return row_configurations
    def has_geometry_changed(self, geometry_data: Dict, excavation_data: List) -> bool:
      """Check if geometry or excavation data has changed since last calculation"""
      import hashlib
      import json
    
    # Create a hash of the current geometry and excavation data
      combined_data = {
        'geometry': geometry_data,
        'excavation': excavation_data
      }
    
      try:
        current_hash = hashlib.md5(
            json.dumps(combined_data, sort_keys=True, default=str).encode()
        ).hexdigest()
        
        if not hasattr(self, 'last_geometry_hash') or self.last_geometry_hash != current_hash:
            print(f"DEBUG: Geometry/excavation data has changed. Old hash: {getattr(self, 'last_geometry_hash', 'None')}, New hash: {current_hash}")
            self.last_geometry_hash = current_hash
            return True
            
        return False
      except Exception as e:
        print(f"DEBUG: Error checking geometry changes: {e}")
        return True  # Assume changed if we can't check

    def force_recalculation(self):
      """Force recalculation on next update by clearing the geometry hash"""
      print("DEBUG: Forcing recalculation by clearing geometry hash")
      if hasattr(self, 'last_geometry_hash'):
        self.last_geometry_hash = None
      self.reset_phase_usage_tracking()
    def get_all_phase_names(self) -> List[str]:
        display_phase_names = []
        for config in self.row_configurations:
            if 'display_phase_name' in config:
                display_name = config['display_phase_name']
            else:
                display_name = config['phase_name']
            if display_name not in display_phase_names:
                display_phase_names.append(display_name)
        return display_phase_names

    def format_element_name_for_display(self, element_name: str, element_type: str) -> str:
      """Format element name for UI display"""
      if element_type.lower() == 'excavation' and element_name.isdigit():
        return f"Stage {element_name}"
      return element_name

# 2. Add a helper method to extract numeric value from display name
    def extract_element_name_from_display(self, display_name: str, element_type: str) -> str:
      """Extract the actual element name from display format"""
      if element_type.lower() == 'excavation' and display_name.startswith('Stage '):
        return display_name.replace('Stage ', '').strip()
      return display_name

# 3. Modify the get_options_for_phase_and_row method
    def get_options_for_phase_and_row(self, selected_display_phase_name: str, row_index: int) -> Dict:
      print(f"DEBUG: Getting options for phase '{selected_display_phase_name}' at row {row_index}")
      db_phase_name = self.phase_display_to_db_mapping.get(selected_display_phase_name, selected_display_phase_name)
      if selected_display_phase_name != db_phase_name:
        print(f"DEBUG: Converted display name '{selected_display_phase_name}' to DB name '{db_phase_name}'")
    
      if db_phase_name in self.phase_to_config_map:
        matching_configs = self.phase_to_config_map[db_phase_name]
        if matching_configs:
            element_type = matching_configs[0]['element_type']
            element_name = self.get_element_name_for_row(db_phase_name, element_type, row_index)
            
            # Format element name for display
            display_element_name = self.format_element_name_for_display(element_name, element_type)
            
            result = {
                'element_type_options': [element_type],
                'element_name_options': [display_element_name]
            }
            print(f"DEBUG: Returning options for phase '{selected_display_phase_name}': {result}")
            return result
    
      print(f"DEBUG: No options found for phase '{selected_display_phase_name}'")
      return {'element_type_options': [], 'element_name_options': []}


    def confirm_phase_selection(self, display_phase_name: str, element_type: str):
        db_phase_name = self.phase_display_to_db_mapping.get(display_phase_name, display_phase_name)
        current_count = self.phase_usage_count.get(db_phase_name, 0)
        ordinal = self.get_ordinal_number(current_count + 1)
        
        print(f"DEBUG: User confirmed phase name='{display_phase_name}' for {ordinal} time")
        if display_phase_name != db_phase_name:
            print(f"DEBUG: Display name '{display_phase_name}' mapped to DB name '{db_phase_name}'")
        
        actual_element_name = self.get_next_element_name_for_phase(db_phase_name, element_type)
        print(f"DEBUG: Confirmed element name: '{actual_element_name}' for phase '{display_phase_name}'")
        return actual_element_name

    def convert_display_to_db_phase_name(self, display_phase_name: str) -> str:
        return self.phase_display_to_db_mapping.get(display_phase_name, display_phase_name)

    def convert_db_to_display_phase_name(self, db_phase_name: str) -> str:
        return self.phase_db_to_display_mapping.get(db_phase_name, db_phase_name)

    def get_element_name_for_row(self, phase_name: str, element_type: str, row_index: int) -> str:
      print(f"DEBUG: Getting element name for phase='{phase_name}' at row {row_index}")
    
      if phase_name in self.phase_element_mapping:
        available_names = self.phase_element_mapping[phase_name]
        phase_occurrence = 0
        for i, config in enumerate(self.row_configurations):
            if i >= row_index:
                break
            if config.get('phase_name') == phase_name:
                phase_occurrence += 1
        print(f"DEBUG: Phase '{phase_name}' occurs {phase_occurrence} times before row {row_index}")
        
        if phase_occurrence < len(available_names):
            selected_name = available_names[phase_occurrence]
            print(f"DEBUG: Selected element name '{selected_name}' for phase '{phase_name}' at occurrence {phase_occurrence}")
            
            # Format for display if it's an excavation
            display_name = self.format_element_name_for_display(selected_name, element_type)
            return display_name
        else:
            selected_name = available_names[-1]
            print(f"DEBUG: Max occurrences reached, returning last element name '{selected_name}' for phase '{phase_name}'")
            display_name = self.format_element_name_for_display(selected_name, element_type)
            return display_name
      else:
        matching_configs = self.phase_to_config_map.get(phase_name, [])
        if matching_configs:
            phase_occurrence = 0
            for i, config in enumerate(self.row_configurations):
                if i >= row_index:
                    break
                if config.get('phase_name') == phase_name:
                    phase_occurrence += 1
            
            if phase_occurrence < len(matching_configs):
                selected_name = matching_configs[phase_occurrence]['element_name']
                print(f"DEBUG: Selected element name '{selected_name}' for phase '{phase_name}' from config at occurrence {phase_occurrence}")
                
                # Format for display if it's an excavation
                display_name = self.format_element_name_for_display(selected_name, element_type)
                return display_name
            else:
                selected_name = matching_configs[-1]['element_name']
                print(f"DEBUG: Max config occurrences reached, returning last element name '{selected_name}' for phase '{phase_name}'")
                display_name = self.format_element_name_for_display(selected_name, element_type)
                return display_name
        else:
            print(f"DEBUG: No mapping found for phase '{phase_name}', using element_type as fallback")
            return element_type 
    def get_default_value_for_field(self, row_index: int, field_name: str) -> str:
      if row_index < len(self.row_configurations):
        config = self.row_configurations[row_index]
        
        if field_name == "PhaseNo":
            return str(row_index + 1)
        elif field_name == "PhaseName":
            return config.get('display_phase_name', config['phase_name'])
        elif field_name == "ElementType":
            return config.get('element_type', '')
        elif field_name == "ElementName":
            element_name = config.get('element_name', '')
            element_type = config.get('element_type', '')
            return self.format_element_name_for_display(element_name, element_type)
        elif field_name == "Action":
            return config.get('action', 'Activate')
      return ""
    def get_fields(self, previous_data: Dict = None) -> List[FormField]:
        """Generate form fields based on calculated row configurations."""
        
        # Calculate rows and their specific options
        row_configs = self.calculate_rows_and_options(previous_data)
        
        if not row_configs:
            return [
                FormField("PhaseNo", "text", required=True),
                FormField("PhaseName", "text", required=True),
                FormField("ElementType", "text", required=True),
                FormField("ElementName", "text", required=True),
                FormField("Action", "dropdown", options=["Activate", "Deactivate"], required=True),
            ]
        
        # Get all possible display phase names for the dropdown
        all_display_phase_names = self.get_all_phase_names()
        
        return [
            FormField("PhaseNo", "text", required=True),
            FormField("PhaseName", "dropdown", required=True, options=all_display_phase_names),
            FormField("ElementType", "dropdown", required=True, options=[]),
            FormField("ElementName", "dropdown", required=True, options=[]),
            FormField("Action", "dropdown", options=["Activate", "Deactivate"], required=True),
        ]
    
    def get_filtered_field_options(self, field_name: str, row_index: int, selected_phase_name: str = None) -> List[str]:
        """
        Get filtered options for a specific field and row.
        Enhanced with smart element name tracking and display name handling.
        """
        
        if field_name.lower() == 'phasename':
            result = self.get_all_phase_names()  # This returns display names
            return result
            
        elif field_name.lower() == 'elementtype':
            if selected_phase_name:
                options = self.get_options_for_phase_and_row(selected_phase_name, row_index)
                result = options['element_type_options']
                return result
            else:
                return []
                
        elif field_name.lower() == 'elementname':
            if selected_phase_name:
                options = self.get_options_for_phase_and_row(selected_phase_name, row_index)
                result = options['element_name_options']
                return result
            else:
                return []
                
        elif field_name.lower() == 'action':
            result = ["Activate", "Deactivate"]
            return result
        else:
            return []
    
    def get_total_rows(self) -> int:
        """Get the total number of rows calculated."""
        total = len(self.row_configurations)
        return total
    
    def get_row_configuration(self, row_index: int) -> Dict:
        """Get the complete configuration for a specific row."""
        
        if row_index < len(self.row_configurations):
            config = self.row_configurations[row_index]
            return config
        else:
            return {}

    def validate(self, data: Dict) -> List[str]:
      """Validate that all phase entries are complete"""
      errors = []

    # Check if all fields are present at the top level
      required_fields = ["PhaseNo", "PhaseName", "ElementType", "ElementName", "Action"]
      for field in required_fields:
        if field not in data or not data[field]:
            errors.append(f"{field} is required")
            return errors  # Early return if top-level missing

    # Check each individual phase entry
      try:
        phases = zip(
            data["PhaseNo"], 
            data["PhaseName"], 
            data["ElementType"], 
            data["ElementName"], 
            data["Action"]
        )
        
        for idx, (phase_no, name, etype, ename, action) in enumerate(phases, 1):
            if not all([phase_no, name, etype, ename, action]):
                errors.append(f"All fields must be filled in Phase {idx}")
      except Exception as e:
        errors.append(f"Validation error: {str(e)}")

      return errors
        
  
    def get_model_element_type(self, element_type: str, element_name: str = "") -> str:
        """
        Determine model element type based on element type and additional properties.
        For struts, checks if type is 'fixed' (Point) or 'node to node' (Line).
        """
        if not element_type:
            return "None"
        
        element_type_lower = element_type.lower()
        
        # Handle strut with type checking
        if element_type_lower == "strut":
            # Ensure strut_type_map exists
            if not hasattr(self, 'strut_type_map'):
                print(f"DEBUG: strut_type_map not initialized! Defaulting to 'Line'")
                return "Line"
            
            # Print all cached strut types for debugging
            print(f"DEBUG: Available strut types in cache: {self.strut_type_map}")
            
            # Get the strut type from cached map
            strut_type = self.strut_type_map.get(element_name, "").lower()
            
            # If not found in map, try using global strut type
            if not strut_type and hasattr(self, 'global_strut_type') and self.global_strut_type:
                strut_type = self.global_strut_type
                print(f"DEBUG: Using global strut type '{strut_type}' for '{element_name}'")
            else:
                print(f"DEBUG: Checking strut '{element_name}' with type '{strut_type}'")
            
            # If still not found, try without underscore or with different format
            if not strut_type:
                # Try alternate formats
                alternate_name = element_name.replace('_', '')
                strut_type = self.strut_type_map.get(alternate_name, "").lower()
                if strut_type:
                    print(f"DEBUG: Found strut type using alternate name '{alternate_name}': '{strut_type}'")
            
            if "fixed" in strut_type:
                print(f"DEBUG: Strut '{element_name}' is 'fixed' type -> ModelElementType = 'Point'")
                return "Point"
            elif "node to node" in strut_type:
                print(f"DEBUG: Strut '{element_name}' is 'node to node' type -> ModelElementType = 'Line'")
                return "Line"
            else:
                # Default to Line if strut type is unclear
                print(f"DEBUG: Strut '{element_name}' type unclear ('{strut_type}'), defaulting to 'Line'")
                return "Line"
        
        # Handle other element types
        elif element_type_lower in ["line load", "erss wall"]:
            return "Line"
        elif element_type_lower == "excavation":
            return "Polygon"
        else:
            return "None"
# Modified save method - update the get_model_element_type calls
    def save(self, cursor, data: Dict) -> None:
      """Save form data to both MySQL database and Excel file."""
      try:
        if not data:
            return

        # MySQL save logic
        common_id = data.get('common_id', '')
        
        # First record - Initial Phase
        cursor.execute(
            """
            INSERT INTO SequenceConstruct 
            (common_id, phase_no, phase_name, element_type, element_name, action, model_element_type)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (common_id, "0", "Initial Phase", "None", "None", "None", "None")
        )

        # Prepare data lists
        phase_nos = data.get('PhaseNo', [])
        phase_names = data.get('PhaseName', [])
        element_types = data.get('ElementType', [])
        element_names = data.get('ElementName', [])
        actions = data.get('Action', [])

        # Ensure all values are lists
        phase_nos = phase_nos if isinstance(phase_nos, list) else [phase_nos]
        phase_names = phase_names if isinstance(phase_names, list) else [phase_names]
        element_types = element_types if isinstance(element_types, list) else [element_types]
        element_names = element_names if isinstance(element_names, list) else [element_names]
        actions = actions if isinstance(actions, list) else [actions]

        # Convert display element names back to actual values for saving
        actual_element_names = []
        for i, (element_type, element_name) in enumerate(zip(element_types, element_names)):
            actual_name = self.extract_element_name_from_display(element_name, element_type)
            actual_element_names.append(actual_name)
        
        element_names = actual_element_names
        
        # Determine the maximum length
        max_len = max(len(phase_nos), len(phase_names), len(element_types), len(element_names), len(actions))

        # Normalize list lengths
        phase_nos += [""] * (max_len - len(phase_nos))
        phase_names += [""] * (max_len - len(phase_names))
        element_types += [""] * (max_len - len(element_types))
        element_names += [""] * (max_len - len(element_names))
        actions += [""] * (max_len - len(actions))

        # Insert records into MySQL
        for i in range(max_len):
            phase_no = phase_nos[i] if i < len(phase_nos) else ""
            phase_name = phase_names[i] if i < len(phase_names) else ""
            element_type = element_types[i] if i < len(element_types) else ""
            element_name = element_names[i] if i < len(element_names) else ""
            action = actions[i] if i < len(actions) else ""
            
            # UPDATED: Pass element_name to get_model_element_type
            model_element_type = self.get_model_element_type(element_type, element_name)
            
            cursor.execute(
                """
                INSERT INTO SequenceConstruct 
                (common_id, phase_no, phase_name, element_type, element_name, action, model_element_type)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (common_id, phase_no, phase_name, element_type, element_name, action, model_element_type)
            )

        # Excel save logic (similar update)
        if getattr(sys, 'frozen', False):
            BASE_DIR = Path(sys.executable).parent / "_internal"
        else:
            BASE_DIR = Path(__file__).resolve().parent.parent.parent
        
        excel_dir = BASE_DIR / "data"
        excel_dir.mkdir(exist_ok=True)
        excel_filename = excel_dir / "Input_Data.xlsx"

        if excel_filename.exists():
            workbook = openpyxl.load_workbook(excel_filename)
        else:
            workbook = openpyxl.Workbook()

        sheet_name = "Construction Sequence"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.delete_rows(1, sheet.max_row)
        else:
            sheet = workbook.create_sheet(sheet_name)

        # Write headers
        headers = ["PhaseNo", "PhaseName", "ElementType", "ElementName", "Action", "ModelElementType"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)

        # Write initial phase
        row = 2
        sheet.cell(row=row, column=1, value="Phase_0")
        sheet.cell(row=row, column=2, value="Initial Phase")
        sheet.cell(row=row, column=3, value="None")
        sheet.cell(row=row, column=4, value="")
        sheet.cell(row=row, column=5, value="None")
        sheet.cell(row=row, column=6, value="")

        # Phase assignment logic
        current_phase_num = 1
        previous_phase_name = None
        phase_assignments = {}
        phase_occurrence_count = {}
        
        # First pass: determine phase number assignments
        for i in range(max_len):
            phase_name = phase_names[i] if i < len(phase_names) else ""
            
            if phase_name not in phase_occurrence_count:
                phase_occurrence_count[phase_name] = 0
            
            phase_key = (phase_name, phase_occurrence_count[phase_name])
            
            if previous_phase_name != phase_name:
                if phase_name and phase_key not in phase_assignments:
                    phase_assignments[phase_key] = current_phase_num
                    current_phase_num += 1
                previous_phase_name = phase_name
                phase_occurrence_count[phase_name] += 1
            else:
                if phase_name and phase_key not in phase_assignments:
                    prev_key = (phase_name, phase_occurrence_count[phase_name] - 1)
                    if prev_key in phase_assignments:
                        phase_assignments[phase_key] = phase_assignments[prev_key]
                phase_occurrence_count[phase_name] += 1

        # Reset occurrence counter for second pass
        phase_occurrence_count = {}
        
        # Second pass: write to Excel
        next_row = 3
        
        for i in range(max_len):
            element_type = element_types[i] if i < len(element_types) else ""
            element_name = element_names[i] if i < len(element_names) else ""
            phase_name = phase_names[i] if i < len(phase_names) else ""
            action = actions[i] if i < len(actions) else ""
            
            # UPDATED: Pass element_name to get_model_element_type
            model_element_type = self.get_model_element_type(element_type, element_name)
            
            if phase_name not in phase_occurrence_count:
                phase_occurrence_count[phase_name] = 0
            
            phase_key = (phase_name, phase_occurrence_count[phase_name])
            phase_num = phase_assignments.get(phase_key, current_phase_num)
            phase_str = f"Phase_{phase_num:02d}"
            
            # Write values to cells
            sheet.cell(row=next_row, column=1, value=phase_str)
            sheet.cell(row=next_row, column=2, value=phase_name)
            sheet.cell(row=next_row, column=3, value=element_type)
            sheet.cell(row=next_row, column=5, value=action)
            sheet.cell(row=next_row, column=6, value=model_element_type)
            
            # Special handling for Excavation element names
            if element_type and element_type.lower() == "excavation" and element_name and element_name.isdigit():
                cell = sheet.cell(row=next_row, column=4, value=int(element_name))
                cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                cell.number_format = '0'
            else:
                sheet.cell(row=next_row, column=4, value=element_name)
            
            phase_occurrence_count[phase_name] += 1
            next_row += 1

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            sheet.column_dimensions[column_letter].auto_size = True

        workbook.save(excel_filename)
        print(f"DEBUG: Excel file saved successfully with strut types properly handled")

      except Exception as e:
        if hasattr(self, 'logger'):
            self.logger.error(f"Error saving sequence construct: {str(e)}")
        import traceback
        traceback.print_exc()
        raise
   
    def import_from_csv(self, csv_file_path: str, cursor) -> List[Dict]:
      """Import sequence data from CSV file"""

      try:
        data = []
        with open(csv_file_path, 'r') as file:
            reader = csv.DictReader(file)
            required_fields = ["PhaseNo", "PhaseName", "ElementType", "ElementName", "Action"]
            
            for row in reader:
                # Add this debug line to check PhaseName values
                
                # Validate required fields
                missing = [field for field in required_fields if field not in row]
                if missing:
                    raise ValueError(f"Missing fields in CSV: {', '.join(missing)}")
                
                # Validate action field
                if row["Action"] not in ["Activate", "Deactivate"]:
                    raise ValueError(f"Invalid Action value: {row['Action']}. Must be 'Activate' or 'Deactivate'")
                
                data.append({
                    "PhaseNo": row["PhaseNo"],
                    "PhaseName": row["PhaseName"],
                    "ElementType": row["ElementType"],
                    "ElementName": row["ElementName"],
                    "Action": row["Action"]
                })
            
        return data
    
      except Exception as e:
        raise ValueError(f"Error reading CSV file: {str(e)}")
    def create_sequence_row(self, row_data: Dict) -> ft.DataRow:
      """Create a DataRow from imported data"""
    # Add debug print
    
      return ft.DataRow(
        cells=[
            ft.DataCell(ft.Text(row_data["PhaseNo"])),
            ft.DataCell(ft.Text(row_data["PhaseName"])),  # Make sure this is correctly formatted
            ft.DataCell(ft.Text(row_data["ElementType"])),
            ft.DataCell(ft.Text(row_data["ElementName"])),
            ft.DataCell(ft.Text(row_data["Action"]))
        ]
    )