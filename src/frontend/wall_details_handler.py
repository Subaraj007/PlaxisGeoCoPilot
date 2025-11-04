# Standard Library
from typing import List, Dict, Optional
import logging

# Third-Party Library
import flet as ft
import sqlite3

# Local Module
from frontend.form_section import FormField
import openpyxl
from pathlib import Path 
import sys
import math

class WallDetailsHandler:
    """Handles dynamic wall details field generation and updates based on wall type selection."""
    
    def __init__(self, geometry_section):
        """
        Initialize the wall details handler.
        
        Args:
            geometry_section: Reference to the parent GeometrySection instance
        """
        self.geometry_section = geometry_section
        self.logger = logging.getLogger(__name__)
        
        # Define wall types - UPDATED
        self.wall_types = [
            "Soldier Pile", 
            "Sheet Pile", 
            "Steel Pipe",
            "Contiguous Bored Pile",
            "Secant Bored Pile",
            "Diaphragm Wall"
        ]
        self._concrete_grades_cache = self.load_concrete_grades_from_excel()
        self._steel_grades_cache = self.load_steel_grades_from_excel()  # ADD THIS LINE

        
    def get_wall_type_field(self, current_value: Optional[str] = None) -> FormField:
        """
        Create the wall type dropdown field.
        
        Args:
            current_value: The currently selected wall type
            
        Returns:
            FormField for wall type selection
        """
        return FormField(
            "Wall Type", 
            "dropdown", 
            options=self.wall_types,
            value=current_value
        )
    
    def get_fields_for_wall_type(self, wall_type: str, form_values: Dict) -> List[FormField]:
        """
        Get the appropriate fields based on wall type.
        
        Args:
            wall_type: The selected wall type
            form_values: Dictionary containing current form values
            
        Returns:
            List of FormField objects for the selected wall type
        """
        if wall_type == "Sheet Pile":
            return self._get_sheet_pile_fields(form_values)
        elif wall_type == "Soldier Pile":
            return self._get_soldier_pile_fields(form_values)
        elif wall_type == "Steel Pipe":
            return self._get_steel_pipe_fields(form_values)
        elif wall_type == "Contiguous Bored Pile":
            return self._get_contiguous_pile_fields(form_values)
        elif wall_type == "Secant Bored Pile":
            return self._get_secant_pile_fields(form_values)
        elif wall_type == "Diaphragm Wall":
            return self._get_diaphragm_wall_fields(form_values)
        else:
            self.logger.warning(f"Unknown wall type: {wall_type}")
            return []
    
    def _get_sheet_pile_fields(self, form_values: Dict) -> List[FormField]:
      """Get fields specific to Sheet Pile."""
      sheet_pile_sizes = self.load_sheet_pile_sizes()
      steel_grades = list(self._steel_grades_cache.keys())
    
      fields = [
        FormField("Material", "dropdown", options=["Steel"],
                 value=form_values.get("Material", "Steel")),
        FormField("Sheet Grade", "dropdown", options=steel_grades,  # NEW FIELD
                 value=form_values.get("Sheet Grade")),
        FormField("Member Size", "dropdown", options=sheet_pile_sizes,
                 value=form_values.get("Member Size")),
        FormField("Connection Type", "dropdown", 
                 options=["Interlock", "Non Interlock"],
                 value=form_values.get("Connection Type"))
    ]
    
    # Add Spacing field if Non Interlock
      if form_values.get("Connection Type") == "Non Interlock":
        fields.append(
            FormField("Spacing", "number", "e.g: 30",
                     value=form_values.get("Spacing"))
        )
    
      return fields    
    def _get_soldier_pile_fields(self, form_values: Dict) -> List[FormField]:
        """Get fields specific to Soldier Pile."""
        section_details = self.geometry_section.section_details or []
        material = form_values.get("Material", "Steel")
        concrete_grades = list(self._concrete_grades_cache.keys())

        fields = [
            FormField("Material", "dropdown", options=["Steel", "Concrete"],
                     value=material)
        ]
        
        if material == "Steel":
            fields.extend([
                FormField("Member Size", "dropdown", options=section_details,
                         value=form_values.get("Member Size")),
                FormField("Spacing", "number", "e.g: 1",
                         value=form_values.get("Spacing"), required=False)
            ])
        elif material == "Concrete":
            fields.extend([
                FormField("Grade", "dropdown", options=concrete_grades,
                         value=form_values.get("Grade")),
                FormField("Spacing", "number", "e.g: 1",
                         value=form_values.get("Spacing")),
                FormField("Shape", "dropdown", options=["Rectangular", "Circular"],
                         value=form_values.get("Shape"))
            ])
            
            # Add shape-specific fields
            shape = form_values.get("Shape")
            if shape == "Rectangular":
                fields.extend([
                    FormField("Width", "number", "e.g: 0.3",
                             value=form_values.get("Width")),
                    FormField("Depth", "number", "e.g: 0.3",
                             value=form_values.get("Depth"))
                ])
            elif shape == "Circular":
                fields.append(
                    FormField("Diameter", "number", "e.g: 0.6",
                             value=form_values.get("Diameter"))
                )
        
        return fields
    
    def load_steel_pipe_sizes(self) -> List[str]:
        """Load steel pipe diameter-thickness combinations from Excel."""
        try:
            # Handle both development and executable environments
            if getattr(sys, 'frozen', False):
                base_dir = Path(sys.executable).parent / "_internal"
            else:
                base_dir = Path(__file__).resolve().parent.parent.parent
                
            file_path = base_dir / "data" / "circular_steel_properties.xlsx"
            
            print(f"DEBUG: Attempting to load steel pipe sizes from: {file_path}")
            print(f"DEBUG: File exists: {file_path.exists()}")
            
            workbook = openpyxl.load_workbook(file_path)
            
            # Try to find the correct sheet
            sheet = None
            possible_names = ['Steel Properties', 'Sheet1', 'Properties', 'Data']
            
            for sheet_name in possible_names:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    print(f"DEBUG: Using sheet: {sheet_name}")
                    break
            
            # If still not found, use the first sheet
            if sheet is None:
                sheet = workbook.active
                print(f"DEBUG: Using active sheet: {sheet.title}")
            
            pipe_sizes = []
            current_diameter = None
            
            # Start from row 3 to skip the two header rows
            start_row = 3
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
                col_a = row[0]  # Column A - Diameter
                col_b = row[1] if len(row) > 1 else None  # Column B - Thickness
                
                # Skip empty rows
                if col_a is None and col_b is None:
                    continue
                
                # If Column A has a numeric value, it's a new diameter
                if col_a is not None and col_a != "":
                    try:
                        float(col_a)
                        current_diameter = col_a
                    except (ValueError, TypeError):
                        continue
                
                # If Column B has a numeric value and we have a current diameter, create combination
                if col_b is not None and col_b != "" and current_diameter is not None:
                    try:
                        float(col_b)
                        pipe_size = f"{current_diameter} x {col_b}"
                        pipe_sizes.append(pipe_size)
                    except (ValueError, TypeError):
                        continue
            
            print(f"DEBUG: Total loaded {len(pipe_sizes)} steel pipe size combinations")
            
            # If no data found, return a default list
            if not pipe_sizes:
                print("WARNING: No steel pipe sizes loaded, using default values")
                return ["0.6 x 10", "0.8 x 12", "1.0 x 15"]
            
            return pipe_sizes
            
        except FileNotFoundError:
            print(f"ERROR: File not found at {file_path}")
            print("Using default steel pipe sizes")
            return ["0.6 x 10", "0.8 x 12", "1.0 x 15"]
        except Exception as e:
            print(f"ERROR loading steel pipe sizes: {str(e)}")
            import traceback
            traceback.print_exc()
            print("Using default steel pipe sizes")
            return ["0.6 x 10", "0.8 x 12", "1.0 x 15"]
    # Add this new method to WallDetailsHandler class in wall_details_handler.py


    def load_concrete_grades_from_excel(self) -> Dict[str, Dict]:
      """
      Load concrete grades and their properties from ConcreteDB.xlsx.
    
      Returns:
        Dictionary mapping grade names to their properties:
        {
            'G12': {'fck': 12, 'fcm': 20, 'Ecm': 27.09, 'density': 25, 'poissons_ratio': 0.2},
            'G16': {'fck': 16, 'fcm': 24, 'Ecm': 28.61, 'density': 25, 'poissons_ratio': 0.2},
            ...
          }
      """
      try:
        # Construct file path
        if getattr(sys, 'frozen', False):
            base_dir = Path(sys.executable).parent / "_internal"
        else:
            base_dir = Path(__file__).resolve().parent.parent.parent
            
        file_path = base_dir / "data" / "ConcreteDB.xlsx"
        
        print(f"DEBUG: Loading concrete grades from: {file_path}")
        print(f"DEBUG: File exists: {file_path.exists()}")
        
        if not file_path.exists():
            print(f"ERROR: ConcreteDB.xlsx not found at {file_path}")
            return self._get_default_concrete_grades()
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook['Sheet1']
        
        concrete_grades = {}
        
        # Start from row 2 (skip header row)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            grade = row[0]  # Column A - Grade
            fck = row[1]    # Column B - fck
            fcm = row[2]    # Column C - fcm
            Ecm = row[3]    # Column D - Ecm
            density = row[4]  # Column E - Density
            poissons_ratio = row[5]  # Column F - Poisson's Ratio
            
            # Skip empty rows
            if grade is None or grade == "":
                continue
            
            # Handle string formulas or non-numeric values
            try:
                grade_str = str(grade)
                
                # Convert fck - since data_only=True, we should have numeric values
                if fck is None or fck == "":
                    # If empty, try to extract from grade string (e.g., "G30" -> 30)
                    if grade_str.startswith('G'):
                        fck_val = float(grade_str[1:])
                    else:
                        fck_val = 0
                else:
                    fck_val = float(fck)
                
                # Convert fcm - with data_only=True, formulas should be evaluated
                if fcm is None or fcm == "":
                    fcm_val = fck_val + 8  # fcm = fck + 8
                    print(f"DEBUG: Row {row_idx} - Calculated fcm = {fck_val} + 8 = {fcm_val}")
                else:
                    fcm_val = float(fcm)
                
                # Convert Ecm - with data_only=True, should have the calculated value from Excel
                if Ecm is None or Ecm == "":
                    # Only calculate if missing
                    Ecm_val = 22 * math.pow(fcm_val / 10, 0.3)
                    print(f"DEBUG: Row {row_idx} - Calculated Ecm = 22 * ({fcm_val}/10)^0.3 = {Ecm_val:.2f}")
                else:
                    # Use the value from Excel (should be pre-calculated by Excel)
                    Ecm_val = float(Ecm)
                
                # Handle density and Poisson's ratio
                density_val = float(density) if density is not None else 25
                poissons_val = float(poissons_ratio) if poissons_ratio is not None else 0.2
                
                # Store the grade and its properties
                concrete_grades[grade_str] = {
                    'fck': fck_val,
                    'fcm': fcm_val,
                    'Ecm': Ecm_val,
                    'density': density_val,
                    'poissons_ratio': poissons_val
                }
                
                print(f"DEBUG: Loaded grade {grade_str}: fck={fck_val}, fcm={fcm_val}, Ecm={Ecm_val:.2f}")
                
            except (ValueError, TypeError) as e:
                print(f"WARNING: Error processing row {row_idx} for grade {grade}: {str(e)}")
                continue
        
        if not concrete_grades:
            print("WARNING: No concrete grades loaded from Excel, using defaults")
            return self._get_default_concrete_grades()
        
        print(f"DEBUG: Successfully loaded {len(concrete_grades)} concrete grades")
        return concrete_grades
        
      except FileNotFoundError:
        print(f"ERROR: ConcreteDB.xlsx not found at {file_path}")
        return self._get_default_concrete_grades()
      except Exception as e:
        print(f"ERROR loading concrete grades: {str(e)}")
        import traceback
        traceback.print_exc()
        return self._get_default_concrete_grades()

    def _get_default_concrete_grades(self) -> Dict[str, Dict]:
      """
      Return default concrete grades if Excel file cannot be loaded.
      """
      print("WARNING: Using default concrete grades")
      return {
        'G12': {'fck': 12, 'fcm': 20, 'Ecm': 27.09, 'density': 25, 'poissons_ratio': 0.2},
        'G16': {'fck': 16, 'fcm': 24, 'Ecm': 28.61, 'density': 25, 'poissons_ratio': 0.2},
        'G20': {'fck': 20, 'fcm': 28, 'Ecm': 29.96, 'density': 25, 'poissons_ratio': 0.2},
        'G25': {'fck': 25, 'fcm': 33, 'Ecm': 31.48, 'density': 25, 'poissons_ratio': 0.2},
        'G30': {'fck': 30, 'fcm': 38, 'Ecm': 32.84, 'density': 25, 'poissons_ratio': 0.2},
        'G35': {'fck': 35, 'fcm': 43, 'Ecm': 34.08, 'density': 25, 'poissons_ratio': 0.2},
        'G40': {'fck': 40, 'fcm': 48, 'Ecm': 35.22, 'density': 25, 'poissons_ratio': 0.2},
        'G45': {'fck': 45, 'fcm': 53, 'Ecm': 36.28, 'density': 25, 'poissons_ratio': 0.2},
        'G50': {'fck': 50, 'fcm': 58, 'Ecm': 37.28, 'density': 25, 'poissons_ratio': 0.2},
        'G55': {'fck': 55, 'fcm': 63, 'Ecm': 38.21, 'density': 25, 'poissons_ratio': 0.2},
        'G60': {'fck': 60, 'fcm': 68, 'Ecm': 39.10, 'density': 25, 'poissons_ratio': 0.2},
        'G70': {'fck': 70, 'fcm': 78, 'Ecm': 40.74, 'density': 25, 'poissons_ratio': 0.2},
        'G80': {'fck': 80, 'fcm': 88, 'Ecm': 42.24, 'density': 25, 'poissons_ratio': 0.2},
        'G90': {'fck': 90, 'fcm': 98, 'Ecm': 43.63, 'density': 25, 'poissons_ratio': 0.2}
    }

    def get_concrete_grade_properties(self, grade: str) -> Dict:
      """
      Get properties for a specific concrete grade.
    
      Args:
        grade: Grade string (e.g., 'G30')
        
      Returns:
        Dictionary with grade properties: {'fck', 'fcm', 'Ecm', 'density', 'poissons_ratio'}
      """
      # Load grades if not already loaded
      if not hasattr(self, '_concrete_grades_cache'):
        self._concrete_grades_cache = self.load_concrete_grades_from_excel()
    
    # Get properties for the requested grade
      if grade in self._concrete_grades_cache:
        return self._concrete_grades_cache[grade]
      else:
        print(f"WARNING: Grade {grade} not found in database, using G30 as default")
        return self._concrete_grades_cache.get('G30', {
            'fck': 30, 'fcm': 38, 'Ecm': 32.84, 'density': 25, 'poissons_ratio': 0.2
        })
  
    def load_sheet_pile_sizes(self) -> List[str]:
        """Load sheet pile designations from Excel."""
        try:
            # Handle both development and executable environments
            if getattr(sys, 'frozen', False):
                base_dir = Path(sys.executable).parent / "_internal"
            else:
                base_dir = Path(__file__).resolve().parent.parent.parent
                
            file_path = base_dir / "data" / "Sheet_Pile_Properties.xlsx"
            
            print(f"DEBUG: Attempting to load sheet pile sizes from: {file_path}")
            print(f"DEBUG: File exists: {file_path.exists()}")
            
            workbook = openpyxl.load_workbook(file_path)
            
            # Use Sheet1
            sheet = workbook['Sheet1']
            print(f"DEBUG: Using sheet: Sheet1")
            
            pile_designations = []
            
            # Start from row 3 to skip the two header rows
            start_row = 3
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
                col_a = row[0]  # Column A - Designation
                
                # Skip empty rows
                if col_a is None or col_a == "":
                    continue
                
                # Add the designation to the list
                pile_designations.append(str(col_a))
                
                if row_num <= start_row + 5:  # Print first few for debugging
                    print(f"DEBUG: Row {row_num}: Designation = {col_a}")
            
            print(f"DEBUG: Total loaded {len(pile_designations)} sheet pile designations")
            
            # If no data found, return a default list
            if not pile_designations:
                print("WARNING: No sheet pile designations loaded, using default values")
                return ["YSP1", "YSP U-5", "FSP IA"]
            
            return pile_designations
            
        except FileNotFoundError:
            print(f"ERROR: File not found at {file_path}")
            print("Using default sheet pile designations")
            return ["YSP1", "YSP U-5", "FSP IA"]
        except Exception as e:
            print(f"ERROR loading sheet pile sizes: {str(e)}")
            import traceback
            traceback.print_exc()
            print("Using default sheet pile designations")
            return ["YSP1", "YSP U-5", "FSP IA"]

    def load_steel_grades_from_excel(self) -> Dict[str, Dict]:
      """
      Load steel grades and their properties from SteelDB.xlsx.
      
      Returns:
          Dictionary mapping grade names to their properties:
          {'Grade1': {'density': X, 'poissons_ratio': Y, 'E': Z}, ...}
      """
      try:
        if getattr(sys, 'frozen', False):
            base_dir = Path(sys.executable).parent / "_internal"
        else:
            base_dir = Path(__file__).resolve().parent.parent.parent
            
        file_path = base_dir / "data" / "SteelDB.xlsx"
        
        print(f"DEBUG: Loading steel grades from: {file_path}")
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook['Sheet1']
        
        steel_grades = {}
        
        # Start from row 2 (skip header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            grade = row[0]      # Column A - Grade
            density = row[1]    # Column B - Density
            poissons = row[2]   # Column C - Poisson's Ratio
            E_value = row[3]    # Column D - E value
            
            if grade is None or grade == "":
                continue
            
            steel_grades[str(grade)] = {
                'density': float(density) if density else 78.5,
                'poissons_ratio': float(poissons) if poissons else 0.3,
                'E': float(E_value) if E_value else 210e9
            }
        
        print(f"DEBUG: Loaded {len(steel_grades)} steel grades")
        return steel_grades
        
      except Exception as e:
        print(f"ERROR loading steel grades: {str(e)}")
        return self._get_default_steel_grades()

    def _get_default_steel_grades(self) -> Dict[str, Dict]:
      """Return default steel grades if Excel cannot be loaded."""
      return {
        'S275': {'density': 78.5, 'poissons_ratio': 0.3, 'E': 210e9},
        'S355': {'density': 78.5, 'poissons_ratio': 0.3, 'E': 210e9}
      }
    def load_sheet_pile_sizes(self) -> List[str]:
        """Load sheet pile designations from Excel."""
        try:
            # Handle both development and executable environments
            if getattr(sys, 'frozen', False):
                base_dir = Path(sys.executable).parent / "_internal"
            else:
                base_dir = Path(__file__).resolve().parent.parent.parent
                
            file_path = base_dir / "data" / "Sheet_Pile_Properties.xlsx"
            
            print(f"DEBUG: Attempting to load sheet pile sizes from: {file_path}")
            print(f"DEBUG: File exists: {file_path.exists()}")
            
            workbook = openpyxl.load_workbook(file_path)
            
            # Use Sheet1
            sheet = workbook['Sheet1']
            print(f"DEBUG: Using sheet: Sheet1")
            
            pile_designations = []
            
            # Start from row 3 to skip the two header rows
            start_row = 3
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
                col_a = row[0]  # Column A - Designation
                
                # Skip empty rows
                if col_a is None or col_a == "":
                    continue
                
                # Add the designation to the list
                pile_designations.append(str(col_a))
                
                if row_num <= start_row + 5:  # Print first few for debugging
                    print(f"DEBUG: Row {row_num}: Designation = {col_a}")
            
            print(f"DEBUG: Total loaded {len(pile_designations)} sheet pile designations")
            
            # If no data found, return a default list
            if not pile_designations:
                print("WARNING: No sheet pile designations loaded, using default values")
                return ["YSP1", "YSP U-5", "FSP IA"]
            
            return pile_designations
            
        except FileNotFoundError:
            print(f"ERROR: File not found at {file_path}")
            print("Using default sheet pile designations")
            return ["YSP1", "YSP U-5", "FSP IA"]
        except Exception as e:
            print(f"ERROR loading sheet pile sizes: {str(e)}")
            import traceback
            traceback.print_exc()
            print("Using default sheet pile designations")
            return ["YSP1", "YSP U-5", "FSP IA"]

    def calculate_sheet_pile_properties(self, designation: str, connection_type: str, steel_grade: str, spacing: float = None) -> Dict:

        """
        Calculate EA and EI for sheet pile based on connection type.
        
        Args:
            designation: Sheet pile designation (e.g., "YSP1", "FSP IA")
            connection_type: Either "Interlock" or "Non Interlock"
            spacing: Spacing between piles (required for Non Interlock, in meters)
            
        Returns:
            Dictionary with calculated properties
        """
        try:
            # Load sheet pile properties from Excel
            if getattr(sys, 'frozen', False):
                base_dir = Path(sys.executable).parent / "_internal"
            else:
                base_dir = Path(__file__).resolve().parent.parent.parent
                
            file_path = base_dir / "data" / "Sheet_Pile_Properties.xlsx"
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook['Sheet1']
            
            # Find the row matching the designation
            A_value = None
            I_value = None
            grade_props = self._steel_grades_cache.get(steel_grade, self._get_default_steel_grades()['S275'])
            E = grade_props['E']
            density = grade_props['density']
            poissons_ratio = grade_props['poissons_ratio']
            for row_num, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                col_a = row[0]  # Column A - Designation
                
                if col_a and str(col_a).strip() == designation.strip():
                    if connection_type == "Interlock":
                        # Column H (index 7) for A, Column J (index 9) for I
                        A_value = row[7] if len(row) > 7 else None
                        I_value = row[9] if len(row) > 9 else None
                    else:  # Non Interlock
                        # Column G (index 6) for A, Column I (index 8) for I
                        A_value = row[6] if len(row) > 6 else None
                        I_value = row[8] if len(row) > 8 else None
                    
                    print(f"DEBUG: Found sheet pile {designation} at row {row_num}")
                    print(f"DEBUG: Connection Type: {connection_type}")
                    print(f"DEBUG: A = {A_value}, I = {I_value}")
                    break
            
            if A_value is None or I_value is None:
                raise ValueError(f"Could not find properties for sheet pile designation: {designation}")
            
            # Convert to float
            try:
                A_value = float(A_value)
                I_value = float(I_value)
            except (ValueError, TypeError):
                raise ValueError(f"Invalid A or I values for {designation}: A={A_value}, I={I_value}")
            
            # Steel Young's modulus
            
            # Convert units based on connection type
            if connection_type == "Interlock":
                # For Interlock: A is in cm²/m, I is in cm⁴/m (per wall)
                # Convert to m²/m and m⁴/m
                A_m2_per_m = A_value / 10000  # cm²/m to m²/m
                I_m4_per_m = I_value / 100000000  # cm⁴/m to m⁴/m
                unit_weight = density * A_m2_per_m
    
                # Calculate EA and EI (already per meter of wall)
                EA = E * A_m2_per_m
                EI = E * I_m4_per_m
                
                print(f"DEBUG: Interlock Sheet Pile:")
                print(f"  A = {A_value} cm^2/m = {A_m2_per_m:.6e} m^2/m")
                print(f"  I = {I_value} cm^4/m = {I_m4_per_m:.6e} m^4/m")
                print(f"  E = {E:.2e} Pa")
                print(f"  EA = {EA:.2e} N/m")
                print(f"  EI = {EI:.2e} Nm^2/m")
                
                return {
                    'EA': EA,
                    'EI': EI,
                    'E': E,
                    'A': A_m2_per_m,
                    'I': I_m4_per_m,
                    'unit_weight': unit_weight,
                    'poissons_ratio': poissons_ratio
                }
                
            else:  # Non Interlock
                if spacing is None or spacing <= 0:
                    raise ValueError("Spacing is required for Non Interlock sheet piles")
                
                # For Non Interlock: A is in cm² (per pile), I is in cm⁴ (per pile)
                # Convert to m² and m⁴
                A_m2 = A_value / 10000  # cm² to m²
                I_m4 = I_value / 100000000  # cm⁴ to m⁴
                
                # Calculate EA and EI per meter of wall
                EA = (E * A_m2) / spacing
                EI = (E * I_m4) / spacing
                unit_weight = (density * A_m2) / spacing
        
                print(f"DEBUG: Non Interlock Sheet Pile:")
                print(f"  A = {A_value} cm^2 (per pile) = {A_m2:.6e} m^2")
                print(f"  I = {I_value} cm^4 (per pile) = {I_m4:.6e} m^4")
                print(f"  E = {E:.2e} Pa")
                print(f"  Spacing = {spacing} m")
                print(f"  EA = {EA:.2e} N/m")
                print(f"  EI = {EI:.2e} Nm^2/m")
                
                return {
                    'EA': EA,
                    'EI': EI,
                    'E': E,
                    'A': A_m2,
                    'I': I_m4,
                    'spacing': spacing,
                    'unit_weight': unit_weight,
                    'poissons_ratio': poissons_ratio                
                }
            
        except Exception as ex:
            print(f"ERROR in calculate_sheet_pile_properties: {str(ex)}")
            import traceback
            traceback.print_exc()
            raise

    def _get_steel_pipe_fields(self, form_values: Dict) -> List[FormField]:
      """Get fields specific to Steel Pipe walls."""
      pipe_sizes = self.load_steel_pipe_sizes()
      steel_grades = list(self._steel_grades_cache.keys())
    
      fields = [
        FormField("Material", "dropdown", options=["Steel"],
                 value=form_values.get("Material", "Steel")),
        FormField("Steel Grade", "dropdown", options=steel_grades,  # NEW FIELD
                 value=form_values.get("Steel Grade")),
        FormField("Diameter", "dropdown", options=pipe_sizes,
                 value=form_values.get("Diameter")),
        FormField("Spacing", "number", "e.g: 1",
                 value=form_values.get("Spacing")),
        FormField("Pipe Type", "dropdown", 
                 options=["Hollow", "Filled with Concrete"],
                 value=form_values.get("Pipe Type"))
    ]
    
    # Add Grade field if Filled with Concrete
      if form_values.get("Pipe Type") == "Filled with Concrete":
        concrete_grades = list(self._concrete_grades_cache.keys())
        fields.append(
            FormField("Grade", "dropdown", options=concrete_grades,
                     value=form_values.get("Grade"))
        )
    
      return fields
    def _get_contiguous_pile_fields(self, form_values: Dict) -> List[FormField]:
        """Get fields specific to Contiguous Bored Pile."""
        concrete_grades = list(self._concrete_grades_cache.keys())

        
        return [
            FormField("Material", "dropdown", options=["Concrete"],
                     value=form_values.get("Material", "Concrete")),
            FormField("Grade", "dropdown", options=concrete_grades,
                     value=form_values.get("Grade")),
            FormField("Diameter", "number", "e.g: 0.6",
                     value=form_values.get("Diameter")),
            FormField("Spacing", "number", "e.g: 1",
                     value=form_values.get("Spacing"))
        ]

    def _get_secant_pile_fields(self, form_values: Dict) -> List[FormField]:
        """Get fields specific to Secant Bored Pile."""
        concrete_grades = list(self._concrete_grades_cache.keys())

        
        return [
            FormField("Material", "dropdown", options=["Concrete"],
                     value=form_values.get("Material", "Concrete")),
            FormField("Grade", "dropdown", options=concrete_grades,
                     value=form_values.get("Grade")),
            FormField("Diameter", "number", "e.g: 0.6",
                     value=form_values.get("Diameter")),
            FormField("Spacing", "number", "e.g: 1",
                     value=form_values.get("Spacing"))
        ]

    def _get_diaphragm_wall_fields(self, form_values: Dict) -> List[FormField]:
        """Get fields specific to Diaphragm Wall."""
        concrete_grades = list(self._concrete_grades_cache.keys())

        
        return [
            FormField("Material", "dropdown", options=["Concrete"],
                     value=form_values.get("Material", "Concrete")),
            FormField("Grade", "dropdown", options=concrete_grades,
                     value=form_values.get("Grade")),
            FormField("Thickness", "number", "e.g: 0.8",
                     value=form_values.get("Thickness"))
        ]

    def handle_wall_type_change(self, e, parent_form) -> None:
        """Handle wall type change event and update the UI accordingly."""
        try:
            if not parent_form or not hasattr(parent_form, 'form_content'):
                self.logger.error("Parent form not initialized properly")
                return
                
            wall_type = e.data
            self.geometry_section.form_values["Wall Type"] = wall_type
            
            self.logger.info(f"Wall type changed to: {wall_type}")
            
            # Store current values before regenerating fields
            current_values = self.geometry_section.form_values.copy()
            current_values["Wall Type"] = wall_type
            
            # Regenerate all fields
            all_fields = self.geometry_section.get_fields()
            if "Steel Grade" in current_values:
               print(f"DEBUG: Preserving Steel Grade: {current_values['Steel Grade']}")
            if "Sheet Grade" in current_values:
               print(f"DEBUG: Preserving Sheet Grade: {current_values['Sheet Grade']}")
            # Find the Wall Details frame
            wall_frame = self._find_wall_details_frame(parent_form)
            
            if not wall_frame:
                self.logger.error("Wall Details frame not found")
                return
            
            # Get wall-specific fields
            wall_fields = self._extract_wall_fields(all_fields)
            
            # Create new wall controls
            wall_controls = self.geometry_section._create_category_controls(
                wall_fields, 
                current_values
            )
            
            # Update the wall frame content
            self._update_wall_frame_content(wall_frame, wall_controls)
            
            # Make frame visible
            wall_frame.visible = True
            
            # Update the UI
            self._update_ui(parent_form)
            
            self.logger.info(f"Successfully updated wall details for type: {wall_type}")
    
        except Exception as ex:
            self.logger.error(f"Error in handle_wall_type_change: {str(ex)}", exc_info=True)

    def handle_material_change(self, e, parent_form) -> None:
        """Handle material change for Soldier Pile."""
        try:
            material = e.data or e.control.value
            self.geometry_section.form_values["Material"] = material
            current_values = self.geometry_section.form_values.copy()
            all_fields = self.geometry_section.get_fields()
            wall_frame = self._find_wall_details_frame(parent_form)
            
            if not wall_frame:
                return
            
            wall_fields = self._extract_wall_fields(all_fields)
            wall_controls = self.geometry_section._create_category_controls(wall_fields, current_values)
            self._update_wall_frame_content(wall_frame, wall_controls)
            self._update_ui(parent_form)
            
        except Exception as ex:
            print(f"ERROR in handle_material_change: {str(ex)}")

    def handle_shape_change(self, e, parent_form) -> None:
        """Handle shape change for Concrete Soldier Pile."""
        try:
            shape = e.data or e.control.value
            self.geometry_section.form_values["Shape"] = shape
            current_values = self.geometry_section.form_values.copy()
            all_fields = self.geometry_section.get_fields()
            wall_frame = self._find_wall_details_frame(parent_form)
            
            if not wall_frame:
                return
            
            wall_fields = self._extract_wall_fields(all_fields)
            wall_controls = self.geometry_section._create_category_controls(wall_fields, current_values)
            self._update_wall_frame_content(wall_frame, wall_controls)
            self._update_ui(parent_form)
            
        except Exception as ex:
            print(f"ERROR in handle_shape_change: {str(ex)}")

    def handle_pipe_type_change(self, e, parent_form) -> None:
        """Handle pipe type change and show/hide Grade field."""
        try:
            pipe_type = e.data or e.control.value
            self.geometry_section.form_values["Pipe Type"] = pipe_type
            current_values = self.geometry_section.form_values.copy()
            all_fields = self.geometry_section.get_fields()
            wall_frame = self._find_wall_details_frame(parent_form)
            
            if not wall_frame:
                return
            
            wall_fields = self._extract_wall_fields(all_fields)
            wall_controls = self.geometry_section._create_category_controls(wall_fields, current_values)
            self._update_wall_frame_content(wall_frame, wall_controls)
            self._update_ui(parent_form)
            
        except Exception as ex:
            print(f"ERROR in handle_pipe_type_change: {str(ex)}")

    def handle_connection_type_change(self, e, parent_form) -> None:
        """Handle connection type change for Sheet Pile."""
        try:
            connection_type = e.data or e.control.value
            self.geometry_section.form_values["Connection Type"] = connection_type
            current_values = self.geometry_section.form_values.copy()
            all_fields = self.geometry_section.get_fields()
            wall_frame = self._find_wall_details_frame(parent_form)
            
            if not wall_frame:
                return
            
            wall_fields = self._extract_wall_fields(all_fields)
            wall_controls = self.geometry_section._create_category_controls(wall_fields, current_values)
            self._update_wall_frame_content(wall_frame, wall_controls)
            self._update_ui(parent_form)
            
        except Exception as ex:
            print(f"ERROR in handle_connection_type_change: {str(ex)}")
    
    def _find_wall_details_frame(self, parent_form) -> Optional[ft.Container]:
        """Find the Wall Details frame in the form."""
        for container in parent_form.form_content.controls:
            if (isinstance(container, ft.Container) and 
                hasattr(container, 'content') and 
                isinstance(container.content, ft.Column) and
                container.content.controls and
                isinstance(container.content.controls[0], ft.Text) and
                container.content.controls[0].value == "Wall Details"):
                return container
        return None
    
    def _extract_wall_fields(self, all_fields: List[FormField]) -> List[FormField]:
        """Extract only the wall-related fields from all fields."""
        wall_field_names = [
            "Wall Type", "Material","Steel Grade",
            "Sheet Grade", "Member Size", "Spacing",
            "Diameter", "Pipe Type", "Grade", "Thickness",
            "Shape", "Width", "Depth", "Connection Type"
        ]
        return [field for field in all_fields if field.label in wall_field_names]
    
    def _update_wall_frame_content(self, wall_frame: ft.Container, 
                                   wall_controls: ft.Column) -> None:
        """Update the wall frame with new controls."""
        if len(wall_frame.content.controls) > 1:
            wall_frame.content.controls[1] = wall_controls
        else:
            wall_frame.content.controls.append(wall_controls)
    
    def _update_ui(self, parent_form) -> None:
        """Update the UI after changes."""
        if parent_form and parent_form.page:
            parent_form.page.update()
        else:
            self.logger.warning("Could not update UI - page reference not found")

    def calculate_hollow_pipe_properties(self, diameter_thickness: str, spacing: float, steel_grade: str) -> Dict:

        """Calculate EA and EI for hollow steel pipe."""
        try:
            # Parse diameter and thickness from the string
            parts = diameter_thickness.split('x')
            if len(parts) != 2:
                raise ValueError(f"Invalid diameter_thickness format: {diameter_thickness}")
            
            diameter = float(parts[0].strip())
            thickness = float(parts[1].strip())
            
            # Load properties from Excel
            if getattr(sys, 'frozen', False):
                base_dir = Path(sys.executable).parent / "_internal"
            else:
                base_dir = Path(__file__).resolve().parent.parent.parent
                
            file_path = base_dir / "data" / "circular_steel_properties.xlsx"
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # Find matching row in Excel
            A_value = None
            I_value = None
            grade_props = self._steel_grades_cache.get(steel_grade, self._get_default_steel_grades()['S275'])
            E = grade_props['E']
            density = grade_props['density']
            poissons_ratio = grade_props['poissons_ratio']
            current_diameter = None
            for row in sheet.iter_rows(min_row=3, values_only=True):
                col_a = row[0]  # Diameter
                col_b = row[1]  # Thickness
                col_c = row[2]  # Area of section (A)
                col_d = row[3]  # Second moment of Inertia (I)
                col_e = row[4]  # Second moment of Inertia (I)

                # Update current diameter if found
                if col_a is not None and col_a != "":
                    try:
                        current_diameter = float(col_a)
                    except (ValueError, TypeError):
                        continue
                
                # Check if this row matches our diameter and thickness
                if current_diameter == diameter and col_b is not None:
                    try:
                        if float(col_b) == thickness:
                            A_value = float(col_d) if col_d else None
                            I_value = float(col_e) if col_e else None
                            break
                    except (ValueError, TypeError):
                        continue
            
            if A_value is None or I_value is None:
                raise ValueError(f"Could not find properties for {diameter_thickness}")
            
            print(f"DEBUG: Found properties - A: {A_value} cm^2, I: {I_value} cm^4")
            
            # Convert A from cm² to m²
            A_m2 = A_value / 10000
            
            # Convert I from cm⁴ to m⁴
            I_m4 = I_value / 100000000
            
            
            # Calculate EA and EI
            EA = (E * A_m2) / spacing
            EI = (E * I_m4) / spacing
            unit_weight = (density * A_m2) / spacing
   
            print(f"DEBUG: Hollow Pipe - Diameter={diameter}mm, Thickness={thickness}mm")
            print(f"DEBUG: A={A_value}cm^2, I={I_value}cm^4, Spacing={spacing}m")
            print(f"DEBUG: EA={EA:.2e} N/m, EI={EI:.2e} Nm^2/m")
            
            return {
                'EA': EA,
                'EI': EI,
                'E': E,
                'A': A_m2,
                'I': I_m4,
                'diameter': diameter / 1000  ,# Convert mm to m for storage
                'unit_weight': unit_weight,
                'poissons_ratio': poissons_ratio
            }
            
            
        except Exception as ex:
            print(f"ERROR in calculate_hollow_pipe_properties: {str(ex)}")
            import traceback
            traceback.print_exc()
            raise

    def calculate_filled_pipe_properties(self, diameter_thickness: str, grade: float, spacing: float, steel_grade: str) -> Dict:
        """Calculate EA and EI for concrete-filled steel pipe using composite section approach."""
        try:
            # Convert to string if not already
            if not isinstance(diameter_thickness, str):
                diameter_thickness = str(diameter_thickness)
            
            # Parse diameter and thickness from the string
            if 'x' not in diameter_thickness:
                raise ValueError(f"Invalid diameter_thickness format: {diameter_thickness}. Expected format: 'DiameterxThickness'")
            
            parts = diameter_thickness.split('x')
            if len(parts) != 2:
                raise ValueError(f"Invalid diameter_thickness format: {diameter_thickness}")
            
            D0 = float(parts[0].strip())  # Outer diameter in mm
            thickness = float(parts[1].strip())  # Thickness in mm
            
            print(f"DEBUG: Parsing filled pipe - D0={D0}mm, thickness={thickness}mm, grade=G{grade}")
            
            # Load steel pipe properties from Excel
            if getattr(sys, 'frozen', False):
                base_dir = Path(sys.executable).parent / "_internal"
            else:
                base_dir = Path(__file__).resolve().parent.parent.parent
                
            file_path = base_dir / "data" / "circular_steel_properties.xlsx"
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # Find matching row in Excel for steel properties
            A0_value = None  # Area of steel section (column D, index 3)
            I0_value = None  # Second moment of inertia of steel (column E, index 4)
            
            current_diameter = None
            for row in sheet.iter_rows(min_row=3, values_only=True):
                col_a = row[0]  # Diameter
                col_b = row[1]  # Thickness
                col_d = row[3]  # Area (A0) - 4th column
                col_e = row[4]  # Inertia (I0) - 5th column
                
                # Update current diameter if found
                if col_a is not None and col_a != "":
                    try:
                        current_diameter = float(col_a)
                    except (ValueError, TypeError):
                        continue
                
                # Check if this row matches our diameter and thickness
                if current_diameter == D0 and col_b is not None:
                    try:
                        if float(col_b) == thickness:
                            A0_value = float(col_d) if col_d else None
                            I0_value = float(col_e) if col_e else None
                            break
                    except (ValueError, TypeError):
                        continue
            
            if A0_value is None or I0_value is None:
                raise ValueError(f"Could not find steel properties for {diameter_thickness}")
            
            print(f"DEBUG: Found steel properties - A0: {A0_value} cm^2, I0: {I0_value} cm^4")
            
            # Convert steel properties from cm units to m units
            A0 = A0_value / 10000  # cm² to m²
            I0 = I0_value / 100000000  # cm⁴ to m⁴
            
            # Steel Young's modulus
            
            # Calculate concrete inner diameter
            D1 = D0 - 2 * thickness  # mm
            D1_m = D1 / 1000  # Convert to meters

            # Get STEEL properties from steel_grade parameter
            steel_grade_props = self._steel_grades_cache.get(steel_grade, self._get_default_steel_grades()['S275'])
            E0 = steel_grade_props['E']  # Steel Young's modulus in Pa
            steel_density = steel_grade_props['density']  # Steel density in kN/m³
            steel_poissons = steel_grade_props['poissons_ratio']  # Steel Poisson's ratio
            
            # Get CONCRETE properties
            grade_str = f"G{int(grade)}" if not str(grade).startswith('G') else str(grade)
            concrete_grade_props = self.get_concrete_grade_properties(grade_str)
            E1 = concrete_grade_props['Ecm'] * 1e9  # Convert GPa to Pa - Concrete Young's modulus
            concrete_density = concrete_grade_props['density']  # Concrete density in kN/m³
            concrete_poissons = concrete_grade_props['poissons_ratio']  # Concrete Poisson's ratio
            
            print(f"DEBUG: Steel Grade {steel_grade}: E={E0:.2e} Pa, density={steel_density} kN/m³")
            print(f"DEBUG: Concrete Grade {grade_str}: Ecm={concrete_grade_props['Ecm']} GPa, density={concrete_density} kN/m³")
            
            # Calculate concrete area: A1 = pi * D1^2 / 4
            A1 = math.pi * (D1_m ** 2) / 4  # m^2
            
            # Calculate concrete moment of inertia: I1 = pi * D1^4 / 64
            I1 = math.pi * (D1_m ** 4) / 64  # m^4
            
            # Calculate composite EA and EI
            EA = (E0 * A0 + E1 * A1) / spacing
            EI = (E0 * I0 + E1 * I1) / spacing
            
            # Unit weight: w = (steel_density * A0 + concrete_density * A1) / spacing
            unit_weight = (steel_density * A0 ) / spacing
            print(f"DEBUG: steel_density={steel_density} kg/m^3, Area steel A0={A0} m^2,spacing={spacing} m,so unit_weight={unit_weight} kN/m") 
            print(f"DEBUG: Filled Pipe Composite Section:")
            print(f"  Outer Diameter D0: {D0} mm, Thickness: {thickness} mm")
            print(f"  Inner Diameter D1: {D1} mm ({D1_m:.4f} m)")
            print(f"  Steel: E0={E0:.2e} Pa, A0={A0:.6e} m^2, I0={I0:.6e} m^4")
            print(f"  Concrete: E1={E1:.2e} Pa, A1={A1:.6e} m^2, I1={I1:.6e} m^4")
            print(f"  Spacing: {spacing} m")
            print(f"  Composite: EA={EA:.2e} N/m, EI={EI:.2e} Nm^2/m")
            print(f"  Unit Weight: {unit_weight:.2f} kN/m")
            
            return {
                'EA': EA,
                'EI': EI,
                'E': E0,  # Return steel E for material definition
                'A': A0 + A1,  # Total composite area
                'I': I0 + I1,  # Total composite inertia
                'diameter': D0 / 1000 , # Outer diameter in meters
                'diameter': D0 / 1000,  # Outer diameter in meters
                'unit_weight': unit_weight,
                'poissons_ratio': steel_poissons
            
            }
            
        except Exception as ex:
            print(f"ERROR in calculate_filled_pipe_properties: {str(ex)}")
            import traceback
            traceback.print_exc()
            raise

    def save_steel_pipe_properties(self, cursor, data: Dict, common_id: str,
                               excel_sheets: Dict, sheets_config: Dict) -> None:
        """Save steel pipe properties to database, Excel, and CSV."""
        try:
            pipe_type = data.get("Pipe Type")
            diameter_str = data.get("Diameter")
            spacing = data.get("Spacing")
            
            if not diameter_str or not spacing:
                print("WARNING: Missing diameter or spacing for Steel Pipe")
                return
            
            # Convert diameter to string if it's not already
            if not isinstance(diameter_str, str):
                diameter_str = str(diameter_str)
            
            print(f"DEBUG: Processing Steel Pipe - Diameter: '{diameter_str}', Type: {pipe_type}, Spacing: {spacing}")
            
            try:
                spacing = float(spacing)
            except (ValueError, TypeError):
                print("ERROR: Invalid spacing value")
                return
            steel_grade = data.get("Steel Grade")  # GET NEW FIELD
            if not steel_grade:
                print("WARNING: Missing Steel Grade")
                return
            # Calculate properties based on pipe type
            if pipe_type == "Hollow":
                # For hollow pipes, diameter_str is in format "DiameterxThickness"
                props = self.calculate_hollow_pipe_properties(diameter_str, spacing, steel_grade)
                
            elif pipe_type == "Filled with Concrete":
                grade = data.get("Grade")
                if not grade:
                    print("WARNING: Missing grade for Filled Steel Pipe")
                    return
                
                try:
                    # Extract numeric grade value (e.g., "G30" -> 30)
                    if isinstance(grade, str) and grade.startswith('G'):
                        grade = float(grade[1:])
                    else:
                        grade = float(grade)
                    
                except (ValueError, TypeError):
                    print("ERROR: Invalid grade value")
                    return
                
                # Ensure diameter_str is in the correct format (DiameterxThickness)
                if 'x' not in diameter_str:
                    print(f"ERROR: Invalid diameter format: '{diameter_str}'. Expected format: 'DiameterxThickness' (e.g., '600 x 10')")
                    return
                
                # Pass the full diameter_thickness string to the calculation function
                print(f"DEBUG: Calling calculate_filled_pipe_properties with diameter_str='{diameter_str}', grade={grade}, spacing={spacing}")
                props = self.calculate_filled_pipe_properties(diameter_str, grade, spacing, steel_grade)
            else:
                print("WARNING: Unknown pipe type")
                return
            
            # Prepare plate properties data
            material_name = f"Steel_Pipe_{pipe_type.replace(' ', '_')}"
            
            plate_data = [
                material_name,  # MaterialName
                "Elastic",  # Elasticity
                True,  # IsIsotropic
                props['EA'],  # EA
                props['EA'],  # EA2 (same as EA for pipes)
                props['EI'],  # EI
                props['diameter'],  # d (diameter in meters)
                props['E'],  # E (Young's modulus)
                props['EA'] / props['diameter'] if props['diameter'] > 0 else 0,  # Gref
                props['unit_weight'],  # UPDATED
                props['poissons_ratio'],  # UPDATED
                17523200  # Colour
            ]
            
            # Save to database - check if table exists first
            try:
                # Try to get the table name from sheets_config
                table_name = sheets_config.get('Plate Properties', {}).get('db_table', 'plate_properties')
                
                cursor.execute(
                    f"INSERT INTO {table_name} "
                    "(common_id, MaterialName, Elasticity, IsIsotropic, EA, EA2, EI, d, E, Gref, w, StrutNu, Colour) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    [common_id] + plate_data
                )
            except sqlite3.OperationalError as db_error:
                print(f"WARNING: Database table issue: {str(db_error)}")
                print("Skipping database insertion for Plate Properties")
            
            # Save to Excel
            excel_sheets["Plate Properties"].append(plate_data)
            
            # Save to CSV
            import csv
            with open(sheets_config["Plate Properties"]["csv_file"], 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([common_id] + plate_data)
            
            print(f"DEBUG: Saved Steel Pipe Plate Properties - EA={props['EA']:.2e}, EI={props['EI']:.2e}")
            
        except Exception as ex:
            print(f"ERROR in save_steel_pipe_properties: {str(ex)}")
            import traceback
            traceback.print_exc()
            raise

    def save_sheet_pile_properties(self, cursor, data: Dict, common_id: str,
                               excel_sheets: Dict, sheets_config: Dict) -> None:
        """Save sheet pile properties to database, Excel, and CSV."""
        try:
            member_size = data.get("Member Size")
            connection_type = data.get("Connection Type")
            spacing = data.get("Spacing")
            
            if not member_size or not connection_type:
                print("WARNING: Missing member size or connection type for Sheet Pile")
                return
            
            print(f"DEBUG: Processing Sheet Pile - Designation: '{member_size}', Connection: {connection_type}, Spacing: {spacing}")
            sheet_grade = data.get("Sheet Grade")  # GET NEW FIELD
            if not sheet_grade:
              print("WARNING: Missing Sheet Grade")
              return
            # For Non Interlock, spacing is required
            if connection_type == "Non Interlock":
                if not spacing:
                    print("ERROR: Spacing is required for Non Interlock sheet piles")
                    return
                try:
                    spacing = float(spacing)
                except (ValueError, TypeError):
                    print("ERROR: Invalid spacing value")
                    return
            else:
                spacing = None  # Not needed for Interlock
            
            # Calculate properties
            props = self.calculate_sheet_pile_properties(member_size, connection_type, sheet_grade, spacing)
            
            # Prepare plate properties data
            material_name = f"Sheet_Pile_{connection_type.replace(' ', '_')}"
            
            plate_data = [
                material_name,  # MaterialName
                "Elastic",  # Elasticity
                True,  # IsIsotropic
                props['EA'],  # EA
                props['EA'],  # EA2 (same as EA for sheet piles)
                props['EI'],  # EI
                0.001,  # d (dummy thickness value for sheet piles)
                props['E'],  # E (Young's modulus)
                props['EA'] / 0.001,  # Gref
                props['unit_weight'],  # UPDATED
                props['poissons_ratio'], # StrutNu (Poisson's ratio)
                17523200  # Colour
            ]
            
            # Save to database - check if table exists first
            try:
                table_name = sheets_config.get('Plate Properties', {}).get('db_table', 'plate_properties')
                
                cursor.execute(
                    f"INSERT INTO {table_name} "
                    "(common_id, MaterialName, Elasticity, IsIsotropic, EA, EA2, EI, d, E, Gref, w, StrutNu, Colour) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    [common_id] + plate_data
                )
            except sqlite3.OperationalError as db_error:
                print(f"WARNING: Database table issue: {str(db_error)}")
                print("Skipping database insertion for Plate Properties")
            
            # Save to Excel
            excel_sheets["Plate Properties"].append(plate_data)
            
            # Save to CSV
            import csv
            with open(sheets_config["Plate Properties"]["csv_file"], 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([common_id] + plate_data)
            
            print(f"DEBUG: Saved Sheet Pile Plate Properties - EA={props['EA']:.2e}, EI={props['EI']:.2e}")
            
        except Exception as ex:
            print(f"ERROR in save_sheet_pile_properties: {str(ex)}")
            import traceback
            traceback.print_exc()
            raise


    # Add this method to the WallDetailsHandler class

    def calculate_unit_weight(self, wall_type: str, density: float, A: float = None, 
                            spacing: float = None, diameter: float = None) -> float:
        """
        Calculate unit weight (w) for concrete walls based on wall type.
        
        Args:
            wall_type: Type of wall (e.g., "Soldier Pile", "Secant Bored Pile")
            density: Concrete density in kN/m³
            A: Cross-sectional area in m² (required for most types)
            spacing: Spacing between piles in m (required for soldier/contiguous piles)
            diameter: Diameter in m (required for secant bored pile)
            
        Returns:
            Unit weight in kN/m
        """
        try:
            # Net density (subtract 20 kN/m³ for buoyancy)
            net_density = density - 20
            
            if wall_type in ["Soldier Pile", "Contiguous Bored Pile"]:
                # w = (density - 20) * A / spacing
                if A is None or spacing is None:
                    raise ValueError(f"A and spacing required for {wall_type}")
                w = (net_density * A) / spacing
                print(f"DEBUG: {wall_type} unit weight: w = ({density}-20) * {A:.6f} / {spacing} = {w:.4f} kN/m")
                
            elif wall_type == "Secant Bored Pile":
                # w = (density - 20) * D
                if diameter is None:
                    raise ValueError("Diameter required for Secant Bored Pile")
                w = net_density * diameter
                print(f"DEBUG: Secant Bored Pile unit weight: w = ({density}-20) * {diameter} = {w:.4f} kN/m")
                
            elif wall_type == "Diaphragm Wall":
                # w = (density - 20) * A
                if A is None:
                    raise ValueError("A required for Diaphragm Wall")
                w = net_density * A
                print(f"DEBUG: Diaphragm Wall unit weight: w = ({density}-20) * {A:.6f} = {w:.4f} kN/m")
                
            else:
                raise ValueError(f"Unknown wall type for unit weight calculation: {wall_type}")
            
            return w
            
        except Exception as ex:
            print(f"ERROR in calculate_unit_weight: {str(ex)}")
            raise

    def calculate_concrete_soldier_pile_properties(self, wall_type: str, grade: float, spacing: float, 
                                               shape: str, **dimensions) -> Dict:
      """
      Calculate properties for concrete piles (Soldier Pile, Contiguous Bored Pile, Secant Bored Pile).
      
      Args:
          wall_type: Type of wall ("Soldier Pile", "Contiguous Bored Pile", "Secant Bored Pile")
          grade: Concrete grade
          spacing: Spacing between piles
          shape: Shape of pile ("Rectangular" or "Circular")
          **dimensions: width, depth (for rectangular) or diameter (for circular)
      """
      try:
        # Calculate E using the formula: E = 22000 * [(Grade + 8) / 10] ^ 0.3
        grade_str = f"G{int(grade)}" if not str(grade).startswith('G') else str(grade)
        grade_props = self.get_concrete_grade_properties(grade_str)
        
        # Extract values from database
        fck = grade_props['fck']
        fcm = grade_props['fcm']
        Ecm = grade_props['Ecm']
        density = grade_props['density']
        poissons_ratio = grade_props['poissons_ratio']
        
        # Convert Ecm from GPa to Pa
        E_Pa = Ecm * 1e9
        
        print(f"DEBUG: Using concrete properties from database:")
        print(f"  Grade: {grade_str}")
        print(f"  fck: {fck} MPa")
        print(f"  fcm: {fcm} MPa")
        print(f"  Ecm: {Ecm} GPa = {E_Pa:.2e} Pa")
        print(f"  Density: {density} kN/m³")
        print(f"  Poisson's Ratio: {poissons_ratio}")
        
        # Rest of the calculation remains the same...
        if shape == "Rectangular":
            width = dimensions.get('width')
            depth = dimensions.get('depth')
            
            if width is None or depth is None:
                raise ValueError("Width and depth are required for rectangular shape")
            
            width = float(width)
            depth = float(depth)
            
            # Calculate A and I for rectangular section
            A = width * depth  # m²
            I = (width * depth * depth * depth) / 12  # m⁴
            
            print(f"DEBUG: Rectangular Concrete Soldier Pile:")
            print(f"  Grade: G{grade}, Width: {width}m, Depth: {depth}m")
            print(f"  E = 22000 * [({grade}+8)/10]^0.3 = {Ecm:.2f} MPa = {E_Pa:.2e} Pa")
            print(f"  A = {width} * {depth} = {A:.6f} m^2")
            print(f"  I = ({width} * {depth}^3) / 12 = {I:.6e} m^4")
            
        elif shape == "Circular":
            diameter = dimensions.get('diameter')
            
            if diameter is None:
                raise ValueError("Diameter is required for circular shape")
            
            diameter = float(diameter)
            
            # Calculate A and I for circular section
            A = (math.pi * diameter * diameter) / 4  # m²
            I = (math.pi * math.pow(diameter, 4)) / 64  # m⁴
            
            print(f"DEBUG: Circular {wall_type}:")
            print(f"  Grade: G{grade}, Diameter: {diameter}m")
            print(f"  E = 22000 * [({grade}+8)/10]^0.3 = {Ecm:.2f} MPa = {E_Pa:.2e} Pa")
            print(f"  A = (pi * {diameter}^2) / 4 = {A:.6f} m^2")
            print(f"  I = (pi * {diameter}^4) / 64 = {I:.6e} m^4")
        else:
            raise ValueError(f"Unknown shape: {shape}")
        
        # Calculate EA and EI per meter of wall
        EA = (E_Pa * A) / spacing
        EI = (E_Pa * I) / spacing
        
        # Calculate unit weight based on wall type
        if wall_type == "Secant Bored Pile":
            # Secant Bored Pile: w = (density - 20) * D
            unit_weight = self.calculate_unit_weight(
                "Secant Bored Pile", 
                density, 
                diameter=dimensions.get('diameter')
            )
        elif wall_type in ["Soldier Pile", "Contiguous Bored Pile"]:
            # Soldier Pile and Contiguous Bored Pile: w = (density - 20) * A / spacing
            unit_weight = self.calculate_unit_weight(
                wall_type, 
                density, 
                A, 
                spacing
            )
        else:
            # Default fallback
            unit_weight = self.calculate_unit_weight(
                "Soldier Pile", 
                density, 
                A, 
                spacing
            )
        
        print(f"  Spacing: {spacing}m")
        print(f"  EA = (E * A) / spacing = {EA:.2e} N/m")
        print(f"  EI = (E * I) / spacing = {EI:.2e} Nm2/m")
        print(f"  Unit Weight = {unit_weight:.4f} kN/m")
        
        return {
            'EA': EA,
            'EI': EI,
            'E': E_Pa,
            'A': A,
            'I': I,
            'spacing': spacing,
            'shape': shape,
            'unit_weight': unit_weight,
            'poissons_ratio': poissons_ratio
        }
        
      except Exception as ex:
        print(f"ERROR in calculate_concrete_soldier_pile_properties: {str(ex)}")
        import traceback
        traceback.print_exc()
        raise
  
    def save_concrete_soldier_pile_properties(self, cursor, data: Dict, common_id: str,
                                          excel_sheets: Dict, sheets_config: Dict) -> None:

      try:
        wall_type = data.get("Wall Type")
        grade = data.get("Grade")
        spacing = data.get("Spacing")
        shape = data.get("Shape")  # Only for Soldier Pile
        
        if not grade or not spacing:
            print(f"WARNING: Missing grade or spacing for {wall_type}")
            return
        
        # Extract numeric grade value (e.g., "G30" -> 30)
        if isinstance(grade, str) and grade.startswith('G'):
            grade = float(grade[1:])
        else:
            grade = float(grade)
        
        spacing = float(spacing)
        
        # Determine shape and get dimensions
        dimensions = {}
        
        if wall_type == "Soldier Pile":
            # Soldier Pile can be Rectangular or Circular
            if not shape:
                print("WARNING: Missing shape for Concrete Soldier Pile")
                return
                
            if shape == "Rectangular":
                width = data.get("Width")
                depth = data.get("Depth")
                if not width or not depth:
                    print("WARNING: Missing width or depth for Rectangular Concrete Soldier Pile")
                    return
                dimensions['width'] = float(width)
                dimensions['depth'] = float(depth)
            elif shape == "Circular":
                diameter = data.get("Diameter")
                if not diameter:
                    print("WARNING: Missing diameter for Circular Concrete Soldier Pile")
                    return
                dimensions['diameter'] = float(diameter)
            else:
                print(f"WARNING: Unknown shape: {shape}")
                return
        else:
            # Contiguous Bored Pile and Secant Bored Pile are always Circular
            shape = "Circular"
            diameter = data.get("Diameter")
            if not diameter:
                print(f"WARNING: Missing diameter for {wall_type}")
                return
            dimensions['diameter'] = float(diameter)
        
        print(f"DEBUG: Processing {wall_type} - Grade: G{grade}, Shape: {shape}, Spacing: {spacing}m")
          
        # Calculate properties using the same function for all concrete piles
        props = self.calculate_concrete_soldier_pile_properties(wall_type, grade, spacing, shape, **dimensions)

        print(f"DEBUG: Calculated {wall_type} Properties - EA={props['EA']:.2e}, EI={props['EI']:.2e}, Unit Weight={props['unit_weight']:.4f} kN/m")
        # Prepare plate properties data with appropriate material name
        if wall_type == "Soldier Pile":
            material_name = f"Concrete_Soldier_Pile_{shape}"
        elif wall_type == "Contiguous Bored Pile":
            material_name = "Contiguous_Bored_Pile"
        elif wall_type == "Secant Bored Pile":
            material_name = "Secant_Bored_Pile"
        else:
            material_name = f"{wall_type.replace(' ', '_')}"
        
        # Use a representative thickness based on the shape
        if shape == "Rectangular":
            representative_d = dimensions['depth']
        else:  # Circular
            representative_d = dimensions['diameter']
        
        plate_data = [
            material_name,  # MaterialName
            "Elastic",  # Elasticity
            True,  # IsIsotropic
            props['EA'],  # EA
            props['EA'],  # EA2 (same as EA)
            props['EI'],  # EI
            representative_d,  # d (representative thickness/diameter)
            props['E'],  # E (Young's modulus in Pa)
            props['EA'] / representative_d if representative_d > 0 else 0,  # Gref
            props['unit_weight'],  # w (weight - not applicable in this context)
            0.2,  # StrutNu (Poisson's ratio for concrete)
            17523200  # Colour
        ]
        
        # Save to database - check if table exists first
        try:
            table_name = sheets_config.get('Plate Properties', {}).get('db_table', 'plate_properties')
            
            cursor.execute(
                f"INSERT INTO {table_name} "
                "(common_id, MaterialName, Elasticity, IsIsotropic, EA, EA2, EI, d, E, Gref, w, StrutNu, Colour) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                [common_id] + plate_data
            )
        except sqlite3.OperationalError as db_error:
            print(f"WARNING: Database table issue: {str(db_error)}")
            print("Skipping database insertion for Plate Properties")
        
        # Save to Excel
        excel_sheets["Plate Properties"].append(plate_data)
        
        # Save to CSV
        import csv
        with open(sheets_config["Plate Properties"]["csv_file"], 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([common_id] + plate_data)
        
        print(f"DEBUG: Saved Concrete Soldier Pile Plate Properties - EA={props['EA']:.2e}, EI={props['EI']:.2e}")
        
      except Exception as ex:
        print(f"ERROR in save_concrete_soldier_pile_properties: {str(ex)}")
        import traceback
        traceback.print_exc()
        raise
  


    def save_diaphragm_wall_properties(self, cursor, data: Dict, common_id: str,
                                   excel_sheets: Dict, sheets_config: Dict) -> None:
      """Save diaphragm wall properties to database, Excel, and CSV."""
      try:
        grade = data.get("Grade")
        thickness = data.get("Thickness")
        
        if not grade or not thickness:
            print("WARNING: Missing grade or thickness for Diaphragm Wall")
            return
        
        # Extract numeric grade value (e.g., "G30" -> 30)
        if isinstance(grade, str) and grade.startswith('G'):
            grade = float(grade[1:])
        else:
            grade = float(grade)
        
        thickness = float(thickness)
        
        print(f"DEBUG: Processing Diaphragm Wall - Grade: G{grade}, Thickness: {thickness}m")
        
        # Calculate properties
        props = self.calculate_diaphragm_wall_properties(grade, thickness)
        print(f"DEBUG: Calculated Diaphragm Wall Properties - EA={props['EA']:.2e}, EI={props['EI']:.2e}, Unit Weight={props['unit_weight']:.4f} kN/m")
        # Prepare plate properties data
        material_name = "Diaphragm_Wall"
        
        plate_data = [
            material_name,  # MaterialName
            "Elastic",  # Elasticity
            True,  # IsIsotropic
            props['EA'],  # EA
            props['EA'],  # EA2 (same as EA)
            props['EI'],  # EI
            props['thickness'],  # d (thickness)
            props['E'],  # E (Young's modulus in Pa)
            props['EA'] / props['thickness'] if props['thickness'] > 0 else 0,  # Gref
            props['unit_weight'],  # w (weight - not applicable in this context)
            0.2,  # StrutNu (Poisson's ratio for concrete)
            17523200  # Colour
        ]
        
        # Save to database - check if table exists first
        try:
            table_name = sheets_config.get('Plate Properties', {}).get('db_table', 'plate_properties')
            
            cursor.execute(
                f"INSERT INTO {table_name} "
                "(common_id, MaterialName, Elasticity, IsIsotropic, EA, EA2, EI, d, E, Gref, w, StrutNu, Colour) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                [common_id] + plate_data
            )
        except sqlite3.OperationalError as db_error:
            print(f"WARNING: Database table issue: {str(db_error)}")
            print("Skipping database insertion for Plate Properties")
        
        # Save to Excel
        excel_sheets["Plate Properties"].append(plate_data)
        
        # Save to CSV
        import csv
        with open(sheets_config["Plate Properties"]["csv_file"], 'a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([common_id] + plate_data)
        
        print(f"DEBUG: Saved Diaphragm Wall Plate Properties - EA={props['EA']:.2e}, EI={props['EI']:.2e}")
        
      except Exception as ex:
        print(f"ERROR in save_diaphragm_wall_properties: {str(ex)}")
        import traceback
        traceback.print_exc()
        raise
    
    def calculate_diaphragm_wall_properties(self, grade: float, thickness: float) -> Dict:

      try:
        grade_str = f"G{int(grade)}" if not str(grade).startswith('G') else str(grade)
        grade_props = self.get_concrete_grade_properties(grade_str)
        
        Ecm = grade_props['Ecm']
        E_Pa = Ecm * 1e9  # Convert to Pa
        density = grade_props['density']
        poissons_ratio = grade_props['poissons_ratio']
        
        print(f"DEBUG: Diaphragm Wall - Using Ecm from database: {Ecm} GPa")
        
        # Calculate A and I for diaphragm wall (per meter width)
        A = thickness * 1.0  # m² (thickness × 1m width)
        I = math.pow(thickness, 3) / 12  # m⁴ (per meter width)
        
        # Calculate EA and EI
        EA = E_Pa * A
        EI = E_Pa * I
        
        # Calculate unit weight: w = (density - 20) * A
        unit_weight = self.calculate_unit_weight(
            "Diaphragm Wall",
            density,
            A=A
        )
        
        print(f"DEBUG: Diaphragm Wall Properties:")
        print(f"  Thickness: {thickness}m")
        print(f"  A = {A:.6f} m^2")
        print(f"  I = {I:.6e} m^4")
        print(f"  EA = {EA:.2e} N/m")
        print(f"  EI = {EI:.2e} Nm^2/m")
        print(f"  Unit Weight = {unit_weight:.4f} kN/m")
        
        return {
            'EA': EA,
            'EI': EI,
            'E': E_Pa,
            'A': A,
            'I': I,
            'thickness': thickness,
            'density': density,
            'poissons_ratio': poissons_ratio,
            'unit_weight': unit_weight
        }
        
      except Exception as ex:
        print(f"ERROR in calculate_diaphragm_wall_properties: {str(ex)}")
        import traceback
        traceback.print_exc()
        raise
    def save_wall_details(self, cursor, data: Dict, common_id: str, 
                     excel_sheets: Dict, sheets_config: Dict) -> None:
      """Save wall details to database, Excel, and CSV files."""
      try:
        wall_type = data.get("Wall Type")
        excavation_type = data.get("Excavation Type", "Single Wall")
        excavation_width = float(data.get("Excavation Width", 0))
        wall_top_level = float(data.get("Wall Top Level", 0))
        toe_level = float(data.get("Toe Level", 0))
        
        # Save ERSS Wall Details
        self._save_erss_wall_details(
            cursor, wall_type, excavation_type, excavation_width,
            wall_top_level, toe_level, common_id, excel_sheets, sheets_config
        )
        
        # Save Plate Properties based on wall type
        if wall_type == "Steel Pipe":
            self.save_steel_pipe_properties(
                cursor, data, common_id, excel_sheets, sheets_config
            )
        elif wall_type == "Sheet Pile":
            self.save_sheet_pile_properties(
                cursor, data, common_id, excel_sheets, sheets_config
            )
        elif wall_type == "Soldier Pile":
            material = data.get("Material")
            
            if material == "Concrete":
                # Handle concrete soldier pile
                self.save_concrete_soldier_pile_properties(
                    cursor, data, common_id, excel_sheets, sheets_config
                )
            elif material == "Steel":
                # Handle steel soldier pile
                member_size = data.get("Member Size") or "NO"
                if member_size and member_size != "NO":
                    if not member_size.startswith("UB"):
                        member_size = f"UB{member_size}"
                    self._save_plate_properties(
                        cursor, wall_type, member_size, common_id, 
                        excel_sheets, sheets_config
                    )
        elif wall_type == "Contiguous Bored Pile":
            # Use the same calculation as concrete soldier pile with circular shape
            self.save_concrete_soldier_pile_properties(
                cursor, data, common_id, excel_sheets, sheets_config
            )
        elif wall_type == "Secant Bored Pile":
            # Use the same calculation as concrete soldier pile with circular shape
            self.save_concrete_soldier_pile_properties(
                cursor, data, common_id, excel_sheets, sheets_config
            )
        elif wall_type == "Diaphragm Wall":
            # Handle diaphragm wall (continuous wall per meter)
            self.save_diaphragm_wall_properties(
                cursor, data, common_id, excel_sheets, sheets_config
            )
        
        print(f"DEBUG: Successfully saved wall details for {wall_type}")
        
      except Exception as ex:
        print(f"ERROR in save_wall_details: {str(ex)}")
        import traceback
        traceback.print_exc()
        raise
    
    def _save_erss_wall_details(self, cursor, wall_type: str, excavation_type: str,
                                excavation_width: float, wall_top_level: float,
                                toe_level: float, common_id: str, 
                                excel_sheets: Dict, sheets_config: Dict) -> None:
        """Save ERSS wall details for left and optionally right wall."""
        
        def save_wall_data(wall_name: str, x_coord: float):
            wall_data = [
                wall_type, wall_name,
                x_coord, wall_top_level,
                x_coord, toe_level
            ]
            print(f"DEBUG: Saving wall data: {wall_data}")
            
            # Save to database
            cursor.execute(
                f"INSERT INTO {sheets_config['ERSS Wall Detail']['db_table']} "
                "(common_id, MaterialName, WallName, x_Top, y_Top, x_Bottom, y_Bottom) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                [common_id] + wall_data
            )
            
            # Save to Excel
            excel_sheets["ERSS Wall Detail"].append(wall_data)
            
            # Save to CSV
            import csv
            with open(sheets_config["ERSS Wall Detail"]["csv_file"], 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([common_id] + wall_data)
        
        # Save left wall
        save_wall_data("Wall_Left", -(excavation_width / 2))
        
        # Save right wall if double wall excavation
        if excavation_type.lower() == "double wall":
            save_wall_data("Wall_Right", (excavation_width / 2))

    def _save_plate_properties(self, cursor, wall_type: str, member_size: str,
                               common_id: str, excel_sheets: Dict, 
                               sheets_config: Dict) -> None:
        """Save plate properties for steel walls."""
        # Ensure member_size has UB prefix for lookup
        lookup_member_size = member_size
        if not member_size.startswith("UB"):
            lookup_member_size = f"UB{member_size}"
        
        # Get comprehensive properties
        plate_props = self.geometry_section.get_comprehensive_plate_properties(lookup_member_size)
        
        if plate_props:
            # Prepare plate properties data
            plate_data = [
                wall_type,  # MaterialName (use Wall Type as material name)
                plate_props["Elasticity"],
                plate_props["IsIsotropic"],
                plate_props["EA"],
                plate_props["EA2"],
                plate_props["EI"],
                plate_props["d"],
                plate_props["E"],
                plate_props["Gref"],
                plate_props["w"],
                plate_props["StrutNu"],
                plate_props["Colour"]
            ]
            
            # Save to Excel
            excel_sheets["Plate Properties"].append(plate_data)
            
            # Save to CSV
            import csv
            with open(sheets_config["Plate Properties"]["csv_file"], 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([common_id] + plate_data)
            
            print(f"DEBUG: Saved Plate Properties for {wall_type} with member size {member_size}")
        else:
            print(f"WARNING: Could not retrieve properties for member size {member_size}")