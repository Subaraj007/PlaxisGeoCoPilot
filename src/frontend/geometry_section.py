# Standard Library 
import csv
import os
import logging
from pathlib import Path
from typing import List, Dict, Optional, Union, Tuple
import sys
import math

# Third-Party Library 
from openpyxl.styles import Alignment
import flet as ft
import mysql.connector
import openpyxl
from openpyxl.utils import get_column_letter

# Local Module 
from frontend.database_config import DatabaseConfig
from frontend.database_connection import DatabaseConnection
from frontend.form_section import FormSection, FormField
from frontend.utilities import resource_path
from frontend.wall_details_handler import WallDetailsHandler
from frontend.lineload_details_handler import LineLoadHandler

class GeometrySection(FormSection):
    """Manages geometry section of excavation form including wall and strut configurations."""

    def __init__(self, db_config: DatabaseConfig):
        self.db_config = db_config
        self.current_num_struts = 0
        self.parent_form = None  
        self.form_values = {}  # Dictionary to store form values
        self.validation_states = {}
        self.field_errors = {}  # Store field-specific validation errors
        self.form_content = None  # Initialize form_content
        self.page = None  # Initialize page reference
        # Check if running as executable
        if getattr(sys, 'frozen', False):
            # Running as exe - use internal/data directory
            self.BASE_DIR = Path(sys.executable).parent / "_internal"
        else:
            # Running as script - use original path
            self.BASE_DIR = Path(__file__).resolve().parent.parent.parent

        self.export_dir = self.BASE_DIR / "data"
        self.steel_member_property_path = resource_path("data/Steel_member_properties.xlsx")
        self.input_data_path = resource_path("data/Input_Data.xlsx")
        self.wall_details_handler = WallDetailsHandler(self)    
        self.line_load_handler = LineLoadHandler(self)      
        # Add debug output
        print(f"Resource path: {self.steel_member_property_path}")
        print(f"File exists: {os.path.exists(self.steel_member_property_path)}")
    def import_from_dict(self, data: dict):
      """Import geometry data directly from dictionary"""
      self.form_values.update(data)
    
    # Process critical values first
      if 'No of Strut' in data:
        self.current_num_struts = int(data['No of Strut'])
    
    # Force UI regeneration
      self.form_content = self.build_section_ui()
    
    # Now populate all fields
      self._populate_all_sections(data)
      
    def build_section_ui(self, initial_data: dict = None):
      """Build the geometry UI with optional initial data"""
      if initial_data:
        self.form_values.update(initial_data)
        
      fields = self.get_fields()
      return self._create_geometry_frames(fields, self.form_values)
    def safe_float(self,value):
      try:
        return float(value)
      except (ValueError, TypeError):
        return None  # or some default value

    def validate_field(self, field_name: str, value: str, all_values: dict) -> str:
      """Validate a single field and return error message if invalid"""
      try:
        if not value and field_name in ["Spacing"]:
            print("DEBUG: Spacing is required")
            return "" 
        if not value and field_name in ["Wall Top Level", "Excavation Depth", 
                                      "Toe Level", "Excavation Width"]:
            return f"{field_name} is required"

        # Convert to float for numeric validations
        if field_name in ["Wall Top Level", "Excavation Depth", "Toe Level", 
                         "Excavation Width", "Ground Water Table"]:
            try:
                value = float(value)
            except ValueError:
                return f"{field_name} must be a valid number"

        # Validation 1: Toe Level <= FEL (Formation/Excavation Level)
        if field_name == "Toe Level":
           wall_top = float(all_values["Wall Top Level"])
           exc_depth = float(all_values["Excavation Depth"])
           fel = wall_top - exc_depth  # Calculate FEL
           if float(value) > fel:  # Toe Level must be <= FEL
              return f"Toe Level must be less than or equal to FEL ({fel})"
        elif field_name == "Stiffness":
           if not isinstance(value, bool):
            return "Invalid stiffness value"
        # Strut Level validations
        elif field_name.startswith("Strut") and "Level" in field_name:
           wall_top = float(all_values["Wall Top Level"])
           exc_depth = float(all_values["Excavation Depth"])
           fel = wall_top - exc_depth  # Calculate FEL
           
           # Get current strut number and level
           strut_num = int(field_name.split()[1])
           current_strut_level = float(value)

           # Validation for Strut 1 
           if strut_num == 1:
               # Strut 1 must be > FEL and < Wall Top Level
               if current_strut_level <= fel:
                   return f"Strut 1 Level must be greater than FEL ({fel})"
               if current_strut_level >= wall_top:
                   return f"Strut 1 Level must be less than Wall Top Level ({wall_top})"

           # Validations for subsequent struts
           else:
               # Check previous strut's level key
               prev_strut_key = f"Strut {strut_num-1} Level"
               
               # Ensure previous strut level exists before comparison
               if prev_strut_key not in all_values:
                   return ""  # Can't validate if previous strut level missing

               prev_level = float(all_values[prev_strut_key])

               # Current strut must be strictly lower than previous strut
               if current_strut_level >= prev_level:
                   return f"Strut {strut_num} Level must be less than Strut {strut_num-1} Level ({prev_level})"

               # Current strut must be above FEL
               if current_strut_level <= fel:
                   return f"Strut {strut_num} Level must be greater than FEL ({fel})"
        elif field_name.startswith("Strut") and "Space" in field_name:
          try:
            space_value = float(value)
            if space_value <= 0:
                return f"{field_name} must be greater than 0"
          except (ValueError, TypeError):
            return f"{field_name} must be a valid number"
    

        # Validation 6: Borehole X coordinate must be within model width
        elif field_name == "X":
            exc_width = float(all_values["Excavation Width"])
            wall_top = float(all_values["Wall Top Level"])
            toe_level = float(all_values["Toe Level"])
            model_width = exc_width/2 + 4 * (wall_top - toe_level)  # Calculate model width
            if abs(float(value)) >= model_width:
                 return f"X must be less than Model width ({model_width})"

# Validation 7: Ground Water Table must be below Wall Top Level
        elif field_name == "Ground Water Table":
           wall_top = float(all_values["Wall Top Level"])
           if float(value) >wall_top:
               return f"Ground Water Table must be greater than Wall Top Level ({wall_top})"
        elif all_values.get("Strut Type") == "Fixed":
            print("DEBUG: Running Fixed Strut validations")
            if field_name == "Strut Space":
                try:
                    strut_space = float(value)
                    if strut_space < 1.0:
                        return "Strut spacing must be ≥ 1.0"
                except ValueError:
                    return "Invalid strut spacing value"
            
            elif field_name == "Struct Length":
                try:
                    strut_length = float(value)
                    excavation_width = float(all_values.get("Excavation Width", 0))
                    
                    if strut_length < 1.0 or strut_length > excavation_width:
                        return f"Strut length must be between 1.0 and {excavation_width}"
                except (ValueError, TypeError):
                    return "Invalid strut length value"
            
            elif field_name == "Angle":
                try:
                    angle = float(value)
                    if angle <= -75 or angle >= 75:
                        return "Angle must be between -75° and 75°"
                except ValueError:
                    return "Invalid angle value"
        return ""  
      except Exception as e:
        return str(e)
    def get_plate_properties(self, cursor, material_name: str) -> Dict[str, any]:
      """Retrieves plate properties from Excel sheet for given material."""
      try:
        # Process material name - remove "UC" prefix if present
        search_material = material_name
        if material_name.startswith("UC"):
            search_material = material_name[2:]  # Remove first two characters "UC"
            print(f"Material starts with 'UC', searching for: {search_material}")
        else:
            print(f"Attempting to get properties for material: {material_name}")
            
        print(f"Opening file: {self.steel_member_property_path}")
        
        # Load the workbook
        workbook = openpyxl.load_workbook(self.steel_member_property_path)
        print(f"Successfully opened workbook, sheets available: {workbook.sheetnames}")
        
        # Select Sheet1
        section_sheet = workbook['Sheet1']
        print(f"Accessed 'Sheet1', searching for material starting from row 7")
        
        # Iterate through rows starting from the 7th row to find the material
        material_found = False
        row_count = 0
        
        for row in section_sheet.iter_rows(min_row=7, values_only=True):
            row_count += 1
            if row_count % 50 == 0:  # Log progress every 50 rows
                print(f"Processed {row_count} rows, still searching...")
                
            if row[0] == search_material:  # Column A contains material_name
                print(f"Found material '{search_material}' at row {row_count+6}")
                material_found = True
                
                # Get value from Column W as "A"
                A = row[22]  # Assuming W is the 23rd column (0-indexed)
                print(f"Retrieved A value from column W: {A}")
                
                # For EA: E * A where E is always 2000
                E = 21000
                EA = E * A
                print(f"Calculated EA = {E} * {A} = {EA}")
                
                # Return properties
                result = {
                    "Elasticity": "Elastic",  # Always "Elastic"
                    "EA": EA,
                    "Colour": "17523200"  # Always "17523200"
                }
                print(f"Returning properties: {result}")
                return result
                
        # If we reach here, material was not found
        if not material_found:
            print(f"Warning: Material '{search_material}' not found after searching {row_count} rows")
            
        # If material not found, return default values
        print("Returning default (None) values for properties")
        return {"Elasticity": None, "EA": None, "Colour": None}
        
      except FileNotFoundError:
        print(f"Error: File not found at path: {self.steel_member_property_path}")
        return {"Elasticity": None, "EA": None, "Colour": None}
      except KeyError as e:
        print(f"Error: Sheet not found: {e}")
        return {"Elasticity": None, "EA": None, "Colour": None}
      except Exception as e:
        print(f"Unexpected error retrieving plate properties: {e}")
        print(f"Error type: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        return {"Elasticity": None, "EA": None, "Colour": None}
    def get_comprehensive_plate_properties(self, material_name: str) -> Dict[str, any]:
      """Retrieves comprehensive plate properties including all required columns."""
      try:
        # Process material name - remove "UC" or "UB" prefix if present
        search_material = material_name
        if material_name.startswith("UC") or material_name.startswith("UB"):
            search_material = material_name[2:]
            
        # Load workbook
        workbook = openpyxl.load_workbook(self.steel_member_property_path)
        section_sheet = workbook['Sheet1']
        
        # Search for material starting from row 7
        for row in section_sheet.iter_rows(min_row=7, values_only=True):
            if row[0] == search_material:  # Column A (index 0)
                # Extract values from the row
                w = row[1]      # Column B (index 1) - Mass per metre
                d = row[7]      # Column H (index 7) - Depth of section
                Ix = row[10]    # Column K (index 10) - Second Moment of Area x-x (cm⁴)
                A = row[22]     # Column W (index 22) - Area (cm²)
                
                # Material constant for steel
                E = 210000  # Young's modulus in N/mm²
                
                # Calculate derived values
                EA = E * A * 100  # Convert A from cm² to mm² (multiply by 100)
                EA2 = EA / 100
                
                # Calculate EI: E × Ix
                # Convert Ix from cm⁴ to mm⁴ (multiply by 10,000)
                EI = E * Ix * 10000  # Result in N·mm²
                
                # Calculate Gref (shear stiffness reference)
                Gref = EA / d if d and d != 0 else 0
                
                return {
                    "Elasticity": "Elastic",
                    "IsIsotropic": True,
                    "EA": EA,
                    "EA2": EA2,
                    "EI": EI,
                    "d": d,
                    "E": E,  # Now properly populated
                    "Gref": Gref,
                    "w": w/100,
                    "StrutNu": 0.2,  # Poisson's ratio for steel
                    "Colour": 17523200
                }
                
        # Material not found
        print(f"Warning: Material '{search_material}' not found")
        return None
        
      except Exception as e:
        print(f"Error retrieving comprehensive plate properties: {e}")
        import traceback
        traceback.print_exc()
        return None
    def get_dependent_fields(self, field_name: str) -> List[str]:
        """Return list of field names that depend on the given field"""
        dependencies = {
            "Wall Top Level": ["Ground Water Table", "Toe Level"] + 
                            [f"Strut {i} Level" for i in range(1, self.current_num_struts + 1)],
            "Excavation Depth": ["Toe Level"] + 
                               [f"Strut {i} Level" for i in range(1, self.current_num_struts + 1)],
            "Strut Type": ["Strut Space","Struct Length","Angle"]
        }
        return dependencies.get(field_name, [])

    def find_field_control(self, field_name: str) -> Optional[ft.Control]:
      """Find the control for a given field name in the form"""
      if not self.parent_form:
        return None
    
    # Check if field_name exists in the form
      for container in self.parent_form.form_content.controls:
        # Safely check if the container has a content attribute and is a Column
        if (hasattr(container, 'content') and 
            isinstance(container.content, ft.Column)):
            
            # Iterate through rows in the column
            for row in container.content.controls:
                if isinstance(row, ft.Row):
                    for field in row.controls:
                        # Check for TextField or Dropdown
                        if (isinstance(field, (ft.TextField, ft.Dropdown)) and 
                            field.label and 
                            field_name in field.label):
                            return field
      return None
          
    def set_parent(self, parent_form):
        """Set the parent form reference and ensure form_content is accessible"""
        self.parent_form = parent_form
        if hasattr(parent_form, 'form_content'):
            self.form_content = parent_form.form_content
        if hasattr(parent_form, 'page'):
            self.page = parent_form.page
        
        return self
    
    def store_current_values(self, controls):
      """Store current values from all form controls, including nested ones."""
      self.form_values.clear()
    
      def traverse(control):
        if isinstance(control, ft.Row):
            for child in control.controls:
                traverse(child)
        elif isinstance(control, ft.Column):
            for child in control.controls:
                traverse(child)
        elif isinstance(control, ft.Container) and hasattr(control, 'content'):
            traverse(control.content)
        elif isinstance(control, (ft.TextField, ft.Dropdown, ft.Checkbox)):
            if control.label:
                # Clean the label to remove any annotations
                clean_label = control.label.split(" (Set")[0].split(" *")[0]
                # Store non-empty values, preserving existing if not empty
                current_value = control.value.strip() if isinstance(control.value, str) else control.value
                if current_value not in (None, "", "None"):
                    self.form_values[clean_label] = current_value
                    print(f"DEBUG: Stored {clean_label} = {current_value}")  # Debug log
    
    # Start traversal from the provided controls
      for control in controls:
        traverse(control)
                        
    def load_section_details(self) -> List[str]:
      """Load section designations from the Excel sheet"""
      try:
        # Load the workbook
        workbook = openpyxl.load_workbook(self.steel_member_property_path)
        
        # Select the Section Details sheet (assuming it's named "Section Details")
        section_sheet = workbook['Sheet1']
        
        # Extract section designations from the first column
        section_designations = []
        
        # Iterate through rows starting from the second row (assuming first row is headers)
        for row in section_sheet.iter_rows(min_row=7, max_col=1, values_only=True):
            if row[0]:  # Check if the designation cell is not empty
                section_designations.append(str(row[0]))
        
        # Sort the section designations
        section_designations.sort()
        
        return section_designations
    
      except FileNotFoundError:
        print(f"Error: Input data file not found at {self.input_data_path}")
        return []
      except KeyError:
        print("Error: 'Section Details' sheet not found in the workbook")
        return []
      except Exception as e:
        print(f"Error loading section details from Excel: {e}")
        return []

    
    def get_fields(self) -> List[FormField]:
      """Generate the form fields based on current state and inputs."""
      # Load section details
      self.section_details = self.load_section_details()

    # Define wall types directly
      wall_types = ["Soldier Pile", "Sheet Pile", "RC Wall"]
    
      current_wall_type = self.form_values.get("Wall Type")
      current_strut_type = self.form_values.get("Strut Type")
      current_excavation_type = self.form_values.get("Excavation Type")

    # Base fields that are always present
      fields = [
        FormField("Excavation Type", "dropdown", 
                 options=["Single wall", "Double wall", "Cassion"],
                 value=self.form_values.get("Excavation Type")).set_on_change(self.handle_excavation_type_change),
        FormField("Wall Top Level", "number", "e.g: 5",
                 value=self.form_values.get("Wall Top Level")),
        FormField("Excavation Depth", "number", "e.g: 10",
                 value=self.form_values.get("Excavation Depth")),
        FormField("Excavation Width", "number", "e.g: 15",
                 value=self.form_values.get("Excavation Width")),
        FormField("Toe Level", "number", "e.g: -15",
                 value=self.form_values.get("Toe Level")),
        FormField("No of Strut", "number", "e.g: 2",
                 value=str(self.current_num_struts) if self.current_num_struts else None
                 ).set_on_change(self.handle_strut_change),
    ]
    
    # MODIFIED: Determine Strut Type options based on Excavation Type
      if current_excavation_type == "Single wall":
        strut_type_options = ["Fixed"]  # Only Fixed for Single wall
        if not current_strut_type:
            current_strut_type = "Fixed"
            self.form_values["Strut Type"] = "Fixed"
      elif current_excavation_type == "Double wall":
        strut_type_options = ["Node to Node"]  # Only Node to Node for Double wall
        if not current_strut_type:
            current_strut_type = "Node to Node"
            self.form_values["Strut Type"] = "Node to Node"
      else:
        # For Cassion or when no excavation type is selected, show both options
        strut_type_options = ["Node to Node", "Fixed"]
    
    # Add Strut Type field with filtered options
      fields.append(
        FormField("Strut Type", "dropdown", 
                 options=strut_type_options,
                 value=current_strut_type)
    )
    
    # Continue with remaining fields
      fields.extend([
        FormField("Excavation Below Strut", "number", "e.g: 1.5",
                 value=self.form_values.get("Excavation Below Strut"), required=False),
        FormField("Over Excavation", "number", "e.g: 0.5",
                 value=self.form_values.get("Over Excavation")),
      ])
    
    # Add wall type field
      wall_type_field = self.wall_details_handler.get_wall_type_field(current_wall_type)
      wall_type_field.set_on_change(self.handle_wall_type_change)
      fields.append(wall_type_field)
    
    # Add wall-specific fields based on wall type
      if current_wall_type:
        wall_specific_fields = self.wall_details_handler.get_fields_for_wall_type(
          current_wall_type, 
          self.form_values
        )
      # Ensure wall_specific_fields are added in the correct order
        for field in wall_specific_fields:
          # Add change handlers for the new grade fields
          if field.label == "Steel Grade":
            field.set_on_change(self.handle_field_change)
          elif field.label == "Sheet Grade":
            field.set_on_change(self.handle_field_change)
    
        fields.extend(wall_specific_fields)
        
    # Add borehole fields
      fields.extend([
        FormField("Borehole X Coordinate", "number", "e.g: 0",
                 value=self.form_values.get("Borehole X Coordinate")),
        FormField("Ground Water Table", "number", "e.g: -2",
                 value=self.form_values.get("Ground Water Table"))
    ])
    
      # Set change handlers for all fields
      for field in fields:
        if field.label == "Wall Type":
            field.set_on_change(self.handle_wall_type_change)
        elif field.label == "No of Strut":
            continue  # Already has handler
        elif field.label == "Strut Type":
            field.set_on_change(self.handle_strut_type_change)
        elif field.label == "Excavation Type":
            continue  # Already has handler
        else:
            field.set_on_change(self.handle_field_change)
    
      return fields
   
    def validate_field(self, field_name: str, value: str, all_values: dict) -> str:
      """Validate a single field and return error message if invalid"""
      try:
        if not value and field_name in ["Spacing"]:
            print("DEBUG: Spacing is required")
            return "" 
        if not value and field_name in ["Wall Top Level", "Excavation Depth", 
                                      "Toe Level", "Excavation Width"]:
            return f"{field_name} is required"

        # Convert to float for numeric validations
        if field_name in ["Wall Top Level", "Excavation Depth", "Toe Level", 
                         "Excavation Width", "Ground Water Table"]:
            try:
                value = float(value)
            except ValueError:
                return f"{field_name} must be a valid number"

        # Validation 1: Toe Level <= FEL (Formation/Excavation Level)
        if field_name == "Toe Level":
           wall_top = float(all_values["Wall Top Level"])
           exc_depth = float(all_values["Excavation Depth"])
           fel = wall_top - exc_depth  # Calculate FEL
           if float(value) > fel:  # Toe Level must be <= FEL
              return f"Toe Level must be less than or equal to FEL ({fel})"
        elif field_name == "Stiffness":
           if not isinstance(value, bool):
            return "Invalid stiffness value"
        # Strut Level validations
        elif field_name.startswith("Strut") and "Level" in field_name:
           wall_top = float(all_values["Wall Top Level"])
           exc_depth = float(all_values["Excavation Depth"])
           fel = wall_top - exc_depth  # Calculate FEL
           
           # Get current strut number and level
           strut_num = int(field_name.split()[1])
           current_strut_level = float(value)

           # Validation for Strut 1 
           if strut_num == 1:
               # Strut 1 must be > FEL and < Wall Top Level
               if current_strut_level <= fel:
                   return f"Strut 1 Level must be greater than FEL ({fel})"
               if current_strut_level >= wall_top:
                   return f"Strut 1 Level must be less than Wall Top Level ({wall_top})"

           # Validations for subsequent struts
           else:
               # Check previous strut's level key
               prev_strut_key = f"Strut {strut_num-1} Level"
               
               # Ensure previous strut level exists before comparison
               if prev_strut_key not in all_values:
                   return ""  # Can't validate if previous strut level missing

               prev_level = float(all_values[prev_strut_key])

               # Current strut must be strictly lower than previous strut
               if current_strut_level >= prev_level:
                   return f"Strut {strut_num} Level must be less than Strut {strut_num-1} Level ({prev_level})"

               # Current strut must be above FEL
               if current_strut_level <= fel:
                   return f"Strut {strut_num} Level must be greater than FEL ({fel})"
        
        # NEW: Strut Space validation for both strut types
        elif field_name.startswith("Strut") and "Space" in field_name:
            if not value:
                strut_type = all_values.get("Strut Type")
                if strut_type in ["Node to Node", "Fixed"]:
                    return f"{field_name} is required for {strut_type} struts"
            try:
                space_value = float(value)
                if space_value <= 0:
                    return f"{field_name} must be greater than 0"
            except (ValueError, TypeError):
                return f"{field_name} must be a valid number"
        
        # NEW: Strut Length validation for Fixed struts
        elif field_name.startswith("Strut") and "Length" in field_name:
            strut_type = all_values.get("Strut Type")
            if strut_type == "Fixed":
                if not value:
                    return f"{field_name} is required for Fixed struts"
                try:
                    strut_length = float(value)
                    excavation_width = float(all_values.get("Excavation Width", 0))
                    
                    if strut_length < 1.0 or strut_length > excavation_width:
                        return f"Strut length must be between 1.0 and {excavation_width}"
                except (ValueError, TypeError):
                    return "Invalid strut length value"
        
        # NEW: Strut Angle validation for Fixed struts
        elif field_name.startswith("Strut") and "Angle" in field_name:
            strut_type = all_values.get("Strut Type")
            if strut_type == "Fixed":
                if not value:
                    return f"{field_name} is required for Fixed struts"
                try:
                    angle = float(value)
                    if angle <= -75 or angle >= 75:
                        return "Angle must be between -75° and 75°"
                except ValueError:
                    return "Invalid angle value"

        # Validation 6: Borehole X coordinate must be within model width
        elif field_name == "X":
            exc_width = float(all_values["Excavation Width"])
            wall_top = float(all_values["Wall Top Level"])
            toe_level = float(all_values["Toe Level"])
            model_width = exc_width/2 + 4 * (wall_top - toe_level)  # Calculate model width
            if abs(float(value)) >= model_width:
                 return f"X must be less than Model width ({model_width})"

        # Validation 7: Ground Water Table must be below Wall Top Level
        elif field_name == "Ground Water Table":
           wall_top = float(all_values["Wall Top Level"])
           if float(value) > wall_top:
               return f"Ground Water Table must be greater than Wall Top Level ({wall_top})"
        
        return ""  
      except Exception as e:
        return str(e)    
    def validate(self, data: dict) -> List[str]:
      """Final validation before form submission including strut-specific validations"""
      errors = list(self.field_errors.values())  # Return any current validation errors
    
      # Get number of struts
      num_struts = 0
      try:
        num_struts = int(data.get("No of Strut", 0))
      except (ValueError, TypeError):
        errors.append("Invalid number of struts")
        return errors
      
      strut_type = data.get("Strut Type")
    
    # Validate strut type is selected when struts exist
      if num_struts > 0 and not strut_type:
        errors.append("Strut Type is required when struts are defined")
        return errors
    
    # Validate each strut based on strut type
      for i in range(1, num_struts + 1):
        # Common validations for all strut types
        if not data.get(f"Strut {i} Level"):
            errors.append(f"Strut {i} Level is required")
        
        if not data.get(f"Strut {i} Material"):
            errors.append(f"Strut {i} Material is required")
        
        # Space is required for both strut types
        if not data.get(f"Strut {i} Space"):
            errors.append(f"Strut {i} Space is required for {strut_type} struts")
        else:
            try:
                space_value = float(data.get(f"Strut {i} Space"))
                if space_value <= 0:
                    errors.append(f"Strut {i} Space must be greater than 0")
            except (ValueError, TypeError):
                errors.append(f"Strut {i} Space must be a valid number")
        
        # Material-specific validations
        material = data.get(f"Strut {i} Material")
        if material == "Steel":
            if not data.get(f"Strut {i} Member Size"):
                errors.append(f"Strut {i} Member Size is required for Steel struts")
        elif material == "Concrete":
            if not data.get(f"Strut {i} EA"):
                errors.append(f"Strut {i} EA is required for Concrete struts")
            if not data.get(f"Strut {i} Stiffness"):
                errors.append(f"Strut {i} Stiffness is required for Concrete struts")
        
        # Fixed strut specific validations
        if strut_type == "Fixed":
            # Length validation
            if not data.get(f"Strut {i} Length"):
                errors.append(f"Strut {i} Length is required for Fixed struts")
            else:
                try:
                    strut_length = float(data.get(f"Strut {i} Length"))
                    excavation_width = float(data.get("Excavation Width", 0))
                    
                    if strut_length < 1.0 or strut_length > excavation_width:
                        errors.append(f"Strut {i} Length must be between 1.0 and {excavation_width}")
                except (ValueError, TypeError):
                    errors.append(f"Strut {i} Length must be a valid number")
            
            # Angle validation
            if not data.get(f"Strut {i} Angle"):
                errors.append(f"Strut {i} Angle is required for Fixed struts")
            else:
                try:
                    angle = float(data.get(f"Strut {i} Angle"))
                    if angle <= -75 or angle >= 75:
                        errors.append(f"Strut {i} Angle must be between -75° and 75°")
                except (ValueError, TypeError):
                    errors.append(f"Strut {i} Angle must be a valid number")
    
      return errors
    
    def _get_strut_fields(self, strut_num):
      strut_fields = [
        FormField(f"Strut {strut_num} Level", "number", 
                 value=self.form_values.get(f"Strut {strut_num} Level")),
        FormField(f"Strut {strut_num} Material", "dropdown",
                 options=["Steel", "Concrete"],
                 value=self.form_values.get(f"Strut {strut_num} Material")),
        FormField(f"Strut {strut_num} Space", "number",
                 value=self.form_values.get(f"Strut {strut_num} Space"))
    ]
      current_strut_type = self.form_values.get("Strut Type")
      if current_strut_type == "Fixed":
        strut_fields.extend([
            FormField(f"Strut {strut_num} Length", "number",
                     value=self.form_values.get(f"Strut {strut_num} Length")),
            FormField(f"Strut {strut_num} Angle", "number",
                     value=self.form_values.get(f"Strut {strut_num} Angle"))
        ])
      for field in strut_fields:
        field.set_on_change(self.handle_field_change)
      material = self.form_values.get(f"Strut {strut_num} Material")
    
      if material == "Steel":
        # Create member size field with proper change handler
        member_size_field = FormField(
            f"Strut {strut_num} Member Size",
            "dropdown",
            options=self.section_details,
            value=self.form_values.get(f"Strut {strut_num} Member Size")
        )
        # Add change handler correctly
        member_size_field.set_on_change(lambda e: self._update_member_size(e, strut_num))
        strut_fields.append(member_size_field)
        
      elif material == "Concrete":
        strut_fields.extend([
            FormField(f"Strut {strut_num} EA", "number",
                     value=self.form_values.get(f"Strut {strut_num} EA")),
            FormField(f"Strut {strut_num} Stiffness", "number",
                     value=self.form_values.get(f"Strut {strut_num} Stiffness"))
        ])

      return strut_fields
    def _update_member_size(self, e, strut_num):
      """Update member size for a specific strut"""
      key = f"Strut {strut_num} Member Size"
      self.form_values[key] = e.control.value
      print(f"DEBUG: Updated {key} to {e.control.value}")
    
    # Use correct page reference
      if self.parent_form and self.parent_form.page:
        self.parent_form.page.update()
      elif hasattr(self, 'page'):
        self.page.update()
      else:
        print("DEBUG: No page reference found for update")
    def _update_strut_value(self, e, strut_num, label):
      clean_label = label.replace(f" {strut_num}", "")  # "Strut Material" -> "Strut"
      key = f"Strut {strut_num} {clean_label}"
      self.form_values[key] = e.control.value
    def handle_strut_material_change(self, e):
      """Handle change in strut material and update member size options."""
      print(f"DEBUG: Entering handle_strut_material_change")
    
    # Ensure section details are loaded
      self.section_details = self.load_section_details()
      print(f"DEBUG: Section Details loaded: {len(self.section_details) if self.section_details else 0} items")
    
      try:
        strut_material = e.control.value if hasattr(e.control, 'value') else e.data
        strut_label = e.control.label if hasattr(e.control, 'label') else None
        
        # Extract strut number from label or infer from context
        if strut_label and "Strut" in strut_label and "Material" in strut_label:
            try:
                strut_num = int(strut_label.split()[1])
            except (ValueError, IndexError):
                print(f"DEBUG: Could not extract strut number from label: {strut_label}")
                return
        else:
            # Try to infer from data
            print(f"DEBUG: No clear strut label, inferring from context")
            return
            
        print(f"DEBUG: Strut {strut_num} Material changed to {strut_material}")
        
        # Update form values
        self.form_values[strut_label] = strut_material
        print(f"DEBUG: Updated form_values: {self.form_values.get(strut_label)}")
        
        # Find the sub-frame containing this strut
        print(f"DEBUG: Searching for strut {strut_num} container")
        container_found = False
        for container_idx, container in enumerate(self.parent_form.form_content.controls):
            if (isinstance(container, ft.Container) and 
                isinstance(container.content, ft.Column) and
                container.content.controls and 
                isinstance(container.content.controls[0], ft.Text) and
                container.content.controls[0].value == "Strut Details"):
                
                print(f"DEBUG: Found 'Strut Details' container at index {container_idx}")
                
                # Search through the sub-frames in the strut frame
                sub_container_found = False
                for sub_idx, sub_container in enumerate(container.content.controls[2:]):  # Skip title and base controls
                    if (isinstance(sub_container, ft.Container) and
                        isinstance(sub_container.content, ft.Column) and
                        sub_container.content.controls and
                        isinstance(sub_container.content.controls[0], ft.Text) and
                        sub_container.content.controls[0].value == f"Strut {strut_num}"):
                        
                        print(f"DEBUG: Found 'Strut {strut_num}' sub-container at index {sub_idx+2}")
                        sub_container_found = True
                        
                        # This is the correct sub-frame - update its fields
                        print(f"DEBUG: Creating new controls for Strut {strut_num}")
                        strut_fields = self._get_strut_fields(strut_num)
                        print(f"DEBUG: Strut fields: {strut_fields}")
                        strut_controls = self._create_category_controls(strut_fields, self.form_values)
                        
                        # Replace the existing controls with new ones
                        print(f"DEBUG: Replacing controls in sub-container")
                        sub_container.content.controls[1] = strut_controls
                        
                        # Update the UI
                        print(f"DEBUG: Updating UI")
                        if self.parent_form and self.parent_form.page:
                            self.parent_form.page.update()
                        break
                
                if not sub_container_found:
                    print(f"WARNING: No sub-container found for Strut {strut_num}")
                
                container_found = True
                break
        
        if not container_found:
            print(f"WARNING: No 'Strut Details' container found")
        
        print(f"DEBUG: Exiting handle_strut_material_change successfully")
    
      except Exception as ex:
        print(f"ERROR in handle_strut_material_change: {str(ex)}")
        import traceback
        traceback.print_exc()
        print(f"DEBUG: Control value: {e.control.value if hasattr(e, 'control') and hasattr(e.control, 'value') else 'N/A'}")
        print(f"DEBUG: Control label: {e.control.label if hasattr(e, 'control') and hasattr(e.control, 'label') else 'N/A'}")
    def handle_field_change(self, e):
      """Handle real-time field validation on change"""
      print(f"DEBUG: Entering handle_field_change")
    
      try:
        field = e.control
        field_name = field.label.split(" (Set")[0].split(" *")[0]
        print(f"DEBUG: Field name: {field_name}, Field type: {type(field).__name__}")
        
        if isinstance(field, ft.Checkbox):
            value = field.value
            print(f"DEBUG: Checkbox value: {value}")
        else:
            value = field.value
            print(f"DEBUG: Field value: {value}")
        
        # Update form values
        previous_value = self.form_values.get(field_name, "Not set")
        self.form_values[field_name] = value
        print(f"DEBUG: Updated form value for '{field_name}' from '{previous_value}' to '{value}'")
        
        if "Space" in field_name:
            print(f"DEBUG: *** SPACE FIELD DETECTED: {field_name} = {value} ***")
            # Force immediate update to ensure the value is stored
            if self.parent_form and self.parent_form.page:
                self.parent_form.page.update()
        
        # Special handling for Excavation Type
        if field_name == "Excavation Type" and value == "Single wall":
            print("DEBUG: Excavation Type is 'Single wall', auto-setting Strut Type to 'Fixed'")
            
            # Update form values
            self.form_values["Strut Type"] = "Fixed"
            
            # Find and update the Strut Type control
            strut_type_control = self.find_field_control("Strut Type")
            if strut_type_control:
                print("DEBUG: Found Strut Type control, updating to 'Fixed'")
                strut_type_control.value = "Fixed"
                strut_type_control.update()
                
                # Trigger strut type change to update the UI with Fixed strut fields
                print("DEBUG: Triggering strut type change handler")
                # Create a mock event for the strut type change
                class MockEvent:
                    def __init__(self, control, data):
                        self.control = control
                        self.data = data
                
                mock_event = MockEvent(strut_type_control, "Fixed")
                self.handle_strut_type_change(mock_event)
            else:
                print("WARNING: Could not find Strut Type control to update")
        
        # Validate the current field
        error = self.validate_field(field_name, value, self.form_values)
        print(f"DEBUG: Validation result for '{field_name}': {'Error: ' + error if error else 'Valid'}")
        
        # Update field appearance based on validation
        if error:
            field.border_color = "red"
            field.hint_text = error
            self.field_errors[field_name] = error
            print(f"DEBUG: Set field '{field_name}' to error state with message: {error}")
        else:
            field.border_color = "black"
            field.hint_text = ""
            self.field_errors.pop(field_name, None)
            print(f"DEBUG: Cleared error state for field '{field_name}'")
        
        # If this field's value affects other fields' validation, revalidate them
        dependent_fields = self.get_dependent_fields(field_name)
        print(f"DEBUG: Found {len(dependent_fields)} dependent fields for '{field_name}': {dependent_fields}")
        
        for dep_field_name in dependent_fields:
            if dep_field_name in self.form_values:
                print(f"DEBUG: Revalidating dependent field '{dep_field_name}'")
                dep_error = self.validate_field(
                    dep_field_name, 
                    self.form_values[dep_field_name], 
                    self.form_values
                )
                
                # Update the dependent field's appearance
                dep_field = self.find_field_control(dep_field_name)
                if dep_field:
                    if dep_error:
                        dep_field.border_color = "red"
                        dep_field.hint_text = dep_error
                        self.field_errors[dep_field_name] = dep_error
                        print(f"DEBUG: Set dependent field '{dep_field_name}' to error state with message: {dep_error}")
                    else:
                        dep_field.border_color = "green"
                        dep_field.hint_text = ""
                        self.field_errors.pop(dep_field_name, None)
                        print(f"DEBUG: Cleared error state for dependent field '{dep_field_name}'")
                    dep_field.update()
                else:
                    print(f"WARNING: Could not find control for dependent field '{dep_field_name}'")
        
        field.update()
        print(f"DEBUG: Exiting handle_field_change successfully")
        
      except Exception as ex:
        print(f"ERROR in handle_field_change: {str(ex)}")
        import traceback
        traceback.print_exc()
        # Additional debug info in case of error
        print(f"DEBUG: Event info - {e}")
        if hasattr(e, 'control'):
            print(f"DEBUG: Control type: {type(e.control).__name__}")
            print(f"DEBUG: Control attributes: {dir(e.control)}")

    def capture_all_ui_values(self):
      """Force capture all current values from UI controls"""
      print("DEBUG: Forcing capture of all UI values")
    
      def traverse_controls(control):
        values = {}
        if isinstance(control, (ft.TextField, ft.Dropdown)):
            if hasattr(control, 'label') and control.label:
                clean_label = control.label.split(" (Set")[0].split(" *")[0]
                values[clean_label] = control.value
                print(f"DEBUG: Captured {clean_label} = {control.value}")
        elif hasattr(control, 'controls'):
            for child in control.controls:
                child_values = traverse_controls(child)
                values.update(child_values)
        elif hasattr(control, 'content'):
            content_values = traverse_controls(control.content)
            values.update(content_values)
        return values
    
      if self.parent_form and self.parent_form.form_content:
        ui_values = traverse_controls(self.parent_form.form_content)
        self.form_values.update(ui_values)
        print(f"DEBUG: Updated form_values with {len(ui_values)} UI values")
    def get_dependent_fields(self, field_name: str) -> List[str]:
       """Return list of field names that depend on the given field"""
       dependencies = {
        "Wall Top Level": ["Ground Water Table", "Toe Level"] + 
                        [f"Strut {i} Level" for i in range(1, self.current_num_struts + 1)],
        "Excavation Depth": ["Toe Level"] + 
                           [f"Strut {i} Level" for i in range(1, self.current_num_struts + 1)],
        "Strut Type": (["Strut Space", "Struct Length", "Angle"] +  # Old fields (for backward compatibility)
                      [f"Strut {i} Space" for i in range(1, self.current_num_struts + 1)] +
                      [f"Strut {i} Length" for i in range(1, self.current_num_struts + 1)] +
                      [f"Strut {i} Angle" for i in range(1, self.current_num_struts + 1)]),
        "Excavation Type": ["Strut Type"],
        "Excavation Width": ([f"Strut {i} Length" for i in range(1, self.current_num_struts + 1)] 
                           if self.form_values.get("Strut Type") == "Fixed" else [])
       }
       return dependencies.get(field_name, [])

    def handle_excavation_type_change(self, e):
      """Handle excavation type change and automatically set Strut Type"""
      try:
        excavation_type = e.data or e.control.value
        print(f"DEBUG: Excavation type changed to: {excavation_type}")
        
        # Update form value
        self.form_values["Excavation Type"] = excavation_type
        
        # Automatically set Strut Type based on Excavation Type
        if excavation_type == "Single wall":
            target_strut_type = "Fixed"
            print("DEBUG: Single wall detected - setting Strut Type to 'Fixed'")
        elif excavation_type == "Double wall":
            target_strut_type = "Node to Node"
            print("DEBUG: Double wall detected - setting Strut Type to 'Node to Node'")
        else:
            # For Cassion or other types, don't auto-set
            print(f"DEBUG: Excavation type '{excavation_type}' - no auto Strut Type setting")
            target_strut_type = None
        
        if target_strut_type:
            # Update form values
            self.form_values["Strut Type"] = target_strut_type
            
            # Find and update the Strut Type control
            strut_type_control = self.find_field_control("Strut Type")
            if strut_type_control:
                print(f"DEBUG: Found Strut Type control, updating to '{target_strut_type}'")
                strut_type_control.value = target_strut_type
                
                # Disable the Strut Type dropdown to prevent manual changes
                strut_type_control.disabled = True
                strut_type_control.update()
                
                # Trigger strut type change to update the UI with appropriate fields
                print("DEBUG: Triggering strut type change handler")
                class MockEvent:
                    def __init__(self, control, data):
                        self.control = control
                        self.data = data
                
                mock_event = MockEvent(strut_type_control, target_strut_type)
                self.handle_strut_type_change(mock_event)
        
        # **NEW: Update Line Load Details frame**
        self.line_load_handler.update_line_load_frame()

        
        # Update the page
        if self.parent_form and self.parent_form.page:
            self.parent_form.page.update()
        
        print(f"DEBUG: Successfully processed excavation type change")
        
      except Exception as ex:
        print(f"ERROR in handle_excavation_type_change: {str(ex)}")
        import traceback
        traceback.print_exc()

    def handle_strut_change(self, e):
       try:
        print(f"DEBUG: handle_strut_change called with data: {e.data}")
        
        # Handle empty/None input - don't change UI visibility
        if not e.data or e.data.strip() == "":
            print("DEBUG: Empty strut input - maintaining current UI state")
            # Clear the form value but don't change current_num_struts or UI
            self.form_values["No of Strut"] = ""
            return
            
        try:
            new_num_struts = int(e.data)
            if new_num_struts < 0:
                raise ValueError("Number of struts cannot be negative")
            if new_num_struts > 100:
                raise ValueError("Maximum number of struts exceeded")
            print(f"DEBUG: Number of struts changed to {new_num_struts}")
        except ValueError as ve:
            print(f"Invalid strut number: {str(ve)}")
            return

        # Preserve existing strut data
        existing_strut_data = {}
        print(f"DEBUG: Preserving data for existing {self.current_num_struts} struts")
        for i in range(1, min(self.current_num_struts, new_num_struts) + 1):
            print(f"DEBUG: Collecting existing data for Strut {i}")
            existing_strut_data[i] = {
                f'Strut {i} Level': self.form_values.get(f'Strut {i} Level'),
                f'Strut {i} Material': self.form_values.get(f'Strut {i} Material'),
            }
            print(f"DEBUG: Strut {i} Level: {existing_strut_data[i][f'Strut {i} Level']}")
            print(f"DEBUG: Strut {i} Material: {existing_strut_data[i][f'Strut {i} Material']}")
            
            if existing_strut_data[i][f'Strut {i} Material'] == 'Concrete':
                existing_strut_data[i][f'Strut {i} Stiffness'] = self.form_values.get(f'Strut {i} Stiffness')
                existing_strut_data[i][f'Strut {i} EA'] = self.form_values.get(f'Strut {i} EA')
                print(f"DEBUG: Concrete Strut {i} - Stiffness: {existing_strut_data[i][f'Strut {i} Stiffness']}, EA: {existing_strut_data[i][f'Strut {i} EA']}")
            elif existing_strut_data[i][f'Strut {i} Material'] == 'Steel':
                existing_strut_data[i][f'Strut {i} Member Size'] = self.form_values.get(f'Strut {i} Member Size')
                print(f"DEBUG: Steel Strut {i} - Member Size: {existing_strut_data[i][f'Strut {i} Member Size']}")

        # Update current number of struts
        self.current_num_struts = new_num_struts
        self.form_values["No of Strut"] = str(new_num_struts)
        print(f"DEBUG: Updated current_num_struts to {self.current_num_struts}")
        print(f"DEBUG: Updated form_values['No of Strut'] to {self.form_values['No of Strut']}")
        
        # Find the Strut Details frame
        strut_frame = None
        for container in self.parent_form.form_content.controls:
            if (hasattr(container, 'content') and 
                isinstance(container.content, ft.Column) and
                container.content.controls and 
                isinstance(container.content.controls[0], ft.Text) and
                container.content.controls[0].value == "Strut Details"):
                strut_frame = container
                print("DEBUG: Found Strut Details frame")
                break
                
        if not strut_frame:
            print("Error: Strut Details frame not found")
            return
            
        # Create the base controls (No of Strut and Strut Type)
        strut_base_fields = ["No of Strut", "Strut Type"]
        base_strut_fields = [f for f in self.get_fields() if f.label in strut_base_fields]
        base_strut_controls = self._create_category_controls(base_strut_fields, self.form_values)
        print(f"DEBUG: Created base strut controls for fields: {strut_base_fields}")
        
        # Create a new content structure for the strut frame
        strut_frame_content = ft.Column([
            ft.Text("Strut Details", size=20, weight=ft.FontWeight.BOLD),
            base_strut_controls
        ])
        
        # Update the strut frame with new content
        strut_frame.content = strut_frame_content
        print("DEBUG: Updated strut frame content with base controls")
        
        # Only add sub-frames if there are struts to display
        if new_num_struts > 0:
            print(f"DEBUG: Adding {new_num_struts} strut sub-frames using helper method")
            # Flatten existing_strut_data for use with _add_strut_subframes
            flattened_data = {}
            for strut_num, strut_data in existing_strut_data.items():
                flattened_data.update(strut_data)
            
            # Use the helper method to add sub-frames
            self._add_strut_subframes(new_num_struts, flattened_data)
        
        # Update visibility of frames and buttons
        frames = {}
        button_containers = []
        for container in self.parent_form.form_content.controls:
            if not isinstance(container, ft.Container):
                continue
                
            if (hasattr(container, 'content') and 
                isinstance(container.content, ft.Column) and
                container.content.controls and 
                isinstance(container.content.controls[0], ft.Text)):
                frame_title = container.content.controls[0].value
                frames[frame_title] = container
            elif (hasattr(container, 'content') and
                isinstance(container.content, ft.ElevatedButton)):
                button_containers.append(container)
        
        print(f"DEBUG: Found {len(frames)} frames and {len(button_containers)} button containers")
                
        frame_order = ["Excavation Details", "Strut Details", "Wall Details", "Borehole Details"]
        
        # Always keep Excavation Details frame visible
        if "Excavation Details" in frames:
            frames["Excavation Details"].visible = True
            print("DEBUG: Ensured Excavation Details frame is visible")
        
        # MODIFIED: Always keep the Strut Details frame visible and its button
        if "Strut Details" in frames:
            frames["Strut Details"].visible = True
            print("DEBUG: Ensured Strut Details frame is visible")
        
        # Always keep the first button (Excavation Details) visible
        if button_containers:
            button_containers[0].visible = True
            print("DEBUG: Ensured Excavation Details button is visible")
        
        # Set visibility based on number of struts
        if new_num_struts > 0:
            # Show Strut Details button if it exists
            strut_button_index = frame_order.index("Strut Details")
            if strut_button_index < len(button_containers):
                button_containers[strut_button_index].visible = True
                print(f"DEBUG: Made Strut Details button visible")
            
            # Hide subsequent frames and buttons (Wall Details, Borehole Details)
            for i in range(strut_button_index + 1, len(button_containers)):
                button_containers[i].visible = False
                print(f"DEBUG: Made button at index {i} invisible")
            for title in frame_order[frame_order.index("Strut Details") + 1:]:
                if title in frames:
                    frames[title].visible = False
                    print(f"DEBUG: Made {title} frame invisible")
        else:
            # When struts = 0: Keep Strut Details frame visible with base controls only
            # The strut frame content has already been updated above with only base controls
            print("DEBUG: Struts = 0, keeping Strut Details frame with base controls only")
            
            # Hide other frames (Wall Details, Borehole Details) but keep buttons visible
            # so user can navigate back if needed
            for title in ["Wall Details", "Borehole Details"]:
                if title in frames:
                    frames[title].visible = False
                    print(f"DEBUG: Made {title} frame invisible (no struts)")
            
            # Keep Strut Details button visible so user can continue
            strut_button_index = frame_order.index("Strut Details")
            if strut_button_index < len(button_containers):
                button_containers[strut_button_index].visible = True
                print(f"DEBUG: Kept Strut Details button visible for re-entry")
                    
        # Update the UI
        self._validate_strut_fields()
        print("DEBUG: Validated strut fields")
        self.parent_form.form_content.update()
        print("DEBUG: Updated form content")
        if self.parent_form and self.parent_form.page:
            self.parent_form.page.update()
            print("DEBUG: Updated page")
            
       except Exception as ex:
        print(f"Error in handle_strut_change: {str(ex)}")
        import traceback
        traceback.print_exc()
    def _validate_strut_fields(self):
      """Validates required strut fields are properly filled."""
      try:
        # Check if required strut fields are filled
        if self.current_num_struts > 0:
            # Check if base strut fields are filled
            if "Strut Type" not in self.form_values or not self.form_values["Strut Type"]:
                return False
                
            for i in range(1, self.current_num_struts + 1):
                # Check common required fields
                if not self.form_values.get(f"Strut {i} Level"):
                    return False
                if not self.form_values.get(f"Strut {i} Material"):
                    return False
                    
                # Material-specific validation
                if self.form_values.get(f"Strut {i} Material") == "Steel":
                    if not self.form_values.get(f"Strut {i} Member Size"):
                        return False
                elif self.form_values.get(f"Strut {i} Material") == "Concrete":
                    if not self.form_values.get(f"Strut {i} Stiffness") or not self.form_values.get(f"Strut {i} EA"):
                        return False
        
        return True
        
      except Exception as ex:
        print(f"Error in _validate_strut_fields: {str(ex)}")
        return False 
    def update_strut_frame(self):
      """Update only the strut details frame while preserving other frames"""
      if not self.parent_form:
        return
        
    # Find and update only the strut frame
      for i, control in enumerate(self.parent_form.form_content.controls):
        if (isinstance(control, ft.Container) and 
            isinstance(control.content, ft.Column) and 
            isinstance(control.content.controls[0], ft.Text) and
            control.content.controls[0].value == "Strut Details"):
            
            # Get all fields including new dynamic strut fields
            all_fields = self.get_fields()
            
            # Filter strut-related fields
            strut_base_fields = ["No of Strut", "Strut Type"]
            strut_fields = [f for f in all_fields if f.label in strut_base_fields]
            strut_fields.extend([f for f in all_fields if f.label.startswith("Strut") and f.label not in strut_base_fields])
            
            # Create new strut controls
            strut_controls = self._create_category_controls(strut_fields)
            
            # Create new strut frame
            new_strut_frame = self.parent_form._create_frame("Strut Details", strut_controls)
            
            # Replace the old frame with the new one
            self.parent_form.form_content.controls[i] = new_strut_frame
            self.parent_form.form_content.update()
            break 
     
    def calculate_coordinates(self, data: Dict) -> Tuple[float, float, float, float]:
      print(f"DEBUG: Starting coordinate calculation with data: {data}")
    
      try:
        wall_top_level = self.safe_float(data.get("Wall Top Level", 0))
        excavation_depth = self.safe_float(data.get("Excavation Depth", 0))
        excavation_width = self.safe_float(data.get("Excavation Width", 0))
        toe_level = self.safe_float(data.get("Toe Level", 0))
        
        print(f"DEBUG: Using values: wall_top={wall_top_level}, exc_depth={excavation_depth}, width={excavation_width}, toe={toe_level}")
        
        if wall_top_level is None or excavation_depth is None or excavation_width is None or toe_level is None:
            print("WARNING: Missing values for coordinate calculation, using defaults")
            return 0, 0, 0, 0
            
        x_min = -(excavation_width/2 + 4*excavation_depth)
        y_min = wall_top_level - 4*excavation_depth
        x_max = (excavation_width/2 + 4*excavation_depth)
        y_max = wall_top_level + excavation_depth
        
        print(f"DEBUG: Calculated coordinates: {x_min}, {y_min}, {x_max}, {y_max}")
        return x_min, y_min, x_max, y_max
      except Exception as e:
        print(f"ERROR in calculate_coordinates: {str(e)}")
        return 0, 0, 0, 0

    def format_strut_type_for_excel(strut_type):
      """Convert strut type to Excel format:
      - 'Node to Node' -> 'n2n'
      - 'Fixed' -> 'fixedend'
      """
      if strut_type == "Node to Node":
          return "n2n"
      elif strut_type == "Fixed":
          return "fixedend"
      else:
          return strut_type  # Return as-is if unknown type
    
    def save(self, cursor, data: Dict) -> None:
      try:
        print(f"DEBUG: Starting save method with data: {data}")
        num_struts = 0
        if "No of Strut" in data:
            try:
                num_struts = int(data.get("No of Strut", 0))
                print(f"DEBUG: Number of struts: {num_struts}")
            except (ValueError, TypeError):
                print(f"Invalid value for 'No of Strut': {data.get('No of Strut')}. Using 0.")
        else:
            print("Warning: 'No of Strut' not found in data. Using 0.")

        print(f"DEBUG: All keys in data dictionary: {list(data.keys())}")
        strut_data = {}
        for i in range(1, num_struts + 1):
            form_level_key = f"Strut {i} Level"
            form_material_key = f"Strut {i} Material"
            form_size_key = f"Strut {i} Member Size"
            form_space_key = f"Strut {i} Space"
            form_length_key = f"Strut {i} Length"  # NEW: Per-strut length
            form_angle_key = f"Strut {i} Angle"    # NEW: Per-strut angle
            
            excel_level_key = f"Strut{i}Level"
            excel_material_key = f"Strut{i}Material" 
            excel_size_key = f"Strut{i}MemberSize"
            excel_space_key = f"Strut{i}Space"  # New Excel key
            excel_length_key = f"Strut{i}Length"   # NEW: Excel keys for length/angle
            excel_angle_key = f"Strut{i}Angle"

            strut_level_value = self.safe_float(data.get(form_level_key))
            strut_material_value = data.get(form_material_key, "Steel")
            strut_size_value = data.get(form_size_key, "N/A")
            strut_space_value = self.safe_float(data.get(form_space_key))
            strut_length_value = self.safe_float(data.get(form_length_key, 0.0))
            strut_angle_value = self.safe_float(data.get(form_angle_key, 0.0))    # NEW
             
            if strut_size_value and strut_size_value != "N/A" and not strut_size_value.startswith("UC"):
                strut_size_value = f"UC{strut_size_value}"
                
            print(f"DEBUG: For {form_level_key}, found value: {strut_level_value}")
            print(f"DEBUG: For {form_material_key}, found value: {strut_material_value}")
            print(f"DEBUG: For {form_size_key}, found value: {strut_size_value}")
            
            strut_data[excel_level_key] = strut_level_value
            strut_data[excel_material_key] = strut_material_value
            strut_data[excel_size_key] = strut_size_value
            strut_data[excel_space_key] = strut_space_value
            strut_data[excel_length_key] = strut_length_value  # NEW
            strut_data[excel_angle_key] = strut_angle_value
            
            if strut_level_value is None:
                if excel_level_key in data:
                    print(f"DEBUG: Found alternative key {excel_level_key} with value {data[excel_level_key]}")
                    strut_data[excel_level_key] = self.safe_float(data[excel_level_key])
                    
            if strut_material_value is None or strut_material_value == "Steel":
                if excel_material_key in data:
                    print(f"DEBUG: Found alternative key {excel_material_key} with value {data[excel_material_key]}")
                    strut_data[excel_material_key] = data[excel_material_key]
                    
            if strut_size_value is None or strut_size_value == "N/A":
                if excel_size_key in data:
                    size_value = data[excel_size_key]
                    if size_value and size_value != "N/A" and not size_value.startswith("UC"):
                        size_value = f"UC{size_value}"
                    print(f"DEBUG: Found alternative key {excel_size_key} with value {size_value}")
                    strut_data[excel_size_key] = size_value

        for i in range(1, num_struts + 1):
            excel_level_key = f"Strut{i}Level"
            excel_material_key = f"Strut{i}Material"
            excel_size_key = f"Strut{i}MemberSize"
            excel_space_key = f"Strut{i}Space"  # New key

            print(f"DEBUG: Final Excel data for Strut {i}:")
            print(f"DEBUG: {excel_level_key}: {strut_data.get(excel_level_key)}")
            print(f"DEBUG: {excel_material_key}: {strut_data.get(excel_material_key)}")
            print(f"DEBUG: {excel_size_key}: {strut_data.get(excel_size_key)}")

        common_id = data.get('common_id')
        excavation_type = data.get("Excavation Type", "Single Wall")
        wall_type = data.get("Wall Type", "Sheet Pile")
        wall_top_level = self.safe_float(data.get("Wall Top Level"))
        excavation_depth = self.safe_float(data.get("Excavation Depth"))
        excavation_width = self.safe_float(data.get("Excavation Width"))
        toe_level = self.safe_float(data.get("Toe Level"))
        
        spacing = data.get("Spacing", "")
        
        
        try:
            if wall_top_level is None or excavation_depth is None or excavation_width is None:
                print("ERROR: Missing critical values for coordinate calculation")
                x_min, y_min, x_max, y_max = 0, 0, 0, 0
            else:
                x_min, y_min, x_max, y_max = self.calculate_coordinates(data)
                print(f"DEBUG: Calculated coordinates: x_min={x_min}, y_min={y_min}, x_max={x_max}, y_max={y_max}")
        except Exception as e:
            print(f"ERROR in coordinate calculation: {str(e)}")
            x_min, y_min, x_max, y_max = 0, 0, 0, 0

        export_dir = self.export_dir
        export_dir.mkdir(exist_ok=True)
        
        base_parameter_names = [
            "ExcavationType", "WallTopLevel", "ExcavationDepth",
            "ExcavationWidth", "ToeLevel", "NoOfStrut"
        ]
        
        strut_parameter_names = []
        for i in range(1, num_struts + 1):
            strut_parameter_names.extend([
                f"Strut{i}Level", 
                f"Strut{i}Material", 
                f"Strut{i}MemberSize",
                f"Strut{i}Space"  # Add Space parameter
        
            ])
            # Add length and angle parameters for Fixed struts
            if data.get("Strut Type") == "Fixed":
                strut_parameter_names.extend([
                    f"Strut{i}Length",
                    f"Strut{i}Angle"
                ])
                
        remaining_parameter_names = [
            "ExcavationBelowStrut", "OverExcavation", "WallType",
            "Material", "MemberSize", "Spacing", "Borehole_x_coordinate",
            "GroundWatertable", "x_min_coordinate", "y_min_coordinate",
            "x_max_coordinate", "y_max_coordinate"
        ]
        
        excel_parameter_names = base_parameter_names + strut_parameter_names + remaining_parameter_names

        db_base_params = [
            "excavation_type", "wall_top_level", "excavation_depth",
            "excavation_width", "toe_level", "no_of_strut"
        ]
        
        db_strut_params = []
        for i in range(1, num_struts + 1):
            db_strut_params.extend([
                f"strut{i}_level", 
                f"strut{i}_material", 
                f"strut{i}_member_size",
                f"strut{i}_space"  # Add Space parameter
        
            ])
            if data.get("Strut Type") == "Fixed":
                db_strut_params.extend([
                    f"Strut{i}Length",
                    f"Strut{i}Angle"
                ])
            
        db_remaining_params = [
            "excavation_below_strut", "over_excavation", "wall_type",
            "material", "member_size", "spacing", "borehole_x_coordinate",
            "ground_watertable", "x_min_coordinate", "y_min_coordinate",
            "x_max_coordinate", "y_max_coordinate"
        ]
        
        db_parameter_names = db_base_params + db_strut_params + db_remaining_params

        sheets_config = {
            "Geometry Info": {
                "headers": ["Parameters", "Value"],
                "parameter_names": excel_parameter_names,
                "db_parameter_names": db_parameter_names,
                "csv_file": export_dir / "geometry.csv",
                "db_table": "geometry"
            },
            "ERSS Wall Detail": {
                "headers": ["MaterialName", "WallName", "x_Top", "y_Top", "x_Bottom", "y_Bottom"],
                "csv_file": export_dir / "erss_wall_details.csv",
                "db_table": "erss_wall_details"
            },
            "Strut Details": {
                "headers": ["MaterialName", "StrutName", "x_Left", "y_Left", "x_Right", "y_Right", "Type", "Direction_x", "Direction_y"],
                "csv_file": export_dir / "strut_details.csv",
                "db_table": "StrutDetails"
            },
            "Plate Properties": {  # ADD THIS
                "headers": ["MaterialName", "Elasticity", "IsIsotropic", "EA", "EA2", "EI", "d", "E", "Gref", "w", "StrutNu", "Colour"],
                "csv_file": export_dir / "plate_properties.csv",
                "db_table": "plate_properties"
            },
            "Anchor Properties": {
                "headers": ["MaterialName", "Elasticity", "EA", "Lspacing", "Colour"],
                "csv_file": export_dir / "anchor_properties.csv",
                "db_table": "anchor_properties"
            },
            "Line Load": {
                "headers": ["LoadName", "x_start", "y_start", "x_end", "y_end", "qx_start", "qy_start", "Distribution"],
                "csv_file": export_dir / "lineload.csv",
                "db_table": "lineload"
            }
        }

        excel_filename = export_dir / "Input_Data.xlsx"
        if not excel_filename.exists():
            workbook = openpyxl.Workbook()
            workbook.active.title = "Project Info"
        else:
            workbook = openpyxl.load_workbook(excel_filename)
            
        excel_sheets = {}
        for sheet_name, config in sheets_config.items():
            if sheet_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(sheet_name)
                sheet.append(config["headers"])
                  
                if sheet_name == "Geometry Info":
                    for row_idx, param_name in enumerate(config["parameter_names"], 2):
                        sheet.cell(row=row_idx, column=1, value=param_name)
            else:
                sheet = workbook[sheet_name]
                if sheet.max_row > 1:
                    sheet.delete_rows(2, sheet.max_row - 1)
                if sheet_name == "Geometry Info":
                    for row_idx, param_name in enumerate(config["parameter_names"], 2):
                        sheet.cell(row=row_idx, column=1, value=param_name)
                        
            excel_sheets[sheet_name] = sheet
            
            for sheet_name, config in sheets_config.items():
                if not config["csv_file"].exists():
                    if sheet_name == "Line Load":
                        # Use the line load handler to initialize CSV
                        self.line_load_handler.initialize_csv_file({sheet_name: config})
                    else:
                        with open(config["csv_file"], 'w', newline='') as f:
                            writer = csv.writer(f)
                            if sheet_name == "Geometry Info":
                                writer.writerow(["common_id"] + config["db_parameter_names"])
                            else:
                                writer.writerow(["common_id"] + config["headers"])

        borehole_x_coordinate = data.get("Borehole X Coordinate")
        if borehole_x_coordinate is not None:
            try:
                borehole_x_coordinate = int(float(borehole_x_coordinate))
            except (ValueError, TypeError):
                print(f"WARNING: Invalid borehole X coordinate: {borehole_x_coordinate}, using 0")
                borehole_x_coordinate = 0
        else:
            print("WARNING: Borehole X coordinate not provided, using 0")
            borehole_x_coordinate = 0

        geometry_data = [
            excavation_type,
            wall_top_level,
            excavation_depth,
            excavation_width,
            toe_level,
            num_struts
        ]
        
        print(f"DEBUG: Initial geometry_data: {geometry_data}")
        excel_geometry_data = list(geometry_data)
        
        for i in range(1, num_struts + 1):
            excel_level_key = f"Strut{i}Level"
            excel_material_key = f"Strut{i}Material"
            excel_size_key = f"Strut{i}MemberSize"
            excel_space_key = f"Strut{i}Space"
            
            strut_level = strut_data.get(excel_level_key, "")
            strut_material = strut_data.get(excel_material_key, "Steel")
            strut_member_size = strut_data.get(excel_size_key, "N/A")
            strut_space = strut_data.get(excel_space_key, "")  # Get space value
            
            excel_geometry_data.extend([
                strut_level,
                strut_material,
                strut_member_size,
                strut_space  # Add space to data array

            ])
            if data.get("Strut Type") == "Fixed":
                excel_length_key = f"Strut{i}Length"
                excel_angle_key = f"Strut{i}Angle"
                strut_length = strut_data.get(excel_length_key, 0.0)
                strut_angle = strut_data.get(excel_angle_key, 0.0)
                
                excel_geometry_data.extend([
                    strut_length,
                    strut_angle
                ])
        # Process MemberSize for UB prefix
        member_size = data.get("Member Size") or "NO"
        if member_size and member_size != "NO" and not member_size.startswith("UB"):
            member_size = f"UB{member_size}"
        
        excavation_below_strut = data.get("Excavation Below Strut")
        over_excavation = data.get("Over Excavation")
            
        remaining_data = [
            excavation_below_strut,
            over_excavation,
            wall_type,
            data.get("Material"),
            member_size,
            spacing,
            borehole_x_coordinate,
            self.safe_float(data.get("Ground Water Table")),
            x_min,
            y_min,
            x_max,
            y_max
        ]
        
        geometry_data.extend(remaining_data)
        excel_geometry_data.extend(remaining_data)
        
        # Add RC Wall specific fields if applicable
        if wall_type == "RC Wall":
            print("DEBUG: Adding RC Wall specific fields")
            stiffness_wall = data.get("Stiffness", False)
            additional_fields = [stiffness_wall, data.get("EA"), data.get("EI"), data.get("V")]
            geometry_data.extend(additional_fields)
            excel_geometry_data.extend(additional_fields)
              
            rc_wall_params = ["stiffness_wall", "EA", "EI", "V"]
            db_remaining_params.extend(rc_wall_params)
            remaining_parameter_names.extend(["StiffnessWall", "EA", "EI", "V"])
        
        # Add Fixed Strut specific fields if applicable
        if data.get("Strut Type") == "Fixed":
            print("DEBUG: Adding Fixed Strut specific fields")
            additional_fields = [
                data.get("Strut Space", 0),
                data.get("Strut Length", 0),
                data.get("Angle", 0)
            ]
            geometry_data.extend(additional_fields)
            excel_geometry_data.extend(additional_fields)
              
            fixed_strut_params = ["strut_space", "strut_length", "angle"]
            db_remaining_params.extend(fixed_strut_params)
            remaining_parameter_names.extend(["StrutSpace", "StrutLength", "Angle"])
        
        # Save to database
        cursor.execute("PRAGMA table_info(geometry)")
        existing_columns = [row[1] for row in cursor.fetchall()]
        print(f"DEBUG: Existing columns in geometry table: {existing_columns}")
        
        if 'common_id' in existing_columns:
            existing_columns.remove('common_id')
            
        filtered_params = []
        filtered_data = []
        db_params_no_strut = db_base_params + db_remaining_params
        geometry_data_no_strut = geometry_data[:len(db_base_params)] + geometry_data[len(db_base_params) + len(db_strut_params):]
        
        for param, value in zip(db_params_no_strut, geometry_data_no_strut):
            if param in existing_columns:
                filtered_params.append(param)
                filtered_data.append(value)
            else:
                print(f"DEBUG: Skipping non-existent column {param}")
                
        strut_db_values = []
        for i in range(1, num_struts + 1):
            excel_level_key = f"Strut{i}Level"
            excel_material_key = f"Strut{i}Material"
            excel_size_key = f"Strut{i}MemberSize"
            
            strut_level = strut_data.get(excel_level_key)
            strut_material = strut_data.get(excel_material_key)
            strut_member_size = strut_data.get(excel_size_key)
            
            db_level_key = f"strut{i}_level"
            db_material_key = f"strut{i}_material"
            db_size_key = f"strut{i}_member_size"
            
            if db_level_key in existing_columns:
                filtered_params.append(db_level_key)
                filtered_data.append(strut_level)
            if db_material_key in existing_columns:
                filtered_params.append(db_material_key)
                filtered_data.append(strut_material)
            if db_size_key in existing_columns:
                filtered_params.append(db_size_key)
                filtered_data.append(strut_member_size)
        
        columns = ", ".join(filtered_params)
        placeholders = ", ".join(["?"] * len(filtered_data))
        query = f"INSERT INTO {sheets_config['Geometry Info']['db_table']} (common_id, {columns}) VALUES (?, {placeholders})"
        print(f"DEBUG: Executing SQL: {query}")
        cursor.execute(query, [common_id] + filtered_data)
        
        # Write to Excel
        geometry_sheet = excel_sheets["Geometry Info"]
        
        for idx, (param_name, param_value) in enumerate(zip(excel_parameter_names, excel_geometry_data), start=2):
            print(f"DEBUG: Writing param {param_name} at row {idx}: {param_value}")
            cell = geometry_sheet.cell(row=idx, column=2, value=param_value)
            
            # Set number formatting for numeric values
            if isinstance(param_value, (int, float)) and param_value is not None:
                if param_value == int(param_value):
                    cell.number_format = '0'
                else:
                    cell.number_format = '0.00'
            
            # Right-align specific parameters and ensure proper number formatting
            if param_name in ["ExcavationBelowStrut", "OverExcavation", "Spacing", "GroundWatertable"]:
                print(f"DEBUG: Aligning {param_name} value to right: {param_value}")
                cell.alignment = Alignment(horizontal='right')
                
                # Convert string values to numbers for these fields to avoid green triangles
                if isinstance(param_value, str) and param_value.strip():
                    try:
                        # Try to convert to float
                        if param_value != "NO":  # Keep "NO" as string for Spacing
                            float_value = float(param_value)
                            cell.value = float_value
                            if float_value == int(float_value):
                                cell.number_format = '0'
                            else:
                                cell.number_format = '0.00'
                    except ValueError:
                        # If conversion fails, keep as string but right-aligned
                        pass
        
        # Write to CSV
        with open(sheets_config["Geometry Info"]["csv_file"], 'a', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([common_id] + excel_geometry_data)
        
        # Now save wall details using the wall details handler
        self.wall_details_handler.save_wall_details(
            cursor=cursor,
            data=data,
            common_id=common_id,
            excel_sheets=excel_sheets,
            sheets_config=sheets_config
        )
        
        plate_props = self.get_plate_properties(cursor, wall_type)
        def format_strut_type_for_excel(strut_type):
            """Convert strut type to Excel format:
            - 'Node to Node' -> 'n2n'
            - 'Fixed' -> 'fixedend'
            """
            if strut_type == "Node to Node":
                return "n2n"
            elif strut_type == "Fixed":
                return "fixedend"
            else:
                return strut_type
        
        excel_strut_type = format_strut_type_for_excel(data.get("Strut Type"))
        
        # Loop through struts
        # Loop through struts
        for i in range(1, num_struts + 1):
            print(f"\nDEBUG: Processing strut {i}")
            excel_level_key = f"Strut{i}Level"
            excel_material_key = f"Strut{i}Material"
            excel_size_key = f"Strut{i}MemberSize"
            excel_space_key = f"Strut{i}Space"
            excel_length_key = f"Strut{i}Length"
            excel_angle_key = f"Strut{i}Angle"
            
            strut_level = strut_data.get(excel_level_key, "MISSING")
            strut_material = strut_data.get(excel_material_key, "Steel")
            strut_member_size = strut_data.get(excel_size_key, "N/A")
            strut_space = strut_data.get(excel_space_key, "")  # Get space for anchor properties
        
            # Handle Steel-specific validation
            if strut_material == "Steel":
                strut_member_size = strut_data.get(excel_size_key)
                if not strut_member_size or strut_member_size == "N/A":
                    raise ValueError(f"Member Size is required for Steel Strut {i}")
        
                plate_props = self.get_plate_properties(cursor, strut_member_size)
                if not plate_props.get("Elasticity") or not plate_props.get("EA"):
                    raise ValueError(
                        f"Missing properties for section {strut_member_size}.\n"
                        "Please ensure it exists in Steel_member_properties.xlsx with Elasticity and EA values."
                    )
        
            # Convert strut level to float
            strut_level_float = 0.0
            if strut_level is not None and strut_level != "MISSING":
                try:
                    strut_level_float = float(strut_level)
                    print(f"DEBUG: Converted strut level to float: {strut_level_float}")
                except ValueError:
                    print(f"Warning: Invalid strut level '{strut_level}' for strut {i}, using 0.0")
            prefixed_member_size = f"S{i}_{strut_member_size}" if strut_member_size else f"S{i}_N/A"
        
            # Initialize direction values
            direction_x = ""
            direction_y = ""
            
            print(f"DEBUG: Strut {i} - excel_strut_type = '{excel_strut_type}'")
            print(f"DEBUG: Strut {i} - Looking for {excel_length_key} in strut_data")
            print(f"DEBUG: Strut {i} - Looking for {excel_angle_key} in strut_data")
            print(f"DEBUG: Strut {i} - strut_data keys: {list(strut_data.keys())}")
            
        # FIX: Check for "fixedend" instead of "Fixed"
            if excel_strut_type == "fixedend":
                # Get per-strut length and angle
                strut_length = strut_data.get(excel_length_key, 0.0)
                strut_angle = strut_data.get(excel_angle_key, 0.0)
                
                # Convert to float safely
                try:
                    strut_length = float(strut_length) if strut_length else 0.0
                except (ValueError, TypeError):
                    strut_length = 0.0
                
                try:
                    strut_angle = float(strut_angle) if strut_angle else 0.0
                except (ValueError, TypeError):
                    strut_angle = 0.0
                
                print(f"DEBUG: Strut {i} - Length: {strut_length}, Angle: {strut_angle}")
                
                # Calculate directions
                direction_x = strut_length * math.cos(math.radians(strut_angle))
                direction_y = strut_length * math.sin(math.radians(strut_angle))
                
                print(f"DEBUG: Strut {i} - Direction X: {direction_x}, Direction Y: {direction_y}")
        
            # Create strut data record - use the same structure for both types
            strut_data_record = [
                prefixed_member_size,  
                f"strut_{i}",
                -(excavation_width / 2),
                strut_level_float,
                (excavation_width / 2),
                strut_level_float,
                excel_strut_type,  # This will be "fixedend" or "n2n"
                direction_x,
                direction_y
            ]
        
            print(f"DEBUG: Strut {i} record: {strut_data_record}")
            
            cursor.execute(
                f"INSERT INTO {sheets_config['Strut Details']['db_table']} "
                "(common_id, MaterialName, StrutName, x_Left, y_Left, x_Right, y_Right, Type, Direction_x, Direction_y) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                [common_id] + strut_data_record
            )
            
            excel_sheets["Strut Details"].append(strut_data_record)
            with open(sheets_config["Strut Details"]["csv_file"], 'a', newline='') as f:
               writer = csv.writer(f)
               writer.writerow([common_id] + strut_data_record)
                
            if strut_material == "Steel":
                print(f"DEBUG: Adding anchor properties for strut {i}")
    
                # Convert numeric values to appropriate types
                elasticity = plate_props.get("Elasticity", 200000)
                ea = float(plate_props.get("EA", 2.1e6))
                
                # Use per-strut space value for Lspacing
                spacing_value = float(strut_space) if strut_space else 0.0
                print(f"DEBUG: Strut {i} Space for anchor properties: {spacing_value}")
                
                colour = float(plate_props.get("Colour", "Blue"))
                prefixed_member_size_anchor = f"S{i}_{strut_member_size}" if strut_member_size else f"S{i}_Unknown_Steel"                
                
                anchor_data = [
                    prefixed_member_size_anchor or "Unknown_Steel",  
                    elasticity,
                    ea,
                    spacing_value,  # This now uses per-strut space value
                    colour
                ]
    
                print(f"DEBUG: Anchor data for strut {i}: {anchor_data}")
    
                cursor.execute(
                    f"INSERT INTO {sheets_config['Anchor Properties']['db_table']} "
                    "(common_id, MaterialName, Elasticity, EA, Lspacing, Colour) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    [common_id] + anchor_data
                )
    
                excel_sheets["Anchor Properties"].append(anchor_data)
    
                with open(sheets_config["Anchor Properties"]["csv_file"], 'a', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow([common_id] + anchor_data)
    
        self.line_load_handler.save_all_lineloads(
            cursor=cursor,
            data=data,
            common_id=common_id,
            sheets_config=sheets_config,
            excel_sheets=excel_sheets
        )
            
        workbook.save(excel_filename)
        print(f"Data saved successfully to MySQL, CSV files, and Excel: {excel_filename}")
        
      except Exception as ex:
        import traceback
        print(f"Error saving data: {str(ex)}")
        print(f"Detailed traceback: {traceback.format_exc()}")
        raise Exception(f"Error saving data: {str(ex)}")
    
    def _create_geometry_frames(self, fields: List[FormField], stored_data: Dict = None) -> List[ft.Container]:
      """Create geometry frames with proper visibility handling for imports."""
      # CRITICAL FIX: Update form_values and current_num_struts from stored_data if available
      if stored_data:
        # Update ALL form values including line load fields
        self.form_values.update(stored_data)
        print(f"DEBUG: Updated form_values with stored_data. Keys: {list(stored_data.keys())}")
        
        if "No of Strut" in stored_data:
            try:
                self.current_num_struts = int(stored_data["No of Strut"])
            except (ValueError, TypeError):
                self.current_num_struts = 0
    
    # Define base categories that don't depend on wall type
      excavation_fields = [
        "Excavation Type", "Wall Top Level", "Excavation Depth",
        "Excavation Width", "Toe Level", "Excavation Below Strut",
        "Over Excavation"
    ]
    
      strut_base_fields = ["No of Strut", "Strut Type"]
      borehole_fields = ["Borehole X Coordinate", "Ground Water Table"]
    
    # Get wall type from stored data or current section's fields
      wall_type = None
      if stored_data and "Wall Type" in stored_data:
        wall_type = stored_data.get("Wall Type")
      else:
        # Try to find Wall Type field in the current fields
        wall_type_field = next((f for f in fields if f.label == "Wall Type"), None)
        if wall_type_field:
            wall_type = wall_type_field.value
    
    # Get strut type from stored data or current section's fields
      strut_type = None
      if stored_data and "Strut Type" in stored_data:
        strut_type = stored_data.get("Strut Type")
      else:
        # Try to find Strut Type field in the current fields
        strut_type_field = next((f for f in fields if f.label == "Strut Type"), None)
        if strut_type_field:
            strut_type = strut_type_field.value

      wall_fields = ["Wall Type"]  # Always include Wall Type
    
      if wall_type == "Steel Pipe":
        # Steel Pipe specific fields
        wall_fields.extend(["Material", "Steel Grade", "Diameter", "Spacing", "Pipe Type"])  # ADDED "Steel Grade"
        # Add Grade if Pipe Type is "Filled with Concrete"
        pipe_type = stored_data.get("Pipe Type") if stored_data else self.form_values.get("Pipe Type")
        if pipe_type == "Filled with Concrete":
            wall_fields.append("Grade")
            
      elif wall_type == "Sheet Pile":
        wall_fields.extend(["Material", "Sheet Grade", "Member Size", "Connection Type"])  # ADDED "Sheet Grade"
        connection_type = stored_data.get("Connection Type") if stored_data else self.form_values.get("Connection Type")
        if connection_type == "Non Interlock":
            wall_fields.append("Spacing")
            
      elif wall_type == "Soldier Pile":
        material = stored_data.get("Material") if stored_data else self.form_values.get("Material")
        wall_fields.append("Material")
        
        if material == "Steel":
            wall_fields.extend(["Member Size", "Spacing"])
        elif material == "Concrete":
            wall_fields.extend(["Grade", "Spacing", "Shape"])
            shape = stored_data.get("Shape") if stored_data else self.form_values.get("Shape")
            if shape == "Rectangular":
                wall_fields.extend(["Width", "Depth"])
            elif shape == "Circular":
                wall_fields.append("Diameter")
                  
      elif wall_type in ["Contiguous Bored Pile", "Secant Bored Pile"]:
        wall_fields.extend(["Material", "Grade", "Diameter", "Spacing"])
        
      elif wall_type == "Diaphragm Wall":
        wall_fields.extend(["Material", "Grade", "Thickness"])
    
      elif stored_data:
        print("DEBUG: No wall type detected, including all available wall fields from stored data")
        for key in stored_data.keys():
            if key in ["Material", "Member Size", "Spacing", "Diameter", "Pipe Type", 
                      "Grade", "Connection Type", "Shape", "Width", "Depth", "Thickness"]:
                if key not in wall_fields:
                    wall_fields.append(key)
    
      print(f"DEBUG: Wall fields to display for '{wall_type}': {wall_fields}")
    
    # Comprehensive Strut Fields Selection
      strut_fields = [
        # Base fields
        "No of Strut",
        "Strut Type",
        
        # Common strut fields
        "Strut Material",
        "Strut Section",
        "Strut Length",
        "Strut Orientation",
        
        # Fixed Strut specific fields
        "Strut Space", 
        "Struct Length",
        "Angle"
    ]
    # Comprehensive Wall Fields Selection
      wall_field_names = [
        "Wall Type",
        "Material", 
        "Steel Grade",    # NEW
        "Sheet Grade",    # NEW
        "Member Size",
        "Diameter",
        "Spacing",
        "Pipe Type",
        "Grade",
        "Connection Type",
        "Shape",
        "Width",
        "Depth",
        "Thickness"
    ]
    # Filter fields that actually exist in the input
      available_strut_fields = [
        f.label for f in fields 
        if f.label in strut_fields
    ]
      available_wall_fields = [
        f.label for f in fields 
        if f.label in wall_field_names  # USE THE NEW LIST
     ]
    # Initialize frames list
      frames = []
    
    # 1. Excavation Details Frame
      excavation_controls = self._create_category_controls(
        [f for f in fields if f.label in excavation_fields],
        stored_data
    )
      excavation_frame = self._create_frame("Excavation Details", excavation_controls)
      frames.append(excavation_frame)
    
    # 2. Strut Details Frame
      strut_controls = self._create_category_controls(
        [f for f in fields if f.label in available_strut_fields],
        stored_data
    )
      strut_frame = self._create_frame("Strut Details", strut_controls)
      frames.append(strut_frame)
    
    # **CRITICAL FIX: 3. Line Load Details Frame - Pass stored_data to preserve values**
      print(f"DEBUG: Creating Line Load frame with stored_data: {stored_data is not None}")
      line_load_controls = self.line_load_handler.create_line_load_controls(self.form_values)  # Always use form_values
      line_load_frame = self._create_frame("Line Load Details", line_load_controls)
      frames.append(line_load_frame)
      
      # 4. Wall Details Frame
      wall_controls = self._create_category_controls(
        [f for f in fields if f.label in wall_fields],
        stored_data
    )
      wall_frame = self._create_frame("Wall Details", wall_controls)
      frames.append(wall_frame)
    
    # 5. Borehole Details Frame
      borehole_controls = self._create_category_controls(
        [f for f in fields if f.label in borehole_fields],
        stored_data
    )
      borehole_frame = self._create_frame("Borehole Details", borehole_controls)
      frames.append(borehole_frame)
    
    # **CRITICAL FIX**: Set frame visibility based on whether we have stored data (import mode)
      if stored_data:
        # **SHOW ALL FRAMES** when importing data - no continue buttons
        excavation_frame.visible = True
        strut_frame.visible = True
        line_load_frame.visible = True  # **ALWAYS SHOW IN IMPORT MODE**
        wall_frame.visible = True
        borehole_frame.visible = True
        show_buttons = False  # Don't show continue buttons during import
        print("DEBUG: Import mode - showing all frames, hiding continue buttons")
      else:
        # Default visibility logic for normal form flow
        excavation_frame.visible = True
        strut_frame.visible = bool(strut_type)
        line_load_frame.visible = bool(self.form_values.get("Excavation Type"))  # Show if excavation type selected
        wall_frame.visible = False
        borehole_frame.visible = False
        show_buttons = True  # Show continue buttons for normal flow
        print("DEBUG: Normal mode - using progressive frame display")
    
    # Create frame containers with continue buttons (only if not importing)
      frame_containers = []
      
      if show_buttons:
        # Normal flow: add frames with continue buttons
        for i, frame in enumerate(frames[:-1]):
            frame_containers.append(frame)
            
            validate_button = ft.ElevatedButton(
                text="Continue",
                on_click=lambda e, idx=i: self._validate_and_show_next_frame(e, idx),
                style=ft.ButtonStyle(
                    color=ft.colors.WHITE,
                    bgcolor=ft.colors.BLUE_600,
                    padding=10,
                )
            )
            
            button_container = ft.Container(
                content=validate_button,
                margin=ft.margin.only(bottom=20),
                visible=frame.visible  # Only visible if the frame is visible
            )
            frame_containers.append(button_container)
        
        # Add the last frame without a button
        frame_containers.append(frames[-1])
      else:
        # **IMPORT MODE**: Just return frames without buttons
        frame_containers = frames
        print(f"DEBUG: Import mode - returning {len(frames)} frames without buttons")
    
    # After creating the base frames, if there are struts, add the sub-frames to the Strut frame
      if self.current_num_struts > 0:
        # Find the Strut frame
        strut_frame = None
        for frame in frames:
            if (hasattr(frame, 'content') and 
                isinstance(frame.content, ft.Column) and
                frame.content.controls and 
                isinstance(frame.content.controls[0], ft.Text) and
                frame.content.controls[0].value == "Strut Details"):
                strut_frame = frame
                break
        
        if strut_frame:
            strut_frame_content = strut_frame.content  # This is a Column
            # Add sub-frames for each strut
            for i in range(1, self.current_num_struts + 1):
                strut_fields = self._get_strut_fields(i)
                # Create the controls for this strut, using stored_data to set values
                strut_controls = self._create_category_controls(strut_fields, stored_data)
                
                strut_sub_frame = ft.Container(
                    content=ft.Column([
                        ft.Text(f"Strut {i}", size=16, weight=ft.FontWeight.BOLD),
                        strut_controls
                    ]),
                    border=ft.border.all(1, ft.colors.GREY_300),
                    border_radius=8,
                    padding=15,
                    margin=ft.margin.only(top=10, bottom=10),
                    width=960
                )
                strut_frame_content.controls.append(strut_sub_frame)
    
      return frame_containers
  
    def _add_strut_subframes(self, num_struts: int, stored_data: Dict = None):
      """
      Helper method to add strut sub-frames to the Strut Details frame.
      This method is used both when building the UI from stored data and when 
      the user changes the number of struts.
      """
      # Find the Strut Details frame
      strut_frame = None
      for container in self.parent_form.form_content.controls:
        if (isinstance(container, ft.Container) and 
            hasattr(container, 'content') and 
            isinstance(container.content, ft.Column) and
            container.content.controls and 
            isinstance(container.content.controls[0], ft.Text) and
            container.content.controls[0].value == "Strut Details"):
            strut_frame = container
            break
            
      if not strut_frame:
        print("Error: Strut Details frame not found")
        return
        
    # Get the strut frame content
      strut_frame_content = strut_frame.content
    
    # Clear existing sub-frames (keep title and base controls)
    # We expect: [0] = title, [1] = base controls, [2+] = sub-frames
      if len(strut_frame_content.controls) > 2:
        strut_frame_content.controls = strut_frame_content.controls[:2]
        
    # Add new sub-frames for each strut
      for i in range(1, num_struts + 1):
        strut_fields = self._get_strut_fields(i)
        print(f"DEBUG: *** Creating Strut {i} fields including Space field ***")
    
        # Add change handlers to each field
        for field in strut_fields:
            field.set_on_change(self.handle_field_change)
            
        # Create the controls for this strut's fields
        strut_controls = self._create_category_controls(strut_fields, stored_data or self.form_values)
        
        # Create a sub-frame for this strut
        strut_sub_frame = ft.Container(
            content=ft.Column([
                ft.Text(f"Strut {i}", size=16, weight=ft.FontWeight.BOLD),
                strut_controls
            ]),
            border=ft.border.all(1, ft.colors.GREY_300),
            border_radius=8,
            padding=15,
            margin=ft.margin.only(top=10, bottom=10),
            width=960
        )
        strut_frame_content.controls.append(strut_sub_frame)
    
    # Update the strut frame
      strut_frame.content = strut_frame_content
      strut_frame.visible = True

 
    def _validate_and_show_next_frame(self, e, current_frame_index: int):
        """Validate current frame and show next frame if validation passes."""
        # Ensure we have access to form_content through parent_form
        if not self.parent_form or not self.parent_form.form_content:
            print("Error: Parent form or form content not initialized")
            return

        form_content = self.parent_form.form_content
        
        # Get all frames (excluding button containers)
        frames = [cont for cont in form_content.controls 
                 if isinstance(cont, ft.Container) 
                 and hasattr(cont, 'content') 
                 and isinstance(cont.content, ft.Column)]
        
        # Get all button containers
        buttons = [cont for cont in form_content.controls 
                  if isinstance(cont, ft.Container) 
                  and hasattr(cont, 'content') 
                  and isinstance(cont.content, ft.ElevatedButton)]
        
        # Validate current frame
        if self._validate_frame(frames[current_frame_index]):
            # Hide current frame's continue button
            if current_frame_index < len(buttons):
                buttons[current_frame_index].visible = False
                
            # Show next frame and its button
            if current_frame_index + 1 < len(frames):
                frames[current_frame_index + 1].visible = True
                if current_frame_index + 1 < len(buttons):
                    buttons[current_frame_index + 1].visible = True
            
            # Update the UI
            form_content.update()

    def _validate_frame(self, frame: ft.Container) -> bool:
     """Validate all required fields in the frame"""
     if not frame or not frame.content or not isinstance(frame.content, ft.Column):
        return False
        
    # Get the actual form controls column (second control after the title)
     form_column = frame.content.controls[1] if len(frame.content.controls) > 1 else None
     if not form_column or not isinstance(form_column, ft.Column):
        return False
    
    # Check all required fields have values
     for row in form_column.controls:
        if isinstance(row, ft.Row):
            for control in row.controls:
                if hasattr(control, 'required') and control.required:
                    if not control.value:
                        return False
     return True
 
    def _create_category_controls(self, fields: List[FormField], stored_data: Dict = None) -> ft.Column:
      form_rows = []
      for i in range(0, len(fields), 3):
        row_fields = []
        for j in range(3):
            if i + j < len(fields):
                field = fields[i + j]
                if field.field_type == "checkbox":
                    control = ft.Checkbox(
                        label=field.label,
                        value=field.value if field.value is not None else False,
                        on_change=field.on_change
                    )
                else:
                 control = field.create_control(
                     width=280,
                     on_change=(
                         self.handle_wall_type_change if field.label == "Wall Type" 
                         else self.handle_strut_type_change if field.label == "Strut Type"
                         else self.handle_pipe_type_change if field.label == "Pipe Type"
                         else self.handle_material_change if field.label == "Material"
                         else self.handle_shape_change if field.label == "Shape"
                         else self.handle_connection_type_change if field.label == "Connection Type"
                         else self.handle_strut_material_change if "Strut" in field.label and "Material" in field.label
                         else (field.on_change if hasattr(field, 'on_change') else None)
                     )
                 )
                if stored_data and field.label in stored_data:
                    control.value = stored_data[field.label]
                row_fields.append(control)
        form_rows.append(
            ft.Row(
                row_fields,
                alignment=ft.MainAxisAlignment.START,
                spacing=20
            )
        )
      return ft.Column(form_rows, spacing=20)
    
   
    def handle_wall_type_change(self, e):
      """Handle wall type change events - delegates to wall details handler"""
      self.wall_details_handler.handle_wall_type_change(e, self.parent_form)

    def handle_material_change(self, e):
      """Delegate material change to wall details handler"""
      self.wall_details_handler.handle_material_change(e, self.parent_form)

    def handle_shape_change(self, e):
      """Delegate shape change to wall details handler"""
      self.wall_details_handler.handle_shape_change(e, self.parent_form)

    def handle_pipe_type_change(self, e):
      """Delegate pipe type change to wall details handler"""
      self.wall_details_handler.handle_pipe_type_change(e, self.parent_form)

    def handle_connection_type_change(self, e):
      """Delegate connection type change to wall details handler"""
      self.wall_details_handler.handle_connection_type_change(e, self.parent_form)
  
    def handle_strut_type_change(self, e):
      """Handle strut type change and update UI with appropriate fields"""
      try:
        strut_type = e.data or e.control.value
        print(f"DEBUG: Strut type changed to: {strut_type}")
        
        # Update form values
        self.form_values["Strut Type"] = strut_type
        
        # Find the Strut Details frame
        strut_frame = None
        for container in self.parent_form.form_content.controls:
            if (hasattr(container, 'content') and 
                isinstance(container.content, ft.Column) and
                container.content.controls and 
                isinstance(container.content.controls[0], ft.Text) and
                container.content.controls[0].value == "Strut Details"):
                strut_frame = container
                break
        
        if not strut_frame:
            print("DEBUG: Strut Details frame not found")
            return
        
        # MODIFIED: Only include base fields in main strut section
        strut_base_fields = ["No of Strut", "Strut Type"]
        
        # Get the updated fields (without Strut Space, Struct Length, Angle in main section)
        all_fields = self.get_fields()
        base_strut_fields = [f for f in all_fields if f.label in strut_base_fields]
        base_strut_controls = self._create_category_controls(base_strut_fields, self.form_values)
        
        # Update the strut frame content
        # Preserve existing sub-frames if they exist
        existing_subframes = []
        if len(strut_frame.content.controls) > 2:
            existing_subframes = strut_frame.content.controls[2:]
        
        # Create new content with updated base controls
        new_content = [
            ft.Text("Strut Details", size=20, weight=ft.FontWeight.BOLD),
            base_strut_controls
        ]
        
        # Add back any existing sub-frames
        new_content.extend(existing_subframes)
        
        strut_frame.content.controls = new_content
        
        # MODIFIED: Update all strut sub-frames to include/exclude Fixed strut fields
        if self.current_num_struts > 0:
            self._update_strut_subframes()
            self._validate_all_strut_fields(strut_type)
        
        # Update the UI
        if self.parent_form and self.parent_form.page:
            self.parent_form.page.update()
        
        print(f"DEBUG: Successfully updated strut type to {strut_type}")
        
      except Exception as ex:
        print(f"ERROR in handle_strut_type_change: {str(ex)}")
        import traceback
        traceback.print_exc()
    def _validate_all_strut_fields(self, strut_type):
       """Validate all strut fields when strut type changes"""
       try:
        for i in range(1, self.current_num_struts + 1):
            # Validate Space field (required for both types)
            space_value = self.form_values.get(f"Strut {i} Space")
            space_error = self.validate_field(f"Strut {i} Space", space_value, self.form_values)
            
            # Find and update the Space field control
            space_control = self.find_field_control_recursive(f"Strut {i} Space")
            if space_control:
                if space_error:
                    space_control.border_color = "red"
                    space_control.hint_text = space_error
                else:
                    space_control.border_color = "black"
                    space_control.hint_text = ""
                space_control.update()
            
            # For Fixed struts, validate Length and Angle fields
            if strut_type == "Fixed":
                # Validate Length field
                length_value = self.form_values.get(f"Strut {i} Length")
                length_error = self.validate_field(f"Strut {i} Length", length_value, self.form_values)
                
                length_control = self.find_field_control_recursive(f"Strut {i} Length")
                if length_control:
                    if length_error:
                        length_control.border_color = "red"
                        length_control.hint_text = length_error
                    else:
                        length_control.border_color = "black"
                        length_control.hint_text = ""
                    length_control.update()
                
                # Validate Angle field
                angle_value = self.form_values.get(f"Strut {i} Angle")
                angle_error = self.validate_field(f"Strut {i} Angle", angle_value, self.form_values)
                
                angle_control = self.find_field_control_recursive(f"Strut {i} Angle")
                if angle_control:
                    if angle_error:
                        angle_control.border_color = "red"
                        angle_control.hint_text = angle_error
                    else:
                        angle_control.border_color = "black"
                        angle_control.hint_text = ""
                    angle_control.update()
        
       except Exception as ex:
        print(f"ERROR in _validate_all_strut_fields: {str(ex)}")
    def _update_strut_subframes(self):
      """Update all existing strut sub-frames based on current strut type"""
      try:
        # Find the Strut Details frame
        strut_frame = None
        for container in self.parent_form.form_content.controls:
            if (hasattr(container, 'content') and 
                isinstance(container.content, ft.Column) and
                container.content.controls and 
                isinstance(container.content.controls[0], ft.Text) and
                container.content.controls[0].value == "Strut Details"):
                strut_frame = container
                break
                
        if not strut_frame or len(strut_frame.content.controls) <= 2:
            return
            
        # Update each strut sub-frame
        for i in range(2, len(strut_frame.content.controls)):
            sub_frame = strut_frame.content.controls[i]
            if (isinstance(sub_frame, ft.Container) and
                isinstance(sub_frame.content, ft.Column) and
                sub_frame.content.controls and
                isinstance(sub_frame.content.controls[0], ft.Text)):
                
                # Extract strut number from title
                title = sub_frame.content.controls[0].value
                if title.startswith("Strut "):
                    try:
                        strut_num = int(title.split()[1])
                        print(f"DEBUG: Updating sub-frame for Strut {strut_num}")
                        
                        # Regenerate strut fields with current strut type
                        strut_fields = self._get_strut_fields(strut_num)
                        strut_controls = self._create_category_controls(strut_fields, self.form_values)
                        
                        # Replace the controls in the sub-frame
                        sub_frame.content.controls[1] = strut_controls
                        
                    except (ValueError, IndexError):
                        print(f"DEBUG: Could not extract strut number from title: {title}")
                        
      except Exception as ex:
        print(f"ERROR in _update_strut_subframes: {str(ex)}")
        import traceback
        traceback.print_exc()
   
    def _create_frame(self, title: str, content: ft.Column) -> ft.Container:
      frame = ft.Container(
        content=ft.Column([
            ft.Text(title, size=20, weight=ft.FontWeight.BOLD),
            content
        ]),
        border=ft.border.all(1, ft.colors.GREY_400),
        border_radius=10,
        padding=20,
        margin=ft.margin.only(bottom=20),
        width=1000
      )
    
    # Make all frames except first one initially invisible
      if title != "Excavation Details":
        frame.visible = False
    
      return frame
  
    
    def import_from_csv(self, csv_file_path: str, cursor) -> None:
      """Import geometry data from CSV file and populate all sections."""
      try:
        # Read CSV data
        with open(csv_file_path, mode='r') as file:
            reader = csv.DictReader(file)
            csv_data = [row for row in reader]
            if not csv_data:
                print("DEBUG: CSV file is empty or has no data rows")
                return
            csv_data = csv_data[0]  # Get first row
            print(f"DEBUG: CSV Data: {csv_data}")

        # Store strut-related fields separately for better handling
        strut_values = {}
        regular_values = {}
        
        # First pass to separate strut values from regular values
        for key, value in csv_data.items():
            if key.startswith("Strut ") and len(key.split()) >= 3:
                strut_values[key] = value
            else:
                regular_values[key] = value
                
        print(f"DEBUG: Regular Values: {regular_values}")
        print(f"DEBUG: Strut Values: {strut_values}")

        # Fix for the " Wall Top Level" field (has a leading space)
        if " Wall Top Level" in regular_values:
            regular_values["Wall Top Level"] = regular_values.pop(" Wall Top Level")
            print("DEBUG: Fixed Wall Top Level field name")

        # Update form values dictionary with regular values
        self.form_values.update(regular_values)
        
        # Process fields in correct section order
        section_order = [
            # 1. Excavation Details
            ['Excavation Type', 'Wall Top Level', 'Excavation Depth', 'Excavation Width', 
             'Toe Level', 'Excavation Below Strut', 'Over Excavation'],
            
            # 2. Strut Details - Only process the count here
            ['No of Strut', 'Strut Type', 'Strut Space'],
            
            # 3. Wall Details
            ['Wall Type', 'Material', 'Member Size', 'Spacing'],
            
            # 4. Borehole Details
            ['Borehole X Coordinate', 'Ground Water Table']
        ]
        
        # First pass: populate basic fields in each section
        for section_fields in section_order:
            for field in section_fields:
                if field in regular_values:
                    control = self.find_field_control(field)
                    if control:
                        control.value = regular_values[field]
                        control.update()
                        print(f"DEBUG: Updated field {field} with value {regular_values[field]}")
                    else:
                        print(f"DEBUG: Could not find control for field {field}")
        
        # Process critical fields that affect UI structure
        if "Excavation Type" in regular_values:
            excavation_type = regular_values["Excavation Type"]
            control = self.find_field_control("Excavation Type")
            if control:
                control.value = excavation_type
                # Trigger change event with page parameter
                event = ft.ControlEvent(
                    target="import",
                    name="excavation_type_change",
                    control=control,
                    data=excavation_type,
                    page=self.parent_form.page if hasattr(self.parent_form, 'page') else None
                )
                self.handle_excavation_type_change(event)
                
        # 1. First process Wall Type as it might affect other fields
        if "Wall Type" in regular_values:
            wall_type_field = self.find_field_control("Wall Type")
            if wall_type_field:
                try:
                    event = ft.ControlEvent(
                        target="import",
                        name="wall_type_change",
                        page=self.parent_form.page if hasattr(self.parent_form, 'page') else None,
                        control=wall_type_field,
                        data=regular_values["Wall Type"]
                    )
                    self.handle_wall_type_change(event)
                    print(f"DEBUG: Processed Wall Type change event with {regular_values['Wall Type']}")
                except Exception as e:
                    print(f"DEBUG: Error processing Wall Type: {e}")
        
        # Process strut type if present
        if "Strut Type" in regular_values:
            strut_type_field = self.find_field_control("Strut Type")
            if strut_type_field:
                try:
                    event = ft.ControlEvent(
                        target="import",
                        name="strut_type_change",
                        page=self.parent_form.page if hasattr(self.parent_form, 'page') else None,
                        control=strut_type_field,
                        data=regular_values["Strut Type"]
                    )
                    self.handle_strut_type_change(event)
                    print(f"DEBUG: Processed Strut Type change event with {regular_values['Strut Type']}")
                    
                    # Force an immediate UI update
                    if self.parent_form and self.parent_form.page:
                        self.parent_form.page.update()
                except Exception as e:
                    print(f"DEBUG: Error processing Strut Type: {e}")
        
        # 2. Process strut count next - this will create the strut UI elements
        if "No of Strut" in regular_values:
            strut_count = int(regular_values["No of Strut"])
            self.current_num_struts = strut_count
            print(f"DEBUG: Setting strut count to {strut_count}")
            
            # Find the strut count field control
            strut_count_field = self.find_field_control("No of Strut")
            print(f"DEBUG: Found strut count field: {strut_count_field}")
            
            if strut_count_field:
                try:
                    event = ft.ControlEvent(
                        target="import",
                        name="strut_change",
                        page=self.parent_form.page if hasattr(self.parent_form, 'page') else None,
                        control=strut_count_field,
                        data=str(strut_count)
                    )
                    print("DEBUG: Created strut change event")
                    self.handle_strut_change(event)
                    print("DEBUG: Handled strut change event")
                    
                    # Force an immediate UI update to ensure strut fields are created
                    if self.parent_form and self.parent_form.page:
                        self.parent_form.page.update()
                        print("DEBUG: Updated page after strut change")
                    
                    # Add a delay to ensure UI has time to render
                    import time
                    time.sleep(0.5)  # Slightly longer delay to ensure UI updates
                    print("DEBUG: Waited for UI to render")
                    
                    # Pre-load section details for all struts
                    self.section_details = self.load_section_details()
                    print("DEBUG: Preloaded section details for materials")
                    
                    # Update form_values with strut data from CSV before populating UI
                    for key, value in strut_values.items():
                        self.form_values[key] = value
                        print(f"DEBUG: Added {key}={value} to form_values")
                    
                    # Now populate each strut's fields
                    for strut_num in range(1, strut_count + 1):
                        print(f"DEBUG: Processing Strut {strut_num} fields")
                        
                        # Set material first (as it might affect available member sizes)
                        material_key = f"Strut {strut_num} Material"
                        material_field = self.find_field_control_recursive(material_key)
                        if material_field and material_key in strut_values:
                            print(f"DEBUG: Setting {material_key} = {strut_values[material_key]}")
                            material_field.value = strut_values[material_key]
                            material_field.update()
                            
                            # Trigger material change event to update available member sizes
                            material_event = ft.ControlEvent(
                                target="import",
                                name="strut_material_change",
                                page=self.parent_form.page if hasattr(self.parent_form, 'page') else None,
                                control=material_field,
                                data=strut_values[material_key]
                            )
                            self.handle_strut_material_change(material_event)
                            
                            # Force UI update after material change
                            if self.parent_form and self.parent_form.page:
                                self.parent_form.page.update()
                                time.sleep(0.2)  # Small delay for UI update
                        
                        # Process all strut fields for this strut
                        for field_type in ["Level", "Material", "Member Size", "Stiffness", "EA","Space"]:
                            field_key = f"Strut {strut_num} {field_type}"
                            
                            # Skip material as we've already processed it
                            if field_type == "Material":
                                continue
                                
                            if field_key in strut_values:
                                # Use recursive search for better field finding
                                control = self.find_field_control_recursive(field_key)
                                if control:
                                    print(f"DEBUG: Setting {field_key} = {strut_values[field_key]}")
                                    control.value = strut_values[field_key]
                                    control.update()
                                else:
                                    print(f"DEBUG: Could not find control for {field_key}")
                            else:
                                print(f"DEBUG: No CSV data for {field_key}")
                except Exception as e:
                    print(f"DEBUG: Error processing strut count: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print("DEBUG: Could not find strut count field control")
        
        # After successful import, rebuild the geometry section with all frames visible
        print("DEBUG: Rebuilding geometry frames for import mode")
        
        # Force show all sections with imported data
        self._show_all_sections()

        # Rebuild geometry section with imported data
        if self.parent_form:
            self.parent_form.update_form_content(1, self.form_values)  # 1=geometry tab index
            print("DEBUG: Updated form content with stored data")

        # Force UI refresh
        if self.parent_form and self.parent_form.page:
            self.parent_form.page.update()
            
        print("DEBUG: Geometry data imported and UI updated successfully!")
        
      except Exception as e:
        print(f"Error importing geometry data: {e}")
        import traceback
        traceback.print_exc()
    def find_field_control_recursive(self, field_name: str) -> Optional[ft.Control]:
      """Recursively search for a field control by name with better nested control handling."""
      print(f"DEBUG: Searching recursively for field: {field_name}")
    
      def search_control(control):
        # Check if this control is the one we're looking for
        if (isinstance(control, (ft.TextField, ft.Dropdown, ft.Checkbox)) and 
            hasattr(control, 'label') and control.label):
            clean_label = control.label.split(" (Set")[0].split(" *")[0]
            if clean_label == field_name:
                print(f"DEBUG: Found field {field_name}")
                return control
        
        # Search in container content
        if isinstance(control, ft.Container) and hasattr(control, 'content'):
            result = search_control(control.content)
            if result:
                return result
                
        # Search in columns
        if isinstance(control, ft.Column) and hasattr(control, 'controls'):
            for child in control.controls:
                result = search_control(child)
                if result:
                    return result
                    
        # Search in rows
        if isinstance(control, ft.Row) and hasattr(control, 'controls'):
            for child in control.controls:
                result = search_control(child)
                if result:
                    return result
        
        return None
    
    # Start search from form content
      if not self.parent_form or not hasattr(self.parent_form, 'form_content'):
        print("DEBUG: Parent form or form_content not available")
        return None
        
      for container in self.parent_form.form_content.controls:
        result = search_control(container)
        if result:
            return result
            
      print(f"DEBUG: Field {field_name} not found")
      return None
    def find_field_control(self, field_name: str) -> Optional[ft.Control]:
      if not self.parent_form:
        return None
    
    # Check all containers in the form content
      for container in self.parent_form.form_content.controls:
        # For containers with columns
        if (hasattr(container, 'content') and 
            isinstance(container.content, ft.Column)):
            # Check all controls in the column
            for control in container.content.controls:
                # For nested columns (like frame content)
                if isinstance(control, ft.Column):
                    for row in control.controls:
                        if isinstance(row, ft.Row):
                            for field in row.controls:
                                if (isinstance(field, (ft.TextField, ft.Dropdown, ft.Checkbox)) and 
                                    hasattr(field, 'label') and field.label):
                                    clean_label = field.label.split(" (Set")[0].split(" *")[0]
                                    if clean_label == field_name:
                                        return field
                # For direct rows
                elif isinstance(control, ft.Row):
                    for field in control.controls:
                        if (isinstance(field, (ft.TextField, ft.Dropdown, ft.Checkbox)) and 
                            hasattr(field, 'label') and field.label):
                            clean_label = field.label.split(" (Set")[0].split(" *")[0]
                            if clean_label == field_name:
                                return field
    
      return None
    def _show_all_sections(self):
      """Force visibility of all relevant sections based on data."""
      if not self.parent_form or not self.parent_form.form_content:
        print("DEBUG: Parent form or form content not available")
        return
        
    # **ENHANCED**: Determine which sections should be visible based on current data
      sections = {
        "Excavation Details": True,  # Always show
        "Strut Details": self.current_num_struts > 0 or "Strut Type" in self.form_values,
        "Wall Details": "Wall Type" in self.form_values,
        "Borehole Details": "Borehole X Coordinate" in self.form_values or "Ground Water Table" in self.form_values
    }
    
      frame_containers = []
      button_containers = []
    
    # Iterate through all controls in the form content
      for container in self.parent_form.form_content.controls:
        if not isinstance(container, ft.Container):
            continue
        
        # Check if this is a frame container (has a title)
        if (hasattr(container, 'content') and 
            isinstance(container.content, ft.Column) and 
            container.content.controls and 
            isinstance(container.content.controls[0], ft.Text)):
            
            title = container.content.controls[0].value
            if title in sections:
                # **FORCE VISIBILITY** for import mode
                container.visible = True
                frame_containers.append(container)
                print(f"DEBUG: Set {title} visibility to True (import mode)")
            else:
                print(f"DEBUG: Frame {title} not in sections list")
        
        # Check if this is a button container
        elif (hasattr(container, 'content') and 
              isinstance(container.content, ft.ElevatedButton)):
            button_containers.append(container)
    
    # **CRITICAL**: Hide all continue buttons when showing all sections (import mode)
      for button in button_containers:
        button.visible = False
        print("DEBUG: Hiding continue button in import mode")
    
    # Update UI
      if self.parent_form.form_content:
        self.parent_form.form_content.update()
        print("DEBUG: Updated form content visibility for import mode")
   
    def _populate_all_sections(self, form_values: Dict):
      """Populate all visible sections with imported data."""
      if not self.parent_form or not self.parent_form.form_content:
        return
        
    # Find all input controls in visible containers
      for container in self.parent_form.form_content.controls:
        if not isinstance(container, ft.Container) or not container.visible:
            continue
            
        if (hasattr(container, 'content') and 
            isinstance(container.content, ft.Column) and
            len(container.content.controls) > 1):
            self._populate_container_controls(container.content.controls[1], form_values)
    
    # Handle wall type and strut type changes after populating values
      if "Wall Type" in form_values:
        wall_type_field = self.find_field_control("Wall Type")
        if wall_type_field:
            self.handle_wall_type_change(ft.ControlEvent(control=wall_type_field, data=form_values["Wall Type"]))
    
      if "Strut Type" in form_values:
        strut_type_field = self.find_field_control("Strut Type")
        if strut_type_field:
            self.handle_strut_type_change(ft.ControlEvent(control=strut_type_field, data=form_values["Strut Type"]))
    
    # Update UI
      if self.parent_form and self.parent_form.page:
        self.parent_form.page.update()

    def _populate_container_controls(self, container, form_values: Dict):
      """Recursively populate controls within a container."""
      if isinstance(container, ft.Column):
        for control in container.controls:
            self._populate_container_controls(control, form_values)
      elif isinstance(container, ft.Row):
        for control in container.controls:
            if isinstance(control, (ft.TextField, ft.Dropdown)):
                clean_label = control.label.split(" (Set")[0].split(" *")[0]
                if clean_label in form_values:
                    control.value = form_values[clean_label]
                    # Clear error state
                    control.border_color = "black"
                    control.hint_text = ""
            elif isinstance(control, ft.Checkbox):
                clean_label = control.label.split(" (Set")[0].split(" *")[0]
                if clean_label in form_values:
                    control.value = form_values[clean_label]
            elif isinstance(control, ft.Container) and hasattr(control, 'content'):
                self._populate_container_controls(control.content, form_values)

    def _get_nested_controls(self, container):
      """Recursively get all controls in a container."""
      controls = []
      if hasattr(container, 'content'):
        content = container.content
        if isinstance(content, ft.Column) or isinstance(content, ft.Row):
            for item in content.controls:
                controls.extend(self._get_nested_controls(item))
        elif isinstance(content, (ft.TextField, ft.Dropdown, ft.Checkbox)):
            controls.append(content)
        else:
            controls.append(content)
      return controls

    def find_control_by_label(self, label: str):
      """Find control by label with deep search"""
      def search_controls(controls):
        for control in controls:
            if isinstance(control, ft.Row):
                for item in control.controls:
                    result = search_controls([item])
                    if result: return result
            elif isinstance(control, ft.Column):
                return search_controls(control.controls)
            elif (hasattr(control, 'label') and 
                 control.label.split(" (Set")[0].split(" *")[0] == label):
                return control
        return None
    
      if not self.parent_form:
        return None
        
      for container in self.parent_form.form_content.controls:
        if isinstance(container, ft.Container):
            result = search_controls(container.content.controls)
            if result: 
                return result
      return None
    def _populate_form_with_data(self, form_values: Dict):
      """Populate the form fields with the imported data."""
      try:
        if not self.parent_form or not self.parent_form.form_content:
            return

        # Process dynamic fields first (like number of struts)
        if "No of Strut" in form_values:
            self.current_num_struts = int(form_values["No of Strut"])
            self.handle_strut_change(ft.ControlEvent(data=str(self.current_num_struts)))

        # Iterate through all containers to find nested controls
        for container in self.parent_form.form_content.controls:
            if isinstance(container, ft.Container) and hasattr(container, 'content'):
                content = container.content
                if isinstance(content, ft.Column):
                    for control in content.controls:
                        # Handle frame titles and form columns
                        if isinstance(control, ft.Column):
                            for row in control.controls:
                                if isinstance(row, ft.Row):
                                    self._process_row(row, form_values)
                        elif isinstance(control, ft.Row):
                            self._process_row(control, form_values)

        # Force UI refresh
        self.parent_form.page.update()

      except Exception as e:
        print(f"Error populating form: {str(e)}")

    def _process_row(self, row: ft.Row, form_values: Dict):
      """Process a row of controls."""
      for control in row.controls:
        if isinstance(control, (ft.TextField, ft.Dropdown)):
            clean_label = control.label.split(" (Set")[0].split(" *")[0]
            if clean_label in form_values:
                control.value = form_values[clean_label]
                control.update()
        elif isinstance(control, ft.Container) and hasattr(control.content, 'controls'):
            for nested_control in control.content.controls:
                if isinstance(nested_control, ft.Row):
                    self._process_row(nested_control, form_values)