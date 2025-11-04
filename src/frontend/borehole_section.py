import csv
import logging
import os
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import flet as ft
import mysql.connector
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

from frontend.database_config import DatabaseConfig
from frontend.database_connection import DatabaseConnection
from frontend.database_operations import DatabaseOperations
from frontend.form_manager import FormManager
from frontend.form_section import FormField, FormSection


class BoreholeSection(FormSection):    
    def __init__(self, db_ops: DatabaseOperations, form_content: ft.Column, options: List[str] = None, form_manager=None):
        self.sets: List[Dict] = []
        self.material_names: List[str] = []          
        self.visible_sets = 0
        self.current_material_index = 0 
        self.selected_formation = None  # Add this to track selected formation
        self.db_ops = db_ops
        self.options = options or []
        self.form_content = form_content
        self.form_manager = form_manager
        # Check if running as executable
        if getattr(sys, 'frozen', False):
            # Running as exe - use internal/data directory
            self.BASE_DIR = Path(sys.executable).parent / "_internal"
        else:
            # Running as script - use original path
            self.BASE_DIR = Path(__file__).resolve().parent.parent.parent

        self.export_dir = self.BASE_DIR / "data"
        self.input_data_path = self.export_dir / "Input_Data.xlsx"
        self.formation_excel_path = self.export_dir / "Borehole_Formation.xlsx"  # Add formation excel path
        self.soil_db_path = form_manager.form_app.soil_db_path if form_manager and hasattr(form_manager, 'form_app') else self.export_dir / "Soil_DB.xlsx"
        self.soil_properties = {}
        self.load_soil_properties()
        print("DEBUG: Materials will be loaded on demand")
        columns = [ft.DataColumn(ft.Text("Actions", size=16, weight=ft.FontWeight.BOLD))]
        for field in self.get_fields():
            columns.append(ft.DataColumn(
                ft.Text(field.label, size=16, weight=ft.FontWeight.BOLD)
            ))
        self.data_table = ft.DataTable(
            columns=[ft.DataColumn(ft.Text(field.label)) for field in self.get_fields()],
            rows=[],
            border=ft.border.all(2, ft.Colors.GREY_400),
            border_radius=10,
            divider_thickness=2, 
            heading_row_height=50,
            data_row_min_height=80,
            data_row_max_height=80,
            column_spacing=20,
        )
        self.form_content.controls.append(self.data_table)

    def load_formation_soil_types(self, formation_name: str) -> List[str]:
        """Load soil types from the formation-specific Excel sheet"""
        try:
            print(f"DEBUG: Loading soil types for formation: {formation_name}")
            
            if not self.formation_excel_path.exists():
                print(f"DEBUG: Formation Excel file not found: {self.formation_excel_path}")
                return []
            
            workbook = openpyxl.load_workbook(self.formation_excel_path)
            print(f"DEBUG: Available sheets in formation workbook: {workbook.sheetnames}")
            
            if formation_name not in workbook.sheetnames:
                print(f"DEBUG: Formation sheet '{formation_name}' not found")
                return []
            
            sheet = workbook[formation_name]
            soil_types = []
            
            # Read soil types from column A, starting from row 2 (skip header)
            for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] and str(row[0]).strip():  # Check if cell has value and is not empty
                    soil_type = str(row[0]).strip()
                    if soil_type not in soil_types:  # Avoid duplicates
                        soil_types.append(soil_type)
                    print(f"DEBUG: Found formation soil type: {soil_type}")
            
            soil_types.sort()
            print(f"DEBUG: Total formation soil types found: {len(soil_types)}")
            return soil_types
            
        except Exception as e:
            print(f"ERROR: Loading formation soil types: {e}")
            import traceback
            traceback.print_exc()
            return []

    def set_selected_formation(self, formation_name: str):
        """Set the selected formation and reload material names"""
        print(f"DEBUG: Setting selected formation to: {formation_name}")
        self.selected_formation = formation_name
        
        # Load material names from the selected formation
        if formation_name:
            self.material_names = self.load_formation_soil_types(formation_name)
            print(f"DEBUG: Loaded {len(self.material_names)} soil types for formation {formation_name}")
            
            # Update existing dropdown options in the data table
            self.update_existing_soil_type_dropdowns()
        else:
            # Fallback to original material loading
            self.material_names = self.load_material_names()

    def update_existing_soil_type_dropdowns(self):
        """Update all existing Soil Type dropdowns with new formation-based options"""
        try:
            if not self.data_table or not self.data_table.rows:
                return
            
            for row in self.data_table.rows:
                if row.cells:
                    # Find the Soil Type cell (first cell)
                    soil_type_cell = row.cells[0]
                    if isinstance(soil_type_cell.content, ft.Dropdown):
                        # Update dropdown options
                        current_value = soil_type_cell.content.value
                        soil_type_cell.content.options = [
                            ft.dropdown.Option(mat) for mat in self.material_names
                        ]
                        
                        # Preserve current value if it exists in new options
                        if current_value in self.material_names:
                            soil_type_cell.content.value = current_value
                        else:
                            soil_type_cell.content.value = self.material_names[0] if self.material_names else ""
            
            self.data_table.update()
            print("DEBUG: Updated existing Soil Type dropdowns with formation-based options")
            
        except Exception as e:
            print(f"ERROR: Updating existing soil type dropdowns: {e}")

    def load_soil_properties(self):
        try:
            print(f"DEBUG: Loading soil properties from: {self.input_data_path}")
            if not self.input_data_path.exists():
                print(f"ERROR: File not found: {self.input_data_path}")
                return
            
            workbook = openpyxl.load_workbook(self.input_data_path)
            print(f"DEBUG: Available sheets in workbook: {workbook.sheetnames}")
            
            if 'Soil Properties' not in workbook.sheetnames:
                print(f"ERROR: 'Soil Properties' sheet not found in workbook.")
                self.soil_properties = {}
                return
            
            soil_sheet = workbook['Soil Properties']
            headers = [cell.value for cell in soil_sheet[1]]
            print(f"DEBUG: Headers in Soil Properties sheet: {headers}")
            
            self.soil_properties = {}
            for row in soil_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:      
                    material_name = row[0]
                    properties = {}
                    for i, header in enumerate(headers):
                        if header and i < len(row):
                            properties[header] = row[i]
                    self.soil_properties[material_name] = properties
                    print(f"DEBUG: Loaded properties for material: {material_name}")
            
            print(f"DEBUG: Total materials with properties: {len(self.soil_properties)}")
            if not self.soil_properties:
                print("WARNING: No soil properties loaded, creating default structure")
                self.soil_properties = {
                    "Default": {
                        "Drain Type": "Drain",
                        "SPT": 30,
                        "Gamma Unsat": 18,
                        "Gamma Sat": 20,
                        "E ref": 30000,
                        "Nu": 0.3,
                        "C '": 0,
                        "Phi '": 30,
                        "Kx": 0.01,
                        "Ky": 0.01,
                        "R inter": 0.67,
                        "K0 Primary": 0.5
                    }
                }
        except Exception as e:
            print(f"ERROR: Loading soil properties: {e}")
            import traceback
            traceback.print_exc()
            self.soil_properties = {
                "Default": {
                    "Drain Type": "Drain",
                    "SPT": 30,
                    "Gamma Unsat": 18,
                    "Gamma Sat": 20,
                    "E ref": 30000,
                    "Nu": 0.3,
                    "C '": 0,
                    "Phi '": 30,
                    "Kx": 0.01,
                    "Ky": 0.01,
                    "R inter": 0.67,
                    "K0 Primary": 0.5
                }
            }

    def load_material_names(self) -> List[str]:
        """Load material names from formation Excel if formation is selected, otherwise from Soil Properties"""
        if self.selected_formation:
            print(f"DEBUG: Loading material names from formation: {self.selected_formation}")
            return self.load_formation_soil_types(self.selected_formation)
        
        # Original logic for loading from Soil Properties sheet
        try:
            print(f"DEBUG: Looking for Excel file at: {self.input_data_path}")
            if not self.input_data_path.exists():
                print(f"ERROR: File not found: {self.input_data_path}")
                return []
            
            workbook = openpyxl.load_workbook(self.input_data_path)
            print(f"DEBUG: Available sheets in workbook: {workbook.sheetnames}")
            
            soil_sheet = workbook['Soil Properties']
            material_names = []
            for row in soil_sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0]:      
                    material_names.append(row[0])
                    print(f"DEBUG: Found material: {row[0]}")
            
            material_names.sort()
            print(f"DEBUG: Total materials found: {len(material_names)}")
            return material_names
        except FileNotFoundError:
            print(f"ERROR: Input data file not found at {self.input_data_path}")
            return []
        except KeyError as e:
            print(f"ERROR: Sheet 'Soil Properties' not found in the workbook. Available sheets: {e}")
            return []
        except Exception as e:
            print(f"ERROR: Loading material names from Excel: {e}")
            return []

    def get_fields(self) -> List[FormField]:
        return [
            FormField("Soil Type", "text", "e.g: Fill, F1"),
            FormField("Drain Type","dropdown",options=["Drain","Undrain"]),
            FormField("SPT", "number", "e.g: 30"),
            FormField("Top Depth", "number", "e.g: 0", required=True),
            FormField("Bottom Depth", "number", "e.g: 5", required=True),
            FormField("Gamma Unsat", "number", "e.g: 16"),
            FormField("Gamma Sat", "number", "e.g: 20"),
            FormField("E ref", "number", "e.g: 30000"),
            FormField("Nu", "number", "e.g: 0.3"),
            FormField("C '", "number", "e.g: 0"),
            FormField("Phi '", "number", "e.g: 30"),
            FormField("Kx", "number", "e.g: 0.01"),
            FormField("Ky", "number", "e.g: 0.01"),
            FormField("R inter", "number", "e.g: 0.67"),
            FormField("K0 Primary", "number", "e.g: 0.5")
        ]

    async def populate_from_ags_data(self, geology_data, borehole_id=None): 
      try:
        # If borehole_id is not provided but available in geology_data, use it
        if borehole_id is None and isinstance(geology_data, dict) and 'borehole_id' in geology_data:
            borehole_id = geology_data['borehole_id']
        
        print(f"Loading geological data from {geology_data} for borehole {borehole_id}")
        
        # Check if this is a new borehole_id and clear table if needed
        if hasattr(self, 'current_borehole_id') and self.current_borehole_id != borehole_id:
            print(f"New borehole detected: {borehole_id} (previous: {self.current_borehole_id}) - clearing table")
            if hasattr(self, 'data_table') and self.data_table:
                self.data_table.rows = []
        
        # Store the current borehole ID for reference
        self.current_borehole_id = borehole_id
        
        # If this is the first time loading data, also clear the table
        if not hasattr(self, 'ags_data_loaded') and hasattr(self, 'data_table') and self.data_table:
            print("First time loading AGS data - clearing table")
            self.data_table.rows = []
            self.ags_data_loaded = True
        
        processed_layers = []
        borehole_spt_data = []  # Store SPT data: (depth, value)
        
        # Also need to handle the case where geology_data is already a list
        # Sort the list by depth (DEEPEST FIRST) if it contains depth information
        if isinstance(geology_data, list):
            processed_layers = sorted(geology_data, key=lambda x: float(x.get('top', 0)) if x.get('top') and str(x.get('top')).replace('.', '').isdigit() else 0, reverse=True)  # FIXED: Deepest first
            print(f"Sorted input list by depth (deepest first)")
        elif isinstance(geology_data, dict):
            processed_layers = [geology_data]
        elif isinstance(geology_data, str) and os.path.exists(geology_data):
            # Process Excel file with proper AGS structure handling
            excel_data = pd.read_excel(geology_data, sheet_name=None)
            
            # Get GEOL data
            geol_df = None
            if 'GEOL' in excel_data:
                geol_df = excel_data['GEOL']
                print(f"Found GEOL sheet with columns: {geol_df.columns.tolist()}")
                
                # Filter out header rows (those starting with '<')
                geol_df = geol_df[~geol_df.iloc[:, 0].str.startswith('<', na=False)]
                
                # Filter for specific borehole if provided
                if borehole_id:
                    hole_id_col = next((col for col in geol_df.columns if 'HOLE_ID' in str(col).upper()), None)
                    if hole_id_col:
                        geol_df = geol_df[geol_df[hole_id_col] == borehole_id]
                        print(f"Filtered GEOL data for borehole {borehole_id}: found {len(geol_df)} rows")
                
                # Sort by GEOL_TOP in descending order (DEEPEST FIRST) BEFORE processing
                if not geol_df.empty:
                    geol_df = geol_df.sort_values('GEOL_TOP', ascending=False)  # FIXED: Deepest first
                    print(f"Sorted GEOL data by depth (deepest first): {geol_df['GEOL_TOP'].tolist()}")
                    
                    # Process GEOL data into layers (now in deepest-first order)
                    for _, row in geol_df.iterrows():
                        # FIXED: Correct depth assignments (already swapped in handle_borehole_selection)
                        geol_top = row.get('GEOL_TOP', 0)    # Top depth = GEOL_TOP
                        geol_base = row.get('GEOL_BASE', 0)  # Bottom depth = GEOL_BASE
                        geol_type = row.get('GEOL_GEOL', 'Unknown')
                        geol_desc = row.get('GEOL_DESC', '')
                        
                        # Clean up soil type
                        if geol_type and not pd.isna(geol_type):
                            geol_type = str(geol_type).replace('"', '').strip()
                        else:
                            geol_type = 'Unknown'
                        
                        # Handle depth values
                        try:
                            geol_top = float(geol_top) if geol_top and not pd.isna(geol_top) else 0
                        except (ValueError, TypeError):
                            geol_top = 0
                            
                        try:
                            geol_base = float(geol_base) if geol_base and not pd.isna(geol_base) else geol_top + 1
                        except (ValueError, TypeError):
                            geol_base = geol_top + 1
                        
                        layer_data = {
                            'soil_type': geol_type,
                            'top': geol_base,       # SWAPPED: Now using base as top
                            'base': geol_top,       # SWAPPED: Now using top as base
                            'description': str(geol_desc) if geol_desc and not pd.isna(geol_desc) else '',
                            'borehole_id': borehole_id
                        }
                        
                        processed_layers.append(layer_data)
            
            # Process ISPT sheet to extract SPT data and SORT BY DEPTH (DEEPEST FIRST)
            if 'ISPT' in excel_data:
                ispt_df = excel_data['ISPT']
                print(f"Found ISPT sheet with columns: {ispt_df.columns.tolist()}")
                
                # Filter out header rows (those starting with '<')
                ispt_df = ispt_df[~ispt_df.iloc[:, 0].str.startswith('<', na=False)]
                
                # Filter for specific borehole if provided
                if borehole_id:
                    hole_id_col = next((col for col in ispt_df.columns if 'HOLE_ID' in str(col).upper()), None)
                    if hole_id_col:
                        ispt_df = ispt_df[ispt_df[hole_id_col] == borehole_id]
                        print(f"Filtered ISPT data for borehole {borehole_id}: found {len(ispt_df)} rows")
                
                # Sort ISPT data by depth (DEEPEST FIRST) BEFORE processing
                top_col = next((col for col in ispt_df.columns if 'ISPT_TOP' in str(col).upper()), None)
                if top_col:
                    ispt_df = ispt_df.sort_values(top_col, ascending=False)  # FIXED: Deepest first
                    print(f"Sorted ISPT data by depth (deepest first): {ispt_df[top_col].tolist()}")
                
                # Extract SPT depths and values (now in deepest-first order)
                spt_col = next((col for col in ispt_df.columns if 'ISPT_NVAL' in str(col).upper()), None)
                
                if top_col and spt_col:
                    for _, row in ispt_df.iterrows():
                        try:
                            depth = float(row[top_col]) if row[top_col] and not pd.isna(row[top_col]) else None
                            spt_val = row[spt_col]
                            
                            if depth is not None:
                                # Handle complex SPT values (e.g., "100 / 200 mm", ">50", etc.)
                                if isinstance(spt_val, str):
                                    spt_val = spt_val.strip()
                                    # Handle refusal values like ">50"
                                    if spt_val.startswith('>'):
                                        spt_val = spt_val[1:].strip()
                                    # Handle fraction values like "100 / 200 mm"
                                    if '/' in spt_val:
                                        spt_val = spt_val.split('/')[0].strip()
                                    # Remove any non-numeric suffixes
                                    spt_val = spt_val.split()[0] if ' ' in spt_val else spt_val
                                    
                                try:
                                    spt_val = float(spt_val)
                                    if spt_val >= 0:  # Only accept non-negative SPT values
                                        borehole_spt_data.append((depth, spt_val))
                                        print(f"Added SPT data: depth={depth}m, value={spt_val}")
                                except (ValueError, TypeError):
                                    print(f"Could not parse SPT value: {row[spt_col]} at depth {depth}m")
                                    continue
                                    
                        except (ValueError, TypeError):
                            continue
                    
                    print(f"Extracted {len(borehole_spt_data)} valid SPT readings from ISPT sheet")
            
            # Match SPT values to geological layers based on depth ranges
            print("Matching SPT values to geological layers...")
            for layer in processed_layers:
                layer_top = layer.get('top', 0)
                layer_base = layer.get('base', layer_top + 1)
                spt_values = []
                spt_depths = []
                
                # Find SPT tests within this layer's depth range
                for depth, spt_val in borehole_spt_data:
                    if layer_top <= depth < layer_base:
                        spt_values.append(spt_val)
                        spt_depths.append(depth)
                
                # Calculate average SPT if we have values
                if spt_values:
                    average_spt = sum(spt_values) / len(spt_values)
                    layer['spt_value'] = round(average_spt, 1)  # Round to 1 decimal place
                    layer['spt_depths'] = spt_depths  # Store depths for reference
                    print(f"Layer {layer['soil_type']} ({layer_top}m-{layer_base}m): "
                          f"Found {len(spt_values)} SPT values, average = {average_spt:.1f}")
                else:
                    layer['spt_value'] = ''  # No SPT data available for this layer
                    print(f"Layer {layer['soil_type']} ({layer_top}m-{layer_base}m): No SPT data found")
        
        # Make sure we have soil properties loaded
        if not hasattr(self, 'soil_properties') or not self.soil_properties:
            self.load_soil_properties()
        
        # Populate table with processed data (now in deepest-first order - BOTTOM to TOP)
        print(f"Found {len(processed_layers)} geological layers to display for borehole {borehole_id}")
        
        for layer in processed_layers:
            if isinstance(layer, dict):
                # Extract soil type - prioritize 'soil_type' key, fallback to 'description'
                soil_type = layer.get('soil_type', layer.get('description', 'Unknown'))
                if not soil_type or pd.isna(soil_type):
                    soil_type = 'Unknown'
                else:
                    soil_type = str(soil_type).strip()
                
                # Extract depths (already swapped in handle_borehole_selection)
                top_depth = layer.get('top', 0)  # This is the deeper value (original base)
                if pd.isna(top_depth):
                    top_depth = 0
                    
                bottom_depth = layer.get('base', top_depth + 1)  # This is the shallower value (original top)
                if pd.isna(bottom_depth):
                    bottom_depth = top_depth + 1
                
                # Get soil properties for this soil type
                props = {}
                if soil_type in self.soil_properties:
                    props = self.soil_properties[soil_type]
                else:
                    # Try to find a partial match or use default
                    matching_key = next((key for key in self.soil_properties.keys() 
                                       if key.lower() in soil_type.lower() or soil_type.lower() in key.lower()), None)
                    if matching_key:
                        props = self.soil_properties[matching_key]
                        print(f"Using partial match '{matching_key}' for soil type '{soil_type}'")
                    else:
                        # Use first available properties as default
                        default_key = next(iter(self.soil_properties)) if self.soil_properties else None
                        if default_key:
                            props = self.soil_properties[default_key]
                            print(f"Using default soil properties '{default_key}' for type: {soil_type}")
                
                # Use calculated SPT if available, else fallback to properties
                spt_value = layer.get('spt_value')
                if spt_value is None or spt_value == '':
                    spt_value = props.get('SPT', "")
                
                # Create initial data for the row
                initial_data = {
                    'Soil Type': soil_type,
                    'Top Depth': top_depth,      # This is the deeper value (original base)
                    'Bottom Depth': bottom_depth, # This is the shallower value (original top)
                    'Drain Type': props.get('Drain Type', ""),
                    'SPT': spt_value,  # Use calculated or fallback value
                    'Gamma Unsat': props.get('Gamma Unsat', ""),
                    'Gamma Sat': props.get('Gamma Sat', ""),
                    'E ref': props.get('E ref', ""),
                    'Nu': props.get('Nu', ""),
                    'C \'': props.get('C \'', ""),
                    'Phi \'': props.get('Phi \'', ""),
                    'Kx': props.get('Kx', ""),
                    'Ky': props.get('Ky', ""),
                    'R inter': props.get('R inter', ""),
                    'K0 Primary': props.get('K0 Primary', ""),
                    'Borehole ID': borehole_id,
                    'Description': layer.get('description', ''),  # Store full description
                    'SPT Depths': layer.get('spt_depths', [])  # Store SPT test depths for reference
                }
                
                # Create and add the row
                new_row = self.create_borehole_row(
                  material_name="",  # Don't use this for soil type
                  initial_data=initial_data  # Soil type comes from here
            )
                self.data_table.rows.append(new_row)
            else:
                print(f"Skipping non-dictionary layer: {layer}")
        
        # Update the table display
        if hasattr(self, 'data_table') and self.data_table:
            self.data_table.update()
        
        print(f"Successfully populated {len(processed_layers)} rows for borehole {borehole_id} (deepest to shallowest order)")
        if borehole_spt_data:
            print(f"Applied SPT data from {len(borehole_spt_data)} test readings")
          
      except Exception as e:
        print(f"Error populating borehole data: {str(e)}")
        import traceback
        traceback.print_exc() 
    def validate(self, data: List[Dict]) -> List[str]:
      """Validate borehole data entries."""
      errors = []
      for i, borehole_set in enumerate(data, 1):
        # Check required numeric fields
        if borehole_set.get("Top Depth") is None:
            errors.append(f"Top Depth is required for Set {i}")
        if borehole_set.get("Bottom Depth") is None:
            errors.append(f"Bottom Depth is required for Set {i}")
        
        # Validate depth relationship
        try:
            top = float(borehole_set.get("Top Depth", 0))
            bottom = float(borehole_set.get("Bottom Depth", 0))
            if bottom >= top:
                errors.append(
                    f"Set {i}: Bottom Depth ({bottom}) must be less than Top Depth ({top})"
                )
        except ValueError:
            errors.append(f"Set {i}: Invalid depth values")
      return errors

    def save(self, cursor, data: List[Dict]) -> None:
        """Save borehole data using DatabaseOperations."""
        if data:
            common_id = data[0].get('common_id')
            self.db_ops.save_borehole_data(cursor, data, common_id)
            # Get the base directory of the script

            # Define the export directory relative to the project root
            export_dir = self.export_dir
            # Save to files
            self.db_ops.save_to_csv(data, export_dir/"borehole_data.csv", data[0].keys())
            self.db_ops.update_excel(export_dir/"Input_Data.xlsx", "Borehole", data)

    def on_top_depth_change(self, e):
      """Handle top depth changes and update previous row's bottom depth"""
      try:
        if not hasattr(self, 'data_table') or not self.data_table or not self.data_table.rows:
            return
        
        # Find which row contains the changed field
        current_row_index = None
        for i, row in enumerate(self.data_table.rows):
            # Skip the first cell (actions) when searching for the changed field
            for cell in row.cells[1:]:
                if isinstance(cell.content, ft.TextField) and cell.content == e.control:
                    current_row_index = i
                    break
            if current_row_index is not None:
                break

        if current_row_index is None:
            return

        new_top_depth = e.control.value
        previous_row_index = current_row_index - 1
        
        # Update previous row's bottom depth if it exists
        if previous_row_index >= 0:
            previous_row = self.data_table.rows[previous_row_index]
            # Bottom depth is at index 5 (accounting for actions column)
            if len(previous_row.cells) > 5:
                bottom_depth_cell = previous_row.cells[5]
                if isinstance(bottom_depth_cell.content, ft.TextField):
                    bottom_depth_cell.content.value = new_top_depth
                    print(f"DEBUG: Updated previous row's bottom depth to: {new_top_depth}")

        self.data_table.update()
        
      except Exception as e:
        print(f"ERROR: Updating previous row's bottom depth: {e}")


# Modify the create_borehole_row method - Replace the Top Depth field creation section
    def create_borehole_row(self, material_name, initial_data=None, row_index=None):
      """Create a new borehole row with proper action buttons and input controls"""
      print(f"\n=== Creating borehole row at index {row_index} ===")
      print(f"DEBUG: material_name parameter: '{material_name}'")
      print(f"DEBUG: initial_data: {initial_data}")

      if not self.material_names:
        self.material_names = self.load_material_names()
        print(f"DEBUG: Loaded material names on demand: {len(self.material_names)} materials")

    # Get geometry data for wall top level
      geometry_data = self.form_manager.get_section_data('geometry')
      wall_top_level = geometry_data.get('Wall Top Level', None) if geometry_data else None

    # Check if this is the first real row
      is_first_real_row = not any(
        any(hasattr(cell.content, 'value') and cell.content.value != "" 
            for cell in row.cells[1:] if hasattr(cell.content, 'value'))
        for row in self.data_table.rows
      )

    # Get previous bottom depth for auto-filling top depth
      previous_bottom_depth = None
      if self.data_table.rows:
        last_row = self.data_table.rows[-1]
        if len(last_row.cells) > 5:
            bottom_depth_cell = last_row.cells[5]
            if isinstance(bottom_depth_cell.content, ft.TextField):
                previous_bottom_depth = bottom_depth_cell.content.value

    # Ensure row_index is properly set
      if row_index is None:
        row_index = len(self.data_table.rows)
        print(f"DEBUG: row_index was None, set to {row_index}")

    # Create action buttons
      action_buttons = ft.Row(
        controls=[
            ft.IconButton(
                icon=ft.icons.REMOVE,
                icon_size=16,
                tooltip="Delete this row",
                on_click=lambda e, idx=row_index: self.delete_row(e, idx),
                style=ft.ButtonStyle(
                    shape=ft.CircleBorder(),
                    padding=5,
                    bgcolor=ft.colors.RED_300
                )
            )
        ],
        spacing=5,
        alignment=ft.MainAxisAlignment.CENTER
    )

    # Start cells with action buttons
      cells = [ft.DataCell(action_buttons)]

    # Create input field cells
      fields = self.get_fields()
      for field in fields:
        field_label = field.label
        value = ""
        read_only = False

        # Set initial values from data or defaults
        if initial_data and field_label in initial_data:
            value = str(initial_data[field_label])
            print(f"Using initial data for {field_label}: {value}")
        else:
            if "Top Depth" in field_label:
                if is_first_real_row and wall_top_level:
                    value = str(wall_top_level)
                    read_only = True  # First row's top depth is read-only
                elif previous_bottom_depth:
                    value = previous_bottom_depth
                    # NOT read-only anymore - user can edit

        # Create appropriate control based on field type
        if "Soil Type" in field_label:
            # Soil Type dropdown logic (unchanged)
            actual_soil_type = None
            if initial_data and 'Soil Type' in initial_data:
                actual_soil_type = str(initial_data['Soil Type']).strip()
            else:
                actual_soil_type = material_name
            
            dropdown_options = self.material_names if self.material_names else ["Default"]
            
            if actual_soil_type and actual_soil_type not in dropdown_options:
                dropdown_options.append(actual_soil_type)
            
            control = ft.Dropdown(
                label=field_label,
                options=[ft.dropdown.Option(mat) for mat in dropdown_options],
                value=actual_soil_type,
                border=ft.InputBorder.UNDERLINE,
                filled=True,
                expand=True,
                hint_text="Select soil type"
            )
            
        elif "Drain Type" in field_label:
            control = ft.Dropdown(
                label=field_label,
                options=[ft.dropdown.Option("Drain"), ft.dropdown.Option("Undrain")],
                value=value if value else "",
                border=ft.InputBorder.UNDERLINE,
                filled=True,
                expand=True
            )
        else:
            # MODIFIED SECTION: Handle both Top Depth and Bottom Depth
            if "Top Depth" in field_label:
                control = ft.TextField(
                    value=value,
                    label=field_label,
                    label_style=ft.TextStyle(color=ft.colors.TRANSPARENT, size=0),
                    border=ft.InputBorder.UNDERLINE,
                    filled=True,
                    read_only=read_only,  # Only first row is read-only
                    expand=True,
                    on_change=self.on_top_depth_change  # NEW: Add handler
                )
            elif "Bottom Depth" in field_label:
                control = ft.TextField(
                    value=value,
                    label=field_label,
                    label_style=ft.TextStyle(color=ft.colors.TRANSPARENT, size=0),
                    border=ft.InputBorder.UNDERLINE,
                    filled=True,
                    read_only=False,
                    expand=True,
                    on_change=self.on_bottom_depth_change  # Existing handler
                )
            else:
                control = ft.TextField(
                    value=value,
                    label=field_label,
                    label_style=ft.TextStyle(color=ft.colors.TRANSPARENT, size=0),
                    border=ft.InputBorder.UNDERLINE,
                    filled=True,
                    read_only=read_only,
                    expand=True
                )

        cells.append(ft.DataCell(control))

      print("=== Row creation complete ===")
      return ft.DataRow(cells=cells)


    # Also update the delete_row method to handle the new bidirectional updates
    def delete_row(self, e, row_index):
      """Delete the row at the specified index and adjust depths"""
      try:
        if row_index is None:
            print("DEBUG: Cannot delete - row index is None")
            return
            
        if not isinstance(row_index, int) or row_index < 0:
            print(f"DEBUG: Invalid row_index: {row_index}")
            return
            
        if row_index >= len(self.data_table.rows):
            print(f"DEBUG: row_index {row_index} is out of bounds")
            return
            
        print(f"DEBUG: Deleting row at index {row_index}")
        
        if len(self.data_table.rows) <= 1:
            print("DEBUG: Cannot delete the last row")
            return
            
        # Get the top depth of the deleted row
        deleted_row = self.data_table.rows[row_index]
        deleted_top_depth = None
        if len(deleted_row.cells) > 4:  # Top Depth is at index 4
            top_depth_cell = deleted_row.cells[4]
            if isinstance(top_depth_cell.content, ft.TextField):
                deleted_top_depth = top_depth_cell.content.value
        
        # Remove the row
        self.data_table.rows.pop(row_index)
        
        # Update counters
        self.visible_sets = max(0, self.visible_sets - 1)
        
        # Re-index all rows
        self.reindex_rows()
        
        # Update the next row's top depth to match deleted row's top depth
        if deleted_top_depth is not None and row_index < len(self.data_table.rows):
            next_row = self.data_table.rows[row_index]
            if len(next_row.cells) > 4:
                top_depth_cell = next_row.cells[4]
                if isinstance(top_depth_cell.content, ft.TextField):
                    top_depth_cell.content.value = deleted_top_depth
                    print(f"DEBUG: Updated next row's top depth to: {deleted_top_depth}")
        
        self.data_table.update()
        self.update_delete_button_state()
        print(f"DEBUG: Successfully deleted row at index {row_index}")
        
      except Exception as ex:
        print(f"ERROR: Deleting row: {ex}")
        import traceback
        traceback.print_exc()
  
    def add_row_above(self, e, row_index):
        """Add a new row above the specified index"""
        try:
            print(f"DEBUG: Adding row above index {row_index}")
            
            if not self.material_names:
                self.material_names = self.load_material_names()

            material_index = self.current_material_index % max(1, len(self.material_names))
            material_name = self.material_names[material_index] if self.material_names else "New Material"

            new_row = self.create_borehole_row(
                material_name=material_name,
                row_index=row_index
            )

            self.data_table.rows.insert(row_index, new_row)
            self.visible_sets += 1
            self.current_material_index += 1

            # Re-index all rows to fix button callbacks
            self.reindex_rows()
            self.data_table.update()
            self.update_delete_button_state()
            
            print(f"DEBUG: Successfully added row above index {row_index}")
            
        except Exception as ex:
            print(f"ERROR: Adding row above: {ex}")
            import traceback
            traceback.print_exc()
    def reindex_rows(self):
        """Re-index all rows to ensure proper button callbacks after insertions/deletions"""
        try:
            for idx, row in enumerate(self.data_table.rows):
                if row.cells and isinstance(row.cells[0].content, ft.Row):
                    action_buttons = row.cells[0].content
                    for button in action_buttons.controls:
                        if hasattr(button, 'on_click') and button.on_click:
                            if button.icon == ft.icons.ADD:
                                button.on_click = lambda e, idx=idx: self.add_row_above(e, idx)
                            elif button.icon == ft.icons.REMOVE:
                                button.on_click = lambda e, idx=idx: self.delete_row(e, idx)
            
            print(f"DEBUG: Reindexed {len(self.data_table.rows)} rows")
            
        except Exception as ex:
            print(f"ERROR: Reindexing rows: {ex}")
            import traceback
            traceback.print_exc()
    def add_borehole_set(self, e):
        """Add a new borehole set to the end of the table"""
        print(f"DEBUG: Starting add_borehole_set.")
        
        if not self.material_names:
            self.material_names = self.load_material_names()
            print(f"DEBUG: Loaded material names on demand: {len(self.material_names)} materials")

        material_index = self.current_material_index % max(1, len(self.material_names))
        material_name = self.material_names[material_index] if self.material_names else "Unknown"

        new_row = self.create_borehole_row(
            material_name=material_name,
            row_index=len(self.data_table.rows)
        )

        self.data_table.rows.append(new_row)
        self.visible_sets += 1
        self.current_material_index += 1

        self.data_table.update()
        self.update_delete_button_state() 
    def on_bottom_depth_change(self, e):
        try:
            if not hasattr(self, 'data_table') or not self.data_table or not self.data_table.rows:
                return
            
            current_row_index = None
            for i, row in enumerate(self.data_table.rows):
                # Skip the first cell (actions) when searching for the changed field
                for cell in row.cells[1:]:
                    if isinstance(cell.content, ft.TextField) and cell.content == e.control:
                        current_row_index = i
                        break
                if current_row_index is not None:
                    break

            if current_row_index is None:
                return

            new_bottom_depth = e.control.value
            next_row_index = current_row_index + 1
            
            if next_row_index < len(self.data_table.rows):
                next_row = self.data_table.rows[next_row_index]
                # Skip the first cell (actions) and look for top depth field (should be at index 4 now)
                if len(next_row.cells) > 4:
                    top_depth_cell = next_row.cells[4]  # Adjusted for actions column
                    if isinstance(top_depth_cell.content, ft.TextField):
                        top_depth_cell.content.value = new_bottom_depth
                        top_depth_cell.content.read_only = True
                        print(f"DEBUG: Updated next row's top depth to: {new_bottom_depth}")

            self.data_table.update()
            
        except Exception as e:
            print(f"ERROR: Updating next row's top depth: {e}")

    def delete_last_row(self, e):
        """Delete the last row in the table"""
        try:
            print("DEBUG: Attempting to delete last row")
            
            if not self.data_table or not self.data_table.rows:
                print("DEBUG: No rows to delete")
                return

            if len(self.data_table.rows) <= 1:
                print("DEBUG: Cannot delete - only one row remaining")
                return

            deleted_row = self.data_table.rows.pop()
            print(f"DEBUG: Deleted row with {len(deleted_row.cells)} cells")
            
            self.visible_sets = max(0, self.visible_sets - 1)
            self.current_material_index = max(0, self.current_material_index - 1)

            self.reindex_rows()
            self.data_table.update()
            self.update_delete_button_state()
            
        except Exception as ex:
            print(f"ERROR: Deleting last row: {ex}")
            import traceback
            traceback.print_exc()

    def update_delete_button_state(self):
        """Update the delete button state based on row count"""
        try:
            if hasattr(self, 'delete_button') and self.delete_button:
                row_count = len(self.data_table.rows) if self.data_table and self.data_table.rows else 0
                should_disable = row_count <= 1
                self.delete_button.disabled = should_disable
                self.delete_button.update()
                print(f"DEBUG: Delete button state updated - rows: {row_count}, disabled: {should_disable}")
                
        except Exception as ex:
            print(f"ERROR: Updating delete button state: {ex}")

    def has_data_in_table(self):
        """Check if the data table has any meaningful data (safe way to check values)"""
        try:
            if not hasattr(self, 'data_table') or not self.data_table or not self.data_table.rows:
                return False
                
            for row in self.data_table.rows:
                # Skip the first cell (action buttons) and check input controls only
                for cell in row.cells[1:]:  # Skip first cell with action buttons
                    if hasattr(cell.content, 'value') and cell.content.value:
                        if isinstance(cell.content, (ft.TextField, ft.Dropdown)):
                            if cell.content.value and str(cell.content.value).strip():
                                return True
            return False
            
        except Exception as ex:
            print(f"ERROR: Checking table data: {ex}")
            return False

    # Updated method for the form manager to use (replaces the problematic condition)
    def should_add_initial_row(self, stored_data):
        """Determine if an initial row should be added to the table"""
        # Don't add if we have stored data
        if stored_data:
            return False
            
        # Don't add if table already has rows
        if hasattr(self, 'data_table') and self.data_table and len(self.data_table.rows) > 0:
            return False
            
        # Add initial row if table is empty
        return True
    def import_from_csv(self, csv_file_path: str, cursor) -> None:
        """Import borehole data from CSV file and populate the DataTable."""
        try:
            # Load material names if needed for import
            if not self.material_names:
                self.material_names = self.load_material_names()
                print(f"DEBUG: Loaded material names for CSV import: {len(self.material_names)} materials")
                
            # Read CSV file
            df = pd.read_csv(csv_file_path)
            
            # Define column mappings based on actual database schema
            column_mappings = {
                'TopDepth': ['Top Depth', 'TopDepth', 'Top_Depth', 'top_depth', 'topdepth', 'Depth From', 'DepthFrom'],
                'BottomDepth': ['Bottom Depth', 'BottomDepth', 'Bottom_Depth', 'bottom_depth', 'bottomdepth', 'Depth To', 'DepthTo'],
                'SoilType': ['Soil Type', 'SoilType', 'Soil_Type', 'soil_type', 'Material'],
                'SPT': ['SPT', 'spt', 'SPT_N'],
                'gammaUnsat': ['Gamma Unsat', 'gammaUnsat', 'Gamma_Unsat', 'UnitWeightUnsat'],
                'gammaSat': ['Gamma Sat', 'gammaSat', 'Gamma_Sat', 'UnitWeightSat'],
                'Eref': ['E ref', 'Eref', 'E_ref', 'YoungModulus'],
                'nu': ['Nu', 'nu', 'Poisson'],
                'cref': ['C \'', 'cref', 'Cohesion'],
                'phi': ['Phi \'', 'phi', 'FrictionAngle'],
                'kx': ['Kx', 'kx', 'HorizontalPermeability'],
                'ky': ['Ky', 'ky', 'VerticalPermeability'],
                'Rinter': ['R inter', 'Rinter', 'InterfaceStrength'],
                'K0Primary': ['K0 Primary', 'K0Primary', 'K0'],
                'DrainType': ['Drain Type', 'DrainType', 'Drainage']
            }
            
            # Find and map the actual column names in the CSV
            mapped_columns = {}
            for db_col, possible_names in column_mappings.items():
                found = False
                for name in possible_names:
                    if name in df.columns:
                        mapped_columns[db_col] = name
                        found = True
                        break
            
            # Validate required fields
            required_fields = ['TopDepth', 'BottomDepth']
            for field in required_fields:
                if field not in mapped_columns:
                    raise ValueError(f"Could not find a column matching {field}")
            
            # Convert numeric columns to appropriate data types
            numeric_cols = ["TopDepth", "BottomDepth", "SPT", "gammaUnsat", 
                           "gammaSat", "Eref", "nu", "cref", "phi", 
                           "kx", "ky", "Rinter", "K0Primary"]
            
            for col in numeric_cols:
                if col in mapped_columns:
                    col_name = mapped_columns[col]
                    df[col_name] = pd.to_numeric(df[col_name], errors='coerce')
            
            # Validate data
            validation_errors = []
            for index, row in df.iterrows():
                set_number = index + 1
                
                # Check required fields
                for field in required_fields:
                    csv_col = mapped_columns[field]
                    if pd.isna(row[csv_col]):
                        validation_errors.append(f"{field} is required for Set {set_number}")
                
                # Validate depth values
                if not validation_errors:
                    try:
                        top_depth = float(row[mapped_columns['TopDepth']])
                        bottom_depth = float(row[mapped_columns['BottomDepth']])
                        
                        if bottom_depth >= top_depth:
                            validation_errors.append(
                                f"Set {set_number}: Bottom Depth ({bottom_depth}) must be less than Top Depth ({top_depth})"
                            )
                    except ValueError:
                        validation_errors.append(f"Set {set_number}: Depth values must be numbers")
            
            if validation_errors:
                raise ValueError("Validation errors in CSV data: " + "; ".join(validation_errors))
            
            # Clear existing rows in the DataTable
            self.data_table.rows.clear()
            
            # Reset counters since we're loading from CSV
            self.current_material_index = 0
            self.visible_sets = 0

            # Populate the DataTable with CSV data
            for index, row in df.iterrows():
                # Prepare data for the DataTable
                initial_data = {
                    "Soil Type": row[mapped_columns['SoilType']],
                    "Drain Type": row.get(mapped_columns.get('DrainType', ''), ""),
                    "SPT": row.get(mapped_columns.get('SPT', ''), ""),
                    "Top Depth": row[mapped_columns['TopDepth']],
                    "Bottom Depth": row[mapped_columns['BottomDepth']],
                    "Gamma Unsat": row.get(mapped_columns.get('gammaUnsat', ''), ""),
                    "Gamma Sat": row.get(mapped_columns.get('gammaSat', ''), ""),
                    "E ref": row.get(mapped_columns.get('Eref', ''), ""),
                    "Nu": row.get(mapped_columns.get('nu', ''), ""),
                    "C '": row.get(mapped_columns.get('cref', ''), ""),
                    "Phi '": row.get(mapped_columns.get('phi', ''), ""),
                    "Kx": row.get(mapped_columns.get('kx', ''), ""),
                    "Ky": row.get(mapped_columns.get('ky', ''), ""),
                    "R inter": row.get(mapped_columns.get('Rinter', ''), ""),
                    "K0 Primary": row.get(mapped_columns.get('K0Primary', ''), "")
                }

                # Create a new row and add it to the DataTable
                new_row = self.create_borehole_row(initial_data["Soil Type"], initial_data)
                self.data_table.rows.append(new_row)
                
                # Update counters
                self.visible_sets += 1
                self.current_material_index += 1

            # Update the UI
            self.data_table.update()
            print("CSV data imported and DataTable updated successfully!")
            
        except pd.errors.EmptyDataError:
            raise ValueError("The CSV file is empty")
        except FileNotFoundError:
            raise ValueError("CSV file not found")
        except Exception as e:
            raise ValueError(f"Error processing CSV: {str(e)}")
    
    
    # Add these methods to your BoreholeSection class
    def create_formation_excel(self, formation_name, soil_type):
      try:
        excel_path = self.export_dir / "Borehole_Formation.xlsx"
        if excel_path.exists():
            workbook = openpyxl.load_workbook(excel_path)
        else:
            workbook = openpyxl.Workbook()
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])

        if formation_name in workbook.sheetnames:
            sheet = workbook[formation_name]
        else:
            sheet = workbook.create_sheet(formation_name)
            # Create headers for all columns (SPT removed)
            headers = [
                'Soil Type', 'Drain Type', 'Gamma Unsat', 'Gamma Sat', 
                'E ref', 'Nu', 'C \'', 'Phi \'', 'Kx', 'Ky', 
                'R inter', 'K0 Primary'
            ]
            
            # Set headers in row 1
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
            
            # Optional: Set column widths for better readability
            column_widths = {
                'A': 15,  # Soil Type
                'B': 12,  # Drain Type
                'C': 12,  # Gamma Unsat
                'D': 12,  # Gamma Sat
                'E': 10,  # E ref
                'F': 8,   # Nu
                'G': 8,   # C '
                'H': 8,   # Phi '
                'I': 8,   # Kx
                'J': 8,   # Ky
                'K': 10,  # R inter
                'L': 12   # K0 Primary
            }
            
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

        # Find the next empty row
        next_row = sheet.max_row + 1 if sheet.max_row > 1 else 2
        
        # Add the soil type to the first column (Soil Type)
        sheet[f'A{next_row}'] = soil_type
       
        workbook.save(excel_path)
        print(f"Successfully saved soil type '{soil_type}' to formation '{formation_name}' in {excel_path}")
        
        # Reload material names to reflect the changes
        self.material_names = self.load_formation_soil_types(formation_name)
        self.update_existing_soil_type_dropdowns()
        return True
        
      except Exception as e:
        print(f"Error creating/updating formation Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False   
    def show_soil_type_popup(self, formation_name):
      added_soil_types = []
    
      def close_popup(e):
        popup_dialog.open = False
        self.form_content.page.update()
    
      def add_soil_type(e):
        soil_type_value = soil_type_input.value.strip()
        if not soil_type_value:
            error_text.value = "Please enter a soil type"
            error_text.visible = True
            success_text.visible = False
            self.form_content.page.update()
            return
        
        if soil_type_value in added_soil_types:
            error_text.value = f"'{soil_type_value}' already added to the list"
            error_text.visible = True
            success_text.visible = False
            self.form_content.page.update()
            return
        
        success = self.create_formation_excel(formation_name, soil_type_value)
        if success:
            added_soil_types.append(soil_type_value)
            update_soil_types_list()
            success_text.value = f"'{soil_type_value}' added successfully!"
            success_text.visible = True
            error_text.visible = False
            soil_type_input.value = ""
            self.form_content.page.update()
        else:
            error_text.value = "Failed to save soil type. Please try again."
            error_text.visible = True
            success_text.visible = False
            self.form_content.page.update()
    
      def update_soil_types_list():
        soil_types_list.controls.clear()
        if added_soil_types:
            soil_types_list.controls.append(
                ft.Text("Added Soil Types:", weight=ft.FontWeight.BOLD, size=14)
            )
            for soil_type in added_soil_types:
                soil_types_list.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.Text(f"• {soil_type}", size=12),
                            # Removed the delete button - users cannot remove soil types
                        ], alignment=ft.MainAxisAlignment.START),
                        padding=ft.padding.only(left=10, right=5, top=2, bottom=2)
                    )
                )
        else:
            soil_types_list.controls.append(
                ft.Text("No soil types added yet", size=12, color=ft.colors.GREY_600)
            )
        self.form_content.page.update()
    
      def finish_adding(e):
        if added_soil_types:
            success_text.value = f"Successfully added {len(added_soil_types)} soil type(s) to {formation_name}"
            success_text.visible = True
            error_text.visible = False
            self.form_content.page.update()
            import threading
            import time
            def delayed_close():
                time.sleep(2)
                popup_dialog.open = False
                if hasattr(self.form_content, 'page'):
                    self.form_content.page.update()
            threading.Thread(target=delayed_close, daemon=True).start()
        else:
            error_text.value = "Please add at least one soil type before finishing"
            error_text.visible = True
            success_text.visible = False
            self.form_content.page.update()
    
      soil_type_input = ft.TextField(
        label="Enter Soil Type",
        hint_text="e.g., Clay, Sand, Silt, etc.",
        border=ft.InputBorder.OUTLINE,
        expand=True,
        autofocus=True,
        on_submit=add_soil_type
      )
    
      error_text = ft.Text(
        color=ft.colors.RED,
        visible=False,
        size=12
      )
    
      success_text = ft.Text(
        color=ft.colors.GREEN,
        visible=False,
        size=12
      )
    
      soil_types_list = ft.Column(
        controls=[ft.Text("No soil types added yet", size=12, color=ft.colors.GREY_600)],
        spacing=5,
        height=150,
        scroll=ft.ScrollMode.AUTO
      )
    
      popup_dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text(f"Add Soil Types to {formation_name}"),
        content=ft.Container(
            content=ft.Column([
                ft.Text(f"Selected Formation: {formation_name}", weight=ft.FontWeight.BOLD),
                ft.Divider(),
                ft.Row([
                    soil_type_input,
                    ft.ElevatedButton(
                        "Add",
                        on_click=add_soil_type,
                        style=ft.ButtonStyle(
                            color=ft.colors.WHITE,
                            bgcolor=ft.colors.GREEN_600
                        ),
                        height=40
                    )
                ], spacing=10),
                error_text,
                success_text,
                ft.Divider(),
                ft.Container(
                    content=soil_types_list,
                    border=ft.border.all(1, ft.colors.GREY_300),
                    border_radius=5,
                    padding=10
                )
            ], 
            tight=True,
            spacing=10
            ),
            width=500,
            height=400
        ),
        actions=[
            ft.TextButton("Cancel", on_click=close_popup),
            ft.ElevatedButton(
                "Finish", 
                on_click=finish_adding,
                style=ft.ButtonStyle(
                    color=ft.colors.WHITE,
                    bgcolor=ft.colors.BLUE_600
                )
            ),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )
    
      self.form_content.page.dialog = popup_dialog
      popup_dialog.open = True
      self.form_content.page.update()
    def on_formation_change(self, e):
        selected_formation = e.control.value
        if selected_formation:
            print(f"Formation selected: {selected_formation}")
            
            # Set the selected formation and update material names
            self.set_selected_formation(selected_formation)
            
            # Show the popup for adding soil types
            self.show_soil_type_popup(selected_formation)
    def import_from_list(self, data):
      self.data_table.rows.clear()
      for idx, item in enumerate(data):
        self.current_material_index = idx
        self.add_borehole_set(item)
    # Add this method to your BoreholeSection class

    def load_soil_db_sheets(self):
      """Load all sheets from Soil_DB.xlsx file"""
      try:
        if not self.soil_db_path.exists():
            print(f"ERROR: Soil DB file not found: {self.soil_db_path}")
            return {}
        
        import openpyxl
        workbook = openpyxl.load_workbook(self.soil_db_path)
        sheets_data = {}
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value))
            
            # Find SPT column index
            spt_col_index = None
            eref_col_index = None
            try:
                spt_col_index = headers.index('SPT') if 'SPT' in headers else None
                eref_col_index = headers.index('Eref') if 'Eref' in headers else None
            except ValueError:
                pass
            
            # Get data rows
            rows_data = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if any(cell.value is not None for cell in row):  # Skip empty rows
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            cell = row[i]
                            
                            # Check if this is the Eref column and contains a formula
                            if header == 'Eref' and hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                                # Convert formula: replace cell references with SPT placeholder
                                formula = cell.value
                                
                                # Replace cell references (like D5, D7) with SPT
                                import re

                                # Pattern to match column letter + row number (e.g., D5, D7)
                                if spt_col_index is not None:
                                    spt_col_letter = openpyxl.utils.get_column_letter(spt_col_index + 1)
                                    # Replace references to SPT column with "SPT"
                                    formula = re.sub(f'{spt_col_letter}\\d+', 'SPT', formula)
                                
                                row_dict[header] = formula
                                print(f"DEBUG: Converted formula in row {row_idx}: {cell.value} -> {formula}")
                            else:
                                row_dict[header] = cell.value if cell.value is not None else ""
                    
                    if row_dict:  # Only add non-empty rows
                        rows_data.append(row_dict)
            
            if rows_data:  # Only add sheets that have data
                sheets_data[sheet_name] = {
                    'headers': headers,
                    'rows': rows_data
                }
        
        print(f"DEBUG: Loaded {len(sheets_data)} sheets from Soil DB")
        return sheets_data
        
      except Exception as e:
        print(f"ERROR: Loading Soil DB sheets: {e}")
        import traceback
        traceback.print_exc()
        return {}

    def create_soil_db_popup(self):
      sheets_data = self.load_soil_db_sheets()
    
      if not sheets_data:
        # Show error dialog if no sheets found
        error_dialog = ft.AlertDialog(
            title=ft.Text("Error"),
            content=ft.Text("No data found in Soil_DB.xlsx file or file not found."),
            actions=[
                ft.TextButton("OK", on_click=lambda e: self.close_dialog(error_dialog))
            ]
        )
        return error_dialog
    
    # Create tabs for each sheet
      sheet_tabs = []
      for sheet_name, sheet_info in sheets_data.items():
        # Filter out SPT column from headers if it exists
        filtered_headers = [header for header in sheet_info['headers'] if header.upper() != 'SPT']
        
        # Create data table for this sheet
        columns = []
        for header in filtered_headers:
            columns.append(ft.DataColumn(ft.Text(header, size=12, weight=ft.FontWeight.BOLD)))
        
        rows = []
        for row_data in sheet_info['rows']:
            cells = []
            for header in filtered_headers:
                value = str(row_data.get(header, ""))
                cells.append(ft.DataCell(ft.Text(value, size=11)))
            rows.append(ft.DataRow(cells=cells))
        
        data_table = ft.DataTable(
            columns=columns,
            rows=rows,
            border=ft.border.all(1, ft.colors.GREY_300),
            border_radius=5,
            heading_row_height=40,
            data_row_min_height=35,
            data_row_max_height=35,
            column_spacing=10,
            horizontal_lines=ft.border.BorderSide(1, ft.colors.GREY_200),
            vertical_lines=ft.border.BorderSide(1, ft.colors.GREY_200),
        )
        
        # Create scrollable container for the table - FIXED VERSION
        table_container = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Text(f"Sheet: {sheet_name}", size=16, weight=ft.FontWeight.BOLD),
                    ft.ElevatedButton(
                        text="Select This Sheet",
                        icon=ft.icons.CHECK,
                        on_click=lambda e, sheet=sheet_name, data=sheet_info: self.load_selected_sheet_data(e, sheet, data),
                        style=ft.ButtonStyle(
                            color=ft.colors.WHITE,
                            bgcolor=ft.colors.GREEN_600,
                        )
                    )
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ft.Container(height=10),  # Small spacer
                # FIXED: Proper scrolling container
                ft.Container(
                    content=ft.Column(
                        controls=[
                            ft.Row(
                                controls=[data_table],
                                scroll=ft.ScrollMode.ALWAYS  # Enable horizontal scroll
                            )
                        ],
                        scroll=ft.ScrollMode.ALWAYS,  # Enable vertical scroll
                        expand=True
                    ),
                    height=350,  # Fixed height for scrolling
                    border=ft.border.all(1, ft.colors.GREY_300),
                    border_radius=5,
                    padding=10,
                    expand=True
                )
            ], 
            spacing=0,
            expand=True
            ),
            padding=10,
            expand=True
        )
        
        sheet_tabs.append(
            ft.Tab(
                text=sheet_name,
                content=table_container
            )
        )
    
    # Create the popup dialog with better sizing
      popup_dialog = ft.AlertDialog(
        title=ft.Text("Select Soil Database Sheet", size=18, weight=ft.FontWeight.BOLD),
        content=ft.Container(
            content=ft.Tabs(
                tabs=sheet_tabs,
                selected_index=0,
                expand=True
            ),
            width=1000,  # Increased width
            height=550,  # Increased height
        ),
        actions=[
            ft.TextButton("Cancel", on_click=lambda e: self.close_dialog(popup_dialog))
        ]
    )
    
      return popup_dialog
    def get_current_borehole_data(self):
      """Extract current data from the borehole table"""
      current_data = []
    
      if not hasattr(self, 'data_table') or not self.data_table or not self.data_table.rows:
        return current_data
    
    # Get field names for mapping
      fields = self.get_fields()
      field_labels = [field.label for field in fields]
    
      for row in self.data_table.rows:
        if row.cells:
            row_data = {}
            for i, cell in enumerate(row.cells):
                if i < len(field_labels):
                    field_name = field_labels[i]
                    
                    # Extract value based on control type
                    if isinstance(cell.content, ft.Dropdown):
                        value = cell.content.value or ""
                    elif isinstance(cell.content, ft.TextField):
                        value = cell.content.value or ""
                    else:
                        value = ""
                    
                    row_data[field_name] = value
            
            # Only add rows that have some data
            if any(value for value in row_data.values()):
                current_data.append(row_data)
    
      return current_data
    def load_selected_sheet_data(self, e, sheet_name, sheet_data):
      try:
        print(f"DEBUG: Loading data from sheet: {sheet_name}")
        self.close_current_dialog()
        
        # Check if there's existing AGS geological data
        has_existing_ags_data = False
        if hasattr(self, 'data_table') and self.data_table and self.data_table.rows:
            # Check if any row has both Top Depth and Bottom Depth filled
            for row in self.data_table.rows:
                if row.cells and len(row.cells) >= 5:  # At least enough cells for depth fields
                    top_depth_cell = row.cells[3]  # Top Depth field
                    bottom_depth_cell = row.cells[4]  # Bottom Depth field
                    
                    top_value = ""
                    bottom_value = ""
                    
                    if isinstance(top_depth_cell.content, ft.TextField):
                        top_value = top_depth_cell.content.value or ""
                    if isinstance(bottom_depth_cell.content, ft.TextField):
                        bottom_value = bottom_depth_cell.content.value or ""
                    
                    if top_value and bottom_value:
                        has_existing_ags_data = True
                        print(f"DEBUG: Found existing AGS data in borehole table")
                        break
        
        # Column mapping - SPT removed from soil database mapping
        column_mapping = {
            'MaterialName': 'Soil Type',
            'DrainageType': 'Drain Type',
            'gammaUnsat': 'Gamma Unsat',
            'gammaSat': 'Gamma Sat',
            'Eref': 'E ref',
            'nu': 'Nu',
            'cref': 'C \'',
            'phi': 'Phi \'',
            'kx': 'Kx',
            'ky': 'Ky',
            'Rinter': 'R inter',
            'K0Primary': 'K0 Primary'
        }
        
        if has_existing_ags_data:
            print(f"DEBUG: AGS data exists - updating matching rows only")
            self.update_existing_rows_with_soil_db_data(sheet_data['rows'], column_mapping)
        else:
            print(f"DEBUG: No AGS data found - loading all sheet data as new rows")
            # Clear existing data and load all sheet data
            if hasattr(self, 'data_table') and self.data_table:
                self.data_table.rows.clear()
            
            self.current_material_index = 0
            self.visible_sets = 0
            
            rows_added = 0
            for row_data in sheet_data['rows']:
                if not row_data:
                    continue
                
                initial_data = {}
                for soil_db_col, borehole_field in column_mapping.items():
                    if soil_db_col in row_data:
                        value = row_data[soil_db_col]
                        if value is not None:
                            initial_data[borehole_field] = str(value)
                        else:
                            initial_data[borehole_field] = ""
                
                # Handle depth fields
                if 'TopDepth' in row_data and row_data['TopDepth'] is not None:
                    initial_data['Top Depth'] = str(row_data['TopDepth'])
                if 'BottomDepth' in row_data and row_data['BottomDepth'] is not None:
                    initial_data['Bottom Depth'] = str(row_data['BottomDepth'])
                
                # Check for alternative depth column names
                depth_column_alternatives = {
                    'Top Depth': ['TopDepth', 'Top_Depth', 'top_depth', 'Depth_From', 'DepthFrom'],
                    'Bottom Depth': ['BottomDepth', 'Bottom_Depth', 'bottom_depth', 'Depth_To', 'DepthTo']
                }
                
                for field_name, alt_columns in depth_column_alternatives.items():
                    if field_name not in initial_data:
                        for alt_col in alt_columns:
                            if alt_col in row_data and row_data[alt_col] is not None:
                                initial_data[field_name] = str(row_data[alt_col])
                                break
                
                # Note: SPT data is not included from soil database
                # SPT will remain empty and be filled separately in the borehole section
                
                soil_type = initial_data.get('Soil Type', f'Material_{rows_added + 1}')
                new_row = self.create_borehole_row(soil_type, initial_data)
                self.data_table.rows.append(new_row)
                self.visible_sets += 1
                self.current_material_index += 1
                rows_added += 1
                print(f"DEBUG: Added row {rows_added} with soil type: {soil_type}")
            
            print(f"DEBUG: Successfully loaded {rows_added} rows from sheet '{sheet_name}'")
        
        self.data_table.update()
        
      except Exception as e:
        print(f"ERROR: Loading sheet data: {e}")
        import traceback
        traceback.print_exc()
        error_dialog = ft.AlertDialog(
            title=ft.Text("Error"),
            content=ft.Text(f"Error loading data from sheet: {str(e)}"),
            actions=[
                ft.TextButton("OK", on_click=lambda e: self.close_dialog(error_dialog))
            ]
        )
        if hasattr(self, 'page') and self.page:
            self.page.dialog = error_dialog
            error_dialog.open = True
            self.page.update()
        elif hasattr(self, 'form_content') and hasattr(self.form_content, 'page'):
            self.form_content.page.dialog = error_dialog
            error_dialog.open = True
            self.form_content.page.update()

    def evaluate_excel_formula(self, formula_str: str, spt_value: float) -> float:
      try:
        # Remove leading '=' if present
        formula = formula_str.strip()
        if formula.startswith('='):
            formula = formula[1:]
        
        # Replace SPT with actual value
        formula = formula.replace('SPT', str(spt_value))
        
        # Convert Excel IF function to Python format
        # Excel: IF(condition, value_if_true, value_if_false)
        # Python: value_if_true if condition else value_if_false
        
        import re

        # Handle nested IF functions by processing from innermost to outermost
        while 'IF(' in formula.upper():
            # Find the innermost IF function
            if_pattern = re.compile(r'IF\s*\(([^()]+)\)', re.IGNORECASE)
            match = if_pattern.search(formula)
            
            if not match:
                # Try to find IF with nested parentheses
                if_pattern = re.compile(r'IF\s*\(([^()]+(?:\([^()]*\)[^()]*)*)\)', re.IGNORECASE)
                match = if_pattern.search(formula)
            
            if match:
                if_content = match.group(1)
                parts = self._split_if_arguments(if_content)
                
                if len(parts) == 3:
                    condition, true_val, false_val = parts
                    # Convert to Python ternary operator
                    python_expr = f"({true_val} if ({condition}) else {false_val})"
                    formula = formula[:match.start()] + python_expr + formula[match.end():]
                else:
                    print(f"WARNING: Could not parse IF function: {match.group(0)}")
                    break
            else:
                break
        
        # Replace Excel operators with Python equivalents
        formula = formula.replace('^', '**')  # Exponentiation
        
        # Evaluate the formula safely
        # Only allow mathematical operations
        allowed_names = {
            '__builtins__': {},
            'abs': abs,
            'min': min,
            'max': max,
            'round': round,
            'pow': pow
        }
        
        result = eval(formula, allowed_names)
        print(f"DEBUG: Evaluated formula '{formula_str}' with SPT={spt_value} -> {result}")
        return float(result)
        
      except Exception as e:
        print(f"ERROR: Evaluating formula '{formula_str}' with SPT={spt_value}: {e}")
        import traceback
        traceback.print_exc()
        return 0.0


    def _split_if_arguments(self, if_content: str) -> list:
      """
      Split IF function arguments, respecting nested parentheses.
      Returns [condition, value_if_true, value_if_false]
      """
      parts = []
      current = ""
      paren_depth = 0
      
      for char in if_content:
        if char == ',' and paren_depth == 0:
            parts.append(current.strip())
            current = ""
        else:
            if char == '(':
                paren_depth += 1
            elif char == ')':
                paren_depth -= 1
            current += char
    
      if current:
        parts.append(current.strip())
    
      return parts


    def update_existing_rows_with_soil_db_data(self, matching_rows, column_mapping):
      """Enhanced version with proper formula evaluation"""
      try:
        if not hasattr(self, 'data_table') or not self.data_table or not self.data_table.rows:
            return
        
        # Create a lookup dictionary for soil DB data
        soil_db_lookup = {}
        for row_data in matching_rows:
            soil_type = str(row_data.get('MaterialName', '')).strip().lower()
            if soil_type:
                soil_db_lookup[soil_type] = row_data
        
        print(f"DEBUG: Created soil DB lookup for types: {list(soil_db_lookup.keys())}")
        
        # Get field labels for reference
        fields = self.get_fields()
        field_labels = [field.label for field in fields]
        
        updated_rows = 0
        
        # Update existing rows where soil types match
        for row in self.data_table.rows:
            if not row.cells:
                continue
            
            # Extract soil type from the second cell (index 1)
            current_soil_type = ""
            
            if len(row.cells) > 1:
                soil_type_cell = row.cells[1]
                
                if isinstance(soil_type_cell.content, ft.Dropdown):
                    current_soil_type = soil_type_cell.content.value or ""
                elif isinstance(soil_type_cell.content, ft.TextField):
                    current_soil_type = soil_type_cell.content.value or ""
                
                current_soil_type = current_soil_type.strip().lower()
                print(f"DEBUG: Current soil type in row: '{current_soil_type}'")
                
                # If this soil type exists in the soil DB, update the row
                if current_soil_type and current_soil_type in soil_db_lookup:
                    soil_db_data = soil_db_lookup[current_soil_type]
                    print(f"DEBUG: Updating row with soil type: {current_soil_type}")
                    
                    # Get SPT value from current row for formula calculation
                    current_spt_value = None
                    spt_field_index = None
                    for idx, field_name in enumerate(field_labels):
                        if field_name == 'SPT':
                            spt_field_index = idx + 1  # +1 for actions cell
                            if spt_field_index < len(row.cells):
                                spt_cell = row.cells[spt_field_index]
                                if isinstance(spt_cell.content, ft.TextField):
                                    try:
                                        current_spt_value = float(spt_cell.content.value) if spt_cell.content.value else None
                                    except (ValueError, TypeError):
                                        current_spt_value = None
                            break
                    
                    # Update each field except Top Depth, Bottom Depth, and SPT
                    for field_index, field_name in enumerate(field_labels):
                        # Skip soil type, depth fields, and SPT to preserve existing data
                        if field_name in ['Soil Type', 'Top Depth', 'Bottom Depth', 'SPT']:
                            continue
                        
                        # Find corresponding soil DB column
                        soil_db_column = None
                        for db_col, borehole_col in column_mapping.items():
                            if borehole_col == field_name:
                                soil_db_column = db_col
                                break
                        
                        # Update the field if corresponding data exists
                        if soil_db_column and soil_db_column in soil_db_data:
                            value = soil_db_data[soil_db_column]
                            
                            # Check if it's a formula for Eref column
                            if field_name == 'E ref' and isinstance(value, str) and value.startswith('='):
                                # Calculate formula using current row's SPT value
                                if current_spt_value is not None:
                                    calculated_value = self.evaluate_excel_formula(value, current_spt_value)
                                    new_value = str(round(calculated_value, 2))
                                    print(f"DEBUG: Calculated Eref = {new_value} from formula with SPT={current_spt_value}")
                                else:
                                    # No SPT value available, show formula
                                    new_value = value
                                    print(f"DEBUG: No SPT value for formula calculation, using: {new_value}")
                            else:
                                new_value = str(value) if value is not None else ""
                            
                            # Calculate correct cell index (+1 to account for actions cell)
                            cell_index = field_index + 1
                            
                            if cell_index < len(row.cells):
                                cell = row.cells[cell_index]
                                
                                if isinstance(cell.content, ft.Dropdown):
                                    if hasattr(cell.content, 'options'):
                                        option_values = [opt.key if hasattr(opt, 'key') else opt.text for opt in cell.content.options]
                                        if new_value in option_values:
                                            cell.content.value = new_value
                                            print(f"DEBUG: Updated dropdown {field_name} to: {new_value}")
                                elif isinstance(cell.content, ft.TextField):
                                    cell.content.value = new_value
                                    print(f"DEBUG: Updated text field {field_name} to: {new_value}")
                    
                    updated_rows += 1
                else:
                    print(f"DEBUG: No match for '{current_soil_type}' in soil DB")
        
        print(f"DEBUG: Successfully updated {updated_rows} existing rows with soil DB data")
        
        # Show success message
        if updated_rows > 0:
            success_dialog = ft.AlertDialog(
                title=ft.Text("Success"),
                content=ft.Text(
                    f"Successfully updated {updated_rows} matching rows with soil database properties.\n"
                    f"AGS geological data (depths and soil types) and SPT values have been preserved.\n"
                    f"Formulas (including IF functions) have been calculated using actual SPT values."
                ),
                actions=[
                    ft.TextButton("OK", on_click=lambda e: self.close_dialog(success_dialog))
                ]
            )
            if hasattr(self, 'form_content') and hasattr(self.form_content, 'page'):
                self.form_content.page.dialog = success_dialog
                success_dialog.open = True
                self.form_content.page.update()
        else:
            no_match_dialog = ft.AlertDialog(
                title=ft.Text("No Matches Found"),
                content=ft.Text(
                    "No matching soil types found between the existing borehole data and the selected soil database sheet.\n"
                    "Make sure the soil type names match exactly."
                ),
                actions=[
                    ft.TextButton("OK", on_click=lambda e: self.close_dialog(no_match_dialog))
                ]
            )
            if hasattr(self, 'form_content') and hasattr(self.form_content, 'page'):
                self.form_content.page.dialog = no_match_dialog
                no_match_dialog.open = True
                self.form_content.page.update()
        
      except Exception as e:
        print(f"ERROR: Updating existing rows with soil DB data: {e}")
        import traceback
        traceback.print_exc()
    def load_all_sheet_data(self, e, sheet_name, sheet_data, current_dialog):
      """Load all data from sheet (called from no-match dialog)"""
      try:
        # Close the current dialog
        self.close_dialog(current_dialog)
        
        # Call the main loading method
        self.load_selected_sheet_data(None, sheet_name, sheet_data)
        
      except Exception as e:
        print(f"ERROR: Loading all sheet data: {e}")

    def load_all_sheet_data_direct(self, sheet_name, sheet_data):
      try:
        if hasattr(self, 'data_table') and self.data_table:
            self.data_table.rows.clear()
        
        self.current_material_index = 0
        self.visible_sets = 0
        
        column_mapping = {
            'MaterialName': 'Soil Type',
            'DrainageType': 'Drain Type',
            'SPT': 'SPT',
            'gammaUnsat': 'Gamma Unsat',
            'gammaSat': 'Gamma Sat',
            'Eref': 'E ref',
            'nu': 'Nu',
            'cref': 'C \'',
            'phi': 'Phi \'',
            'kx': 'Kx',
            'ky': 'Ky',
            'Rinter': 'R inter',
            'K0Primary': 'K0 Primary'
        }
        
        rows_added = 0
        for row_data in sheet_data['rows']:
            if not row_data:
                continue
            
            initial_data = {}
            
            # Only map columns that exist in the sheet data
            for soil_db_col, borehole_field in column_mapping.items():
                if soil_db_col in row_data:
                    value = row_data[soil_db_col]
                    if value is not None:
                        initial_data[borehole_field] = str(value)
                    else:
                        initial_data[borehole_field] = ""
            
            # Check if Top Depth and Bottom Depth exist in the sheet data
            # If they exist, use them. If not, leave them empty (don't auto-generate)
            if 'TopDepth' in row_data and row_data['TopDepth'] is not None:
                initial_data['Top Depth'] = str(row_data['TopDepth'])
            if 'BottomDepth' in row_data and row_data['BottomDepth'] is not None:
                initial_data['Bottom Depth'] = str(row_data['BottomDepth'])
            
            # Check for alternative column names
            depth_column_alternatives = {
                'Top Depth': ['TopDepth', 'Top_Depth', 'top_depth', 'Depth_From', 'DepthFrom'],
                'Bottom Depth': ['BottomDepth', 'Bottom_Depth', 'bottom_depth', 'Depth_To', 'DepthTo']
            }
            
            for field_name, alt_columns in depth_column_alternatives.items():
                if field_name not in initial_data:  # Only if not already set
                    for alt_col in alt_columns:
                        if alt_col in row_data and row_data[alt_col] is not None:
                            initial_data[field_name] = str(row_data[alt_col])
                            break
            
            soil_type = initial_data.get('Soil Type', f'Material_{rows_added + 1}')
            new_row = self.create_borehole_row(soil_type, initial_data)
            self.data_table.rows.append(new_row)
            
            self.visible_sets += 1
            self.current_material_index += 1
            rows_added += 1
        
        self.data_table.update()
        print(f"DEBUG: Loaded {rows_added} rows directly")
        
      except Exception as e:
        print(f"ERROR: Loading all sheet data direct: {e}")
        import traceback
        traceback.print_exc()
        
        # Show success dialog
        success_dialog = ft.AlertDialog(
            title=ft.Text("Success"),
            content=ft.Text(
                f"Successfully loaded {rows_added} materials from sheet '{sheet_name}' "
                f"into the borehole table. Depth fields left empty for manual entry."
            ),
            actions=[
                ft.TextButton("OK", on_click=lambda e: self.close_dialog(success_dialog))
            ]
        )
        
        if hasattr(self, 'page') and self.page:
            self.page.dialog = success_dialog
            success_dialog.open = True
            self.page.update()
        elif hasattr(self, 'form_content') and hasattr(self.form_content, 'page'):
            self.form_content.page.dialog = success_dialog
            success_dialog.open = True
            self.form_content.page.update()
            
      except Exception as e:
        print(f"ERROR: Loading sheet data: {e}")
        import traceback
        traceback.print_exc()
        
        error_dialog = ft.AlertDialog(
            title=ft.Text("Error"),
            content=ft.Text(f"Error loading data from sheet: {str(e)}"),
            actions=[
                ft.TextButton("OK", on_click=lambda e: self.close_dialog(error_dialog))
            ]
        )
        
        if hasattr(self, 'page') and self.page:
            self.page.dialog = error_dialog
            error_dialog.open = True
            self.page.update()
        elif hasattr(self, 'form_content') and hasattr(self.form_content, 'page'):
            self.form_content.page.dialog = error_dialog
            error_dialog.open = True
            self.form_content.page.update() 
    def open_soil_db_popup(self, e):
      """Open the Soil DB selection popup"""
      popup = self.create_soil_db_popup()
    
    # Store reference to current dialog
      self.current_dialog = popup
    
    # Show the dialog
      if hasattr(self, 'page') and self.page:
        self.page.dialog = popup
        popup.open = True
        self.page.update()
      elif hasattr(self, 'form_content') and hasattr(self.form_content, 'page'):
        self.form_content.page.dialog = popup
        popup.open = True
        self.form_content.page.update()

    def close_dialog(self, dialog):
      """Close a specific dialog"""
      try:
        dialog.open = False
        if hasattr(self, 'page') and self.page:
            self.page.update()
        elif hasattr(self, 'form_content') and hasattr(self.form_content, 'page'):
            self.form_content.page.update()
      except Exception as e:
        print(f"DEBUG: Error closing specific dialog: {e}")

    def close_current_dialog(self):
      """Close the currently open dialog"""
      try:
        if hasattr(self, 'page') and self.page and hasattr(self.page, 'dialog') and self.page.dialog:
            self.page.dialog.open = False
            self.page.update()
        elif hasattr(self, 'form_content') and hasattr(self.form_content, 'page') and hasattr(self.form_content.page, 'dialog') and self.form_content.page.dialog:
            self.form_content.page.dialog.open = False
            self.form_content.page.update()
      except Exception as e:
        print(f"DEBUG: Error closing dialog: {e}")

    def on_formation_change(self, e):
      """Handle formation dropdown change"""
      selected_formation = e.control.value
      print(f"DEBUG: Formation changed to: {selected_formation}")
      self.set_selected_formation(selected_formation)

# Add this method to store page reference
    def set_page_reference(self, page):
      """Store reference to the page for dialog management"""
      self.page = page